from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from database import (
    AdminDB,
    AgentAssignmentDB,
    AgentDeletionDB,
    CartDB,
    CouponDB,
    GiftThresholdDB,
    LotteryConfigDB,
    LotteryDB,
    OrderDB,
    RewardDB,
    UserProfileDB,
)
from ..dependencies import build_staff_scope
from ..utils import convert_sqlite_timestamp_to_unix, format_device_time_ms, format_export_range_label

MANAGE_ORDER_STATUS_ALIASES: Dict[str, str] = {
    "未付款": "未付款",
    "unpaid": "未付款",
    "pending": "待确认",
    "待确认": "待确认",
    "pending_confirm": "待确认",
    "processing": "待确认",
    "confirmed": "待配送",
    "待配送": "待配送",
    "awaiting_delivery": "待配送",
    "paid": "待配送",
    "shipped": "配送中",
    "配送中": "配送中",
    "delivering": "配送中",
    "delivered": "已完成",
    "已完成": "已完成",
    "completed": "已完成",
    "cancelled": "已取消",
    "已取消": "已取消",
}


def resolve_staff_order_scope(
    staff: Dict[str, Any],
    scope: Dict[str, Any],
    agent_param: Optional[str],
) -> Tuple[Optional[str], Optional[List[str]], Optional[List[str]], Optional[List[str]], Optional[List[str]], str]:
    selected_agent_id = scope.get("agent_id")
    selected_address_ids = scope.get("address_ids")
    selected_building_ids = scope.get("building_ids")
    exclude_address_ids: Optional[List[str]] = None
    exclude_building_ids: Optional[List[str]] = None

    if staff.get("type") == "agent":
        return (
            staff.get("agent_id"),
            selected_address_ids,
            selected_building_ids,
            None,
            None,
            "self",
        )

    filter_value = (agent_param or "").strip() or "self"
    lower = filter_value.lower()

    if lower == "all":
        return None, None, None, None, None, "all"

    if lower == "self":
        assignments = AgentAssignmentDB.list_agents_with_buildings(include_disabled=True)
        address_set: Set[str] = set()
        building_set: Set[str] = set()
        for entry in assignments:
            for record in entry.get("buildings") or []:
                addr_id = record.get("address_id")
                bld_id = record.get("building_id")
                if addr_id:
                    address_set.add(addr_id)
                if bld_id:
                    building_set.add(bld_id)
        return (
            None,
            None,
            None,
            list(address_set) if address_set else None,
            list(building_set) if building_set else None,
            "self",
        )

    target = AdminDB.get_admin_by_agent_id(filter_value, include_disabled=True, include_deleted=True)
    if not target or (target.get("role") or "").lower() != "agent":
        raise HTTPException(status_code=400, detail="指定的代理不存在")

    assignments = AgentAssignmentDB.get_buildings_for_agent(filter_value)
    address_ids = list({record.get("address_id") for record in assignments if record.get("address_id")}) or None
    building_ids = [record.get("building_id") for record in assignments if record.get("building_id")]

    if target.get("deleted_at") and not assignments:
        deletion_records = AgentDeletionDB.list_active_records()
        for record in deletion_records:
            if record.get("agent_id") == filter_value:
                address_ids = record.get("address_ids") or None
                building_ids = record.get("building_ids") or None
                break

    return filter_value, address_ids, building_ids, None, None, filter_value


def compute_unified_order_status(order: Dict[str, Any]) -> str:
    ps = order.get("payment_status") if isinstance(order, dict) else None
    st = order.get("status") if isinstance(order, dict) else None
    if not ps and not st:
        return "未付款"
    if st == "cancelled":
        return "已取消"
    if ps == "processing":
        return "待确认"
    if ps != "succeeded":
        return "未付款"
    if st == "shipped":
        return "配送中"
    if st == "delivered":
        return "已完成"
    return "待配送"


def normalize_manage_order_status(target_status: Any) -> Optional[str]:
    if target_status is None:
        return None
    text = str(target_status).strip()
    if not text:
        return None
    if text in MANAGE_ORDER_STATUS_ALIASES:
        return MANAGE_ORDER_STATUS_ALIASES[text]
    return MANAGE_ORDER_STATUS_ALIASES.get(text.lower())


def _refresh_order_snapshot(order_id: str, fallback: Dict[str, Any]) -> Dict[str, Any]:
    return OrderDB.get_order_by_id(order_id) or fallback


def _apply_payment_status_change(
    order_id: str,
    order: Dict[str, Any],
    new_status: str,
) -> Tuple[bool, List[str], Dict[str, Any]]:
    current_status = str(order.get("payment_status") or "pending").strip() or "pending"
    if current_status == new_status:
        return True, [], order

    if new_status == "succeeded":
        ok, missing_items = OrderDB.complete_payment_and_update_stock(order_id)
        if not ok:
            return False, missing_items, order

        order_owner_id = LotteryConfigDB.normalize_owner(order.get("agent_id"))
        try:
            CartDB.update_cart(order["student_id"], {})
        except Exception:
            pass
        try:
            if isinstance(order.get("shipping_info"), dict):
                UserProfileDB.upsert_shipping(order["student_id"], order["shipping_info"])
        except Exception:
            pass
        try:
            draw = LotteryDB.get_draw_by_order(order_id)
            if draw and draw.get("prize_name") != "谢谢参与":
                RewardDB.add_reward_from_order(
                    user_identifier=order["student_id"],
                    prize_name=draw.get("prize_name"),
                    prize_product_id=draw.get("prize_product_id"),
                    quantity=int(draw.get("prize_quantity") or 1),
                    source_order_id=order_id,
                    owner_id=order_owner_id,
                    prize_group_id=draw.get("prize_group_id"),
                    prize_product_name=draw.get("prize_product_name"),
                    prize_variant_id=draw.get("prize_variant_id"),
                    prize_variant_name=draw.get("prize_variant_name"),
                    prize_unit_price=draw.get("prize_unit_price"),
                )
        except Exception:
            pass
        try:
            items = order.get("items") or []
            items_subtotal = 0.0
            for item in items:
                if isinstance(item, dict) and not (item.get("is_lottery") or item.get("is_auto_gift")):
                    try:
                        items_subtotal += float(item.get("subtotal", 0) or 0)
                    except Exception:
                        pass
            applicable_thresholds = GiftThresholdDB.get_applicable_thresholds(items_subtotal, order_owner_id)
            for threshold in applicable_thresholds:
                gift_coupon = threshold.get("gift_coupon", 0) == 1
                coupon_amount = threshold.get("coupon_amount", 0)
                applicable_times = threshold.get("applicable_times", 0)
                if gift_coupon and coupon_amount > 0 and applicable_times > 0:
                    for _ in range(applicable_times):
                        CouponDB.issue_coupons(
                            user_identifier=order["student_id"],
                            amount=coupon_amount,
                            quantity=1,
                            expires_at=None,
                            owner_id=order_owner_id,
                        )
        except Exception:
            pass
        try:
            coupon_id = order.get("coupon_id")
            discount_amount = float(order.get("discount_amount") or 0)
            if coupon_id and discount_amount > 0:
                CouponDB.delete_coupon(coupon_id)
        except Exception:
            pass
        return True, [], _refresh_order_snapshot(order_id, {**order, "payment_status": new_status})

    ok, missing_items = OrderDB.update_payment_status_with_inventory(order_id, new_status)
    if not ok:
        return False, missing_items, order

    if new_status in ["pending", "failed"]:
        try:
            coupon_id = order.get("coupon_id")
            discount_amount = float(order.get("discount_amount") or 0)
            if coupon_id and discount_amount > 0:
                CouponDB.unlock_for_order(coupon_id, order_id)
        except Exception:
            pass

    return True, [], _refresh_order_snapshot(order_id, {**order, "payment_status": new_status})


def _apply_order_status_change(
    order_id: str,
    order: Dict[str, Any],
    new_status: str,
) -> Tuple[bool, List[str], Dict[str, Any]]:
    current_status = str(order.get("status") or "").strip()
    if current_status == new_status:
        return True, [], order
    ok, missing_items = OrderDB.update_order_status_with_inventory(order_id, new_status)
    if not ok:
        return False, missing_items, order
    return True, [], _refresh_order_snapshot(order_id, {**order, "status": new_status})


def apply_manage_order_status_transition(order_id: str, target_status: Any) -> Tuple[bool, List[str], Dict[str, Any]]:
    order = OrderDB.get_order_by_id(order_id)
    if not order:
        return False, ["订单不存在"], {}

    normalized_target = normalize_manage_order_status(target_status)
    if not normalized_target:
        return False, ["无效的订单状态"], {}

    old_unified_status = compute_unified_order_status(order)

    if normalized_target == "未付款":
        ok, missing_items, order = _apply_payment_status_change(order_id, order, "pending")
        if not ok:
            return False, missing_items, {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "pending")
        if not ok:
            return False, missing_items, {}
    elif normalized_target == "待确认":
        ok, missing_items, order = _apply_payment_status_change(order_id, order, "processing")
        if not ok:
            return False, missing_items, {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "pending")
        if not ok:
            return False, missing_items, {}
    elif normalized_target == "待配送":
        current_payment_status = str(order.get("payment_status") or "pending").strip() or "pending"
        if current_payment_status != "succeeded":
            if current_payment_status != "processing":
                ok, missing_items, order = _apply_payment_status_change(order_id, order, "processing")
                if not ok:
                    return False, missing_items, {}
            ok, missing_items, order = _apply_payment_status_change(order_id, order, "succeeded")
            if not ok:
                return False, missing_items, {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "confirmed")
        if not ok:
            return False, missing_items, {}
    elif normalized_target == "配送中":
        current_payment_status = str(order.get("payment_status") or "pending").strip() or "pending"
        if current_payment_status != "succeeded":
            return False, ["请先将订单状态设置为待配送"], {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "shipped")
        if not ok:
            return False, missing_items, {}
    elif normalized_target == "已完成":
        current_payment_status = str(order.get("payment_status") or "pending").strip() or "pending"
        if current_payment_status != "succeeded":
            return False, ["请先将订单状态设置为待配送"], {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "delivered")
        if not ok:
            return False, missing_items, {}
    elif normalized_target == "已取消":
        current_payment_status = str(order.get("payment_status") or "pending").strip() or "pending"
        if current_payment_status not in {"pending", "failed"}:
            ok, missing_items, order = _apply_payment_status_change(order_id, order, "pending")
            if not ok:
                return False, missing_items, {}
        ok, missing_items, order = _apply_order_status_change(order_id, order, "cancelled")
        if not ok:
            return False, missing_items, {}

    new_unified_status = compute_unified_order_status(order)
    return True, [], {
        "order_id": order_id,
        "old_unified_status": old_unified_status,
        "new_unified_status": new_unified_status,
        "status": order.get("status"),
        "payment_status": order.get("payment_status"),
    }


def resolve_order_timestamp_ms(order: Dict[str, Any]) -> Optional[float]:
    if not isinstance(order, dict):
        return None
    raw_ts = order.get("created_at_timestamp")
    if isinstance(raw_ts, (int, float)):
        return float(raw_ts) * 1000
    created_at_str = order.get("created_at")
    if created_at_str:
        try:
            return float(convert_sqlite_timestamp_to_unix(created_at_str, order.get("id"))) * 1000
        except Exception:
            return None
    return None


def build_agent_name_map() -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    agents = AdminDB.list_admins(role="agent", include_disabled=True, include_deleted=True)
    for agent in agents:
        agent_id = agent.get("agent_id")
        if not agent_id:
            continue
        mapping[agent_id] = agent.get("name") or agent.get("id") or agent_id
    return mapping


def resolve_scope_label(selected_filter: Optional[str], staff: Dict[str, Any], agent_name_map: Dict[str, str]) -> str:
    if staff.get("type") == "agent":
        return staff.get("name") or staff.get("id") or "我的订单"
    lower = (selected_filter or "self").lower()
    if lower == "all":
        return "全部订单"
    if lower == "self":
        return "管理员订单"
    return f"{agent_name_map.get(selected_filter, selected_filter)} 的订单"


def resolve_order_owner_label(order: Dict[str, Any], agent_name_map: Dict[str, str], staff: Dict[str, Any], is_admin_role: bool) -> str:
    agent_id = order.get("agent_id") if isinstance(order, dict) else None
    if is_admin_role:
        if agent_id:
            return agent_name_map.get(agent_id) or agent_id
        return staff.get("name") or staff.get("id") or "管理员"
    if agent_id:
        return agent_name_map.get(agent_id) or agent_id
    return staff.get("name") or staff.get("id") or "我的订单"


def build_export_row(
    order: Dict[str, Any],
    agent_name_map: Dict[str, str],
    staff: Dict[str, Any],
    is_admin_role: bool,
    tz_offset_minutes: Optional[int],
) -> List[str]:
    shipping = order.get("shipping_info") if isinstance(order, dict) and isinstance(order.get("shipping_info"), dict) else {}
    owner_label = resolve_order_owner_label(order, agent_name_map, staff, is_admin_role)
    address_parts = [shipping.get("dormitory"), shipping.get("building")]
    base_address = " ".join([part for part in address_parts if part]) or shipping.get("full_address") or ""
    detail_segments = [
        shipping.get("room"),
        shipping.get("address_detail"),
        shipping.get("detail"),
        shipping.get("extra"),
    ]
    detail_address = " ".join([seg for seg in detail_segments if seg]) or ""

    total_value = order.get("total_amount")
    total_text = f"{float(total_value):.2f}" if isinstance(total_value, (int, float)) else str(total_value or "")

    items = order.get("items") if isinstance(order, dict) and isinstance(order.get("items"), list) else []
    item_summary_parts: List[str] = []
    for item in items:
        if not item:
            continue
        markers: List[str] = []
        try:
            if item.get("is_auto_gift"):
                markers.append("赠品")
            if item.get("is_lottery"):
                markers.append("抽奖")
        except Exception:
            pass
        marker_text = f"[{'+'.join(markers)}]" if markers else ""
        base_name = (
            (item.get("name") or item.get("product_name") or item.get("title") or "未命名商品")
            if isinstance(item, dict)
            else "未命名商品"
        )
        variant = f"({item.get('variant_name')})" if isinstance(item, dict) and item.get("variant_name") else ""
        quantity = ""
        try:
            qty_val = int(item.get("quantity", 0))
            if qty_val:
                quantity = f"x{qty_val}"
        except Exception:
            quantity = ""
        item_summary_parts.append(
            " ".join(part for part in [marker_text, f"{base_name}{variant}".strip(), quantity] if part).strip()
        )
    item_summary = "\n".join([part for part in item_summary_parts if part])

    created_ms = resolve_order_timestamp_ms(order)
    created_at_text = format_device_time_ms(created_ms, tz_offset_minutes) if created_ms is not None else ""
    unified_status = compute_unified_order_status(order)

    return [
        str(order.get("id") or ""),
        owner_label or "",
        order.get("student_id") or order.get("user_id") or "",
        shipping.get("phone") or "",
        base_address,
        detail_address,
        total_text,
        item_summary,
        unified_status,
        created_at_text,
    ]


def write_export_workbook(rows: List[List[str]], file_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导出"
    header = ["订单号", "归属", "用户名", "电话", "地址", "详细地址", "订单金额", "订单信息", "订单状态", "创建时间"]
    ws.append(header)
    for row in rows:
        ws.append(row)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        try:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        except ValueError:
            max_len = 0
        adjusted_width = min(max(max_len + 4, 12), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(file_path)


def prepare_export_scope(
    staff: Dict[str, Any], agent_filter_value: Optional[str]
) -> Tuple[Optional[str], Optional[List[str]], Optional[List[str]], Optional[List[str]], Optional[List[str]], str, bool]:
    scope = build_staff_scope(staff)
    if staff.get("type") == "agent":
        return (
            scope.get("agent_id"),
            scope.get("address_ids"),
            scope.get("building_ids"),
            None,
            None,
            "self",
            False,
        )

    (
        selected_agent_id,
        selected_address_ids,
        selected_building_ids,
        exclude_address_ids,
        exclude_building_ids,
        selected_filter,
    ) = resolve_staff_order_scope(staff, scope, agent_filter_value)
    enforce_admin_only = bool(scope.get("filter_admin_orders")) if (selected_filter or "").lower() == "self" else False
    return (
        selected_agent_id,
        selected_address_ids,
        selected_building_ids,
        exclude_address_ids,
        exclude_building_ids,
        selected_filter,
        enforce_admin_only,
    )


def serialize_export_job(job: Dict[str, Any], staff_prefix: str) -> Dict[str, Any]:
    if not job:
        return {}
    download_url = None
    is_valid = False
    expires_at = job.get("expires_at")
    now = datetime.now()
    if job.get("status") == "completed" and job.get("download_token"):
        try:
            if expires_at:
                expire_dt = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
                is_valid = expire_dt > now
            else:
                is_valid = True
        except Exception:
            is_valid = False
        if is_valid:
            download_url = f"{staff_prefix}/orders/export/download/{job.get('id')}?token={job.get('download_token')}"

    range_label = format_export_range_label(job.get("start_time_ms"), job.get("end_time_ms"), job.get("client_tz_offset"))

    return {
        "id": job.get("id"),
        "status": job.get("status"),
        "created_at": job.get("created_at"),
        "expires_at": expires_at,
        "exported_count": job.get("exported_count"),
        "total_count": job.get("total_count"),
        "range_label": range_label,
        "agent_filter": job.get("agent_filter"),
        "scope_label": job.get("scope_label"),
        "status_filter": job.get("status_filter"),
        "keyword": job.get("keyword"),
        "download_url": download_url,
        "filename": job.get("filename"),
        "message": job.get("message"),
        "is_active": is_valid,
    }
