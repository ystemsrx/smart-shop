# /backend/main.py
import os
import re
import asyncio
from contextlib import asynccontextmanager
import logging
from collections import Counter
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple, Set
from pathlib import Path
from fastapi import FastAPI, HTTPException, Depends, Request, Response, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from starlette.responses import Response
from sse_starlette.sse import EventSourceResponse
from pydantic import BaseModel
import uvicorn
import json
import random
from PIL import Image
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 导入自定义模块
from database import (
    init_database, cleanup_old_chat_logs, get_db_connection,
    UserDB, ProductDB, CartDB, ChatLogDB, AdminDB, CategoryDB, OrderDB, AddressDB, BuildingDB, UserProfileDB,
    VariantDB, SettingsDB, LotteryDB, RewardDB, CouponDB, AutoGiftDB, GiftThresholdDB, LotteryConfigDB, AgentStatusDB,
    PaymentQrDB, DeliverySettingsDB, OrderExportDB
)
from database import AgentAssignmentDB, AgentDeletionDB
from auth import (
    AuthManager, get_current_user_optional, get_current_user_required,
    get_current_admin, set_auth_cookie, clear_auth_cookie,
    get_current_user_from_cookie, get_current_admin_required_from_cookie,
    get_current_user_required_from_cookie, success_response, error_response,
    get_current_staff_required_from_cookie, get_current_super_admin_required_from_cookie,
    get_current_staff_from_cookie, get_current_agent_from_cookie, is_super_admin_role, AuthError
)
from config import get_settings


settings = get_settings()


def is_truthy(value: Optional[Any]) -> bool:
    """将不同类型的输入转换为布尔值，识别常见真值表示"""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {'1', 'true', 'yes', 'on'}


def is_non_sellable(product: Dict[str, Any]) -> bool:
    """统一判断商品是否标记为非卖品"""
    if not isinstance(product, dict):
        return False
    try:
        return is_truthy(product.get('is_not_for_sale'))
    except Exception:
        return False

def convert_sqlite_timestamp_to_unix(created_at_str: str, order_id: str = None) -> int:
    """
    将SQLite的CURRENT_TIMESTAMP字符串转换为Unix时间戳（秒）
    SQLite返回的是UTC时间，需要正确处理时区
    """
    try:
        from datetime import datetime, timezone
        import time
        
        # SQLite的CURRENT_TIMESTAMP返回UTC时间，需要明确指定为UTC时区
        if " " in created_at_str:
            # SQLite格式：YYYY-MM-DD HH:MM:SS (UTC)
            dt = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
            # 明确指定为UTC时区
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            # ISO格式：YYYY-MM-DDTHH:MM:SS
            dt = datetime.fromisoformat(created_at_str)
            if dt.tzinfo is None:
                # 如果没有时区信息，假设为UTC
                dt = dt.replace(tzinfo=timezone.utc)
        
        # 转换为时间戳（秒）
        timestamp = int(dt.timestamp())
        
        # 调试日志
        current_timestamp = int(time.time())
        age_minutes = (current_timestamp - timestamp) // 60
        order_info = f"订单 {order_id}" if order_id else "时间"
        # logger在后面定义，这里暂时用print
        print(f"DEBUG: {order_info} 时间转换: {created_at_str} (UTC) -> {timestamp}, 创建于 {age_minutes} 分钟前")
        
        return timestamp
        
    except Exception as e:
        order_info = f"订单 {order_id}" if order_id else "时间"
        print(f"WARNING: {order_info} 转换失败: {e}, 原始时间: {created_at_str}")
        # 如果转换失败，使用当前时间戳减去1小时（确保倒计时能正常显示）
        import time
        return int(time.time() - 3600)

def format_device_time_ms(ms_value: Optional[float], tz_offset_minutes: Optional[int], fmt: str = "%Y-%m-%d %H:%M:%S") -> str:
    """根据设备时区偏移格式化毫秒时间戳。"""
    if ms_value is None:
        return ""
    try:
        seconds = float(ms_value) / 1000.0
        dt_utc = datetime.utcfromtimestamp(seconds)
        if tz_offset_minutes is not None:
            dt_local = dt_utc - timedelta(minutes=int(tz_offset_minutes))
        else:
            dt_local = datetime.fromtimestamp(seconds)
        return dt_local.strftime(fmt)
    except Exception:
        return ""


def format_export_range_label(start_ms: Optional[float], end_ms: Optional[float], tz_offset_minutes: Optional[int]) -> str:
    """生成导出范围的友好描述。"""
    if start_ms is None and end_ms is None:
        return "全部时间"
    start_label = format_device_time_ms(start_ms, tz_offset_minutes, "%Y-%m-%d") if start_ms is not None else ""
    end_label = format_device_time_ms(end_ms, tz_offset_minutes, "%Y-%m-%d") if end_ms is not None else ""
    if start_label and end_label:
        return f"{start_label} 至 {end_label}"
    return start_label or end_label or "全部时间"


def build_export_filename(start_ms: Optional[float], end_ms: Optional[float]) -> str:
    """根据时间范围生成导出文件名。"""
    start_part = format_device_time_ms(start_ms, None, "%Y%m%d") if start_ms is not None else "all"
    end_part = format_device_time_ms(end_ms, None, "%Y%m%d") if end_ms is not None else "all"
    timestamp_part = datetime.now().strftime("%Y%m%dT%H%M%S")
    return f"orders_{start_part}-{end_part}_{timestamp_part}.xlsx"

# 配置日志
log_level = settings.log_level.upper()
logging.basicConfig(
    level=getattr(logging, log_level, logging.INFO),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

auth_logger = logging.getLogger("auth")
auth_logger.setLevel(getattr(logging, log_level, logging.INFO))

logger = logging.getLogger(__name__)

ALLOWED_ORIGINS = settings.allowed_origins
STATIC_ALLOWED_ORIGINS = [origin for origin in ALLOWED_ORIGINS if origin != "*"]
ALLOW_ALL_ORIGINS = "*" in ALLOWED_ORIGINS
STATIC_CACHE_MAX_AGE = settings.static_cache_max_age

# FastAPI应用实例

@asynccontextmanager
async def app_lifespan(app: FastAPI):
    background_tasks: List[asyncio.Task] = []
    try:
        background_tasks = await run_startup_tasks()
        yield
    finally:
        for task in background_tasks:
            task.cancel()
        if background_tasks:
            await asyncio.gather(*background_tasks, return_exceptions=True)

app = FastAPI(
    title="宿舍智能小商城API",
    description="基于FastAPI的宿舍智能小商城后端系统",
    version="1.0.0",
    lifespan=app_lifespan
)


def build_staff_scope(staff: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """根据工作人员信息计算可访问的订单/商品范围"""
    scope = {
        "owner_ids": None,
        "address_ids": None,
        "building_ids": None,
        "is_super_admin": False,
        "agent_id": None,
        "filter_admin_orders": False  # 新增：标记是否只查询管理员的订单（agent_id IS NULL）
    }
    if not staff:
        return scope

    scope["is_super_admin"] = is_super_admin_role(staff.get('role'))

    if staff.get('type') == 'agent':
        assignments = AgentAssignmentDB.get_buildings_for_agent(staff['id'])
        building_ids = [item['building_id'] for item in assignments if item.get('building_id')]
        address_ids = list({item['address_id'] for item in assignments if item.get('address_id')})
        scope.update({
            "owner_ids": [staff['id']],
            "address_ids": address_ids,
            "building_ids": building_ids,
            "agent_id": staff['id']
        })
    else:
        # admin统一使用'admin'作为owner_id，且只查询agent_id为NULL的订单
        scope.update({
            "owner_ids": ['admin'],
            "filter_admin_orders": True  # 管理员只查询agent_id IS NULL的订单
        })

    return scope


def require_agent_with_scope(request: Request) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Ensure the requester is an agent and return (agent, scope)."""
    agent = get_current_agent_from_cookie(request)
    if not agent:
        raise HTTPException(status_code=401, detail="需要代理权限")
    scope = build_staff_scope(agent)
    return agent, scope


def get_owner_id_for_staff(staff: Dict[str, Any]) -> Optional[str]:
    """Return the owner_id used to segregate staff resources."""
    if not staff:
        return None
    if staff.get('type') == 'agent':
        return staff.get('id')
    else:
        # admin统一使用'admin'作为owner_id
        return 'admin'


def get_owner_id_from_scope(scope: Optional[Dict[str, Any]]) -> Optional[str]:
    if not scope:
        return None
    agent_id = scope.get('agent_id')
    if agent_id:
        return agent_id
    else:
        # 如果没有agent_id，说明是admin用户，使用'admin'作为owner_id
        return 'admin'


def check_address_and_building(address_id: Optional[str], building_id: Optional[str]) -> Dict[str, Any]:
    """校验地址与楼栋状态，返回可供前端与后端共用的结构"""
    result: Dict[str, Any] = {
        "is_valid": False,
        "reason": "missing_address",
        "message": "请先选择配送地址",
        "address": None,
        "building": None,
        "address_id": address_id,
        "building_id": building_id,
        "should_force_reselect": True
    }

    if not address_id:
        return result

    address = AddressDB.get_by_id(address_id)
    if not address:
        result.update({
            "reason": "address_missing",
            "message": "地址不存在，请联系管理员",
            "address": None
        })
        return result

    result["address"] = address

    address_enabled = str(address.get('enabled', 1)).strip().lower() in ('1', 'true')
    if not address_enabled:
        result.update({
            "reason": "address_disabled",
            "message": "该地址未启用，请重新选择"
        })
        return result

    if not building_id:
        result.update({
            "reason": "missing_building",
            "message": "请先选择配送地址"
        })
        return result

    building = BuildingDB.get_by_id(building_id)
    if not building:
        result.update({
            "reason": "building_missing",
            "message": "楼栋不存在或未启用，请重新选择",
            "building": None
        })
        return result

    result["building"] = building

    if building.get('address_id') != address_id:
        result.update({
            "reason": "building_mismatch",
            "message": "配送地址信息已失效，请重新选择"
        })
        return result

    building_enabled = str(building.get('enabled', 1)).strip().lower() in ('1', 'true')
    if not building_enabled:
        result.update({
            "reason": "building_disabled",
            "message": "楼栋未启用，请重新选择"
        })
        return result

    result.update({
        "is_valid": True,
        "reason": None,
        "message": "",
        "should_force_reselect": False
    })
    return result


def expire_agent_tokens_for_address(address_id: str, agent_ids: Optional[List[str]] = None) -> int:
    """让指定地址下代理的登录token立即失效"""
    if not address_id and not agent_ids:
        return 0
    ids = agent_ids if agent_ids is not None else AgentAssignmentDB.get_agent_ids_for_address(address_id)
    expired = 0
    seen: Set[str] = set()
    for agent_id in ids or []:
        if not agent_id or agent_id in seen:
            continue
        seen.add(agent_id)
        if AdminDB.bump_token_version(agent_id):
            expired += 1
    if expired:
        logger.info(f"地址 {address_id} 已使 {expired} 个代理登录状态失效")
    return expired


def fix_legacy_product_ownership():
    """修复旧系统遗留的owner_id为None的商品，分配给统一的'admin'"""
    logger.info("开始检查并修复旧商品的归属...")
    
    try:
        # 获取所有owner_id为None的商品
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM products WHERE owner_id IS NULL OR owner_id = ''")
            count = cursor.fetchone()[0]
            
            if count == 0:
                logger.info("没有发现需要修复的商品")
                return
                
            logger.info(f"发现 {count} 个需要修复归属的商品")
            
            # 确保系统中有管理员存在
            cursor.execute("SELECT COUNT(*) FROM admins WHERE (role = 'super_admin' OR role = 'admin') AND is_active = 1")
            admin_count = cursor.fetchone()[0]
            
            if admin_count > 0:
                logger.info("将使用统一的'admin'作为默认归属")
                
                # 更新所有owner_id为None的商品，统一设置为'admin'
                cursor.execute(
                    "UPDATE products SET owner_id = ? WHERE owner_id IS NULL OR owner_id = ''",
                    ('admin',)
                )
                
                updated_count = cursor.rowcount
                conn.commit()
                logger.info(f"成功修复 {updated_count} 个商品的归属")
            else:
                logger.warning("未找到可用的管理员账户，跳过商品归属修复")
                
    except Exception as e:
        logger.error(f"修复商品归属时发生错误: {e}")
        raise


def migrate_admin_products_to_unified_owner():
    """将现有admin拥有的商品迁移到统一的'admin'owner_id"""
    logger.info("开始迁移现有admin商品到统一的owner_id...")
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # 获取所有活跃的管理员ID
            cursor.execute("SELECT id FROM admins WHERE (role = 'super_admin' OR role = 'admin') AND is_active = 1")
            admin_ids = [row[0] for row in cursor.fetchall()]
            
            if not admin_ids:
                logger.info("没有找到活跃的管理员，跳过迁移")
                return
            
            # 查找所有属于这些管理员的商品
            placeholders = ', '.join(['?' for _ in admin_ids])
            cursor.execute(f"SELECT COUNT(*) FROM products WHERE owner_id IN ({placeholders})", admin_ids)
            count = cursor.fetchone()[0]
            
            if count == 0:
                logger.info("没有发现需要迁移的管理员商品")
                return
                
            logger.info(f"发现 {count} 个需要迁移到统一owner_id的管理员商品")
            
            # 更新所有属于管理员的商品，统一设置为'admin'
            cursor.execute(f"UPDATE products SET owner_id = ? WHERE owner_id IN ({placeholders})", ['admin'] + admin_ids)
            
            updated_count = cursor.rowcount
            conn.commit()
            logger.info(f"成功迁移 {updated_count} 个管理员商品到统一的'admin'归属")
                
    except Exception as e:
        logger.error(f"迁移管理员商品归属时发生错误: {e}")
        raise


def fix_legacy_config_ownership():
    """修复旧系统遗留的配置数据owner_id为None的问题，分配给统一的'admin'"""
    logger.info("开始检查并修复旧配置数据的归属...")
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # 统计需要修复的各类配置数据
            config_tables = {
                'lottery_prizes': '抽奖奖项',
                'auto_gift_items': '自动赠品',
                'gift_thresholds': '满额门槛',
                'coupons': '优惠券',
                'settings': '系统设置'
            }
            
            total_fixed = 0
            fix_summary = []
            
            for table, description in config_tables.items():
                try:
                    # 检查表是否存在且有owner_id列
                    cursor.execute(f"PRAGMA table_info({table})")
                    columns = [row[1] for row in cursor.fetchall()]
                    
                    if 'owner_id' not in columns:
                        logger.debug(f"表 {table} 没有 owner_id 列，跳过")
                        continue
                    
                    # 统计需要修复的记录数
                    cursor.execute(f"SELECT COUNT(*) FROM {table} WHERE owner_id IS NULL OR owner_id = ''")
                    count = cursor.fetchone()[0]
                    
                    if count > 0:
                        # 执行修复
                        cursor.execute(f"UPDATE {table} SET owner_id = 'admin' WHERE owner_id IS NULL OR owner_id = ''")
                        fixed_count = cursor.rowcount
                        total_fixed += fixed_count
                        fix_summary.append(f"{description}: {fixed_count}项")
                        logger.info(f"修复 {table} 表中 {fixed_count} 项配置的owner_id")
                
                except Exception as e:
                    logger.warning(f"修复表 {table} 时出错: {e}")
                    continue
            
            if total_fixed > 0:
                conn.commit()
                logger.info(f"配置数据修复完成，共修复 {total_fixed} 项：{', '.join(fix_summary)}")
            else:
                logger.info("没有发现需要修复的配置数据")
                
    except Exception as e:
        logger.error(f"修复配置数据归属时发生错误: {e}")
        raise


def resolve_owner_id_for_staff(staff: Dict[str, Any], requested_owner_id: Optional[str]) -> Optional[str]:
    """Resolve the final owner_id a staff member is allowed to use."""
    if staff.get('type') == 'agent':
        return staff.get('id')
    
    # 对于admin：统一使用'admin'作为owner_id
    if requested_owner_id:
        # 如果指定了owner_id，验证其有效性
        owner_record = AdminDB.get_admin(requested_owner_id, include_disabled=True)
        if not owner_record:
            raise HTTPException(status_code=400, detail="指定的代理不存在")
        role = (owner_record.get('role') or '').lower()
        if role != 'agent' and not is_super_admin_role(role):
            raise HTTPException(status_code=400, detail="owner_id 必须为代理账号")
        return requested_owner_id
    else:
        # admin创建的商品统一归属为'admin'
        return 'admin'


def ensure_product_accessible(staff: Dict[str, Any], product_id: str) -> Dict[str, Any]:
    product = ProductDB.get_product_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    if not staff_can_access_product(staff, product):
        raise HTTPException(status_code=403, detail="无权操作该商品")
    return product


async def store_product_image(category: str, base_name: str, image: UploadFile) -> Tuple[str, str]:
    if not image:
        raise HTTPException(status_code=400, detail="未上传图片")
    safe_category = (category or "misc").strip() or "misc"
    category_dir = os.path.join(items_dir, safe_category)
    os.makedirs(category_dir, exist_ok=True)

    timestamp = int(datetime.now().timestamp())
    # 统一使用.webp扩展名
    filename = f"{base_name}_{timestamp}.webp"
    file_path = os.path.join(category_dir, filename)

    # 读取并转换图片为webp格式
    content = await image.read()
    try:
        # 使用PIL打开图片
        img = Image.open(io.BytesIO(content))
        
        # 如果图片有RGBA模式，转换为RGB模式（webp支持更好）
        if img.mode in ('RGBA', 'LA', 'P'):
            # 创建白色背景
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # 保存为webp格式，使用用户指定的参数
        img.save(file_path, "WEBP", quality=40, method=6, optimize=True)
        
    except Exception as e:
        logger.error(f"图片处理失败: {e}")
        raise HTTPException(status_code=400, detail=f"图片处理失败: {str(e)}")

    relative_path = f"items/{safe_category}/{filename}"
    return relative_path, file_path


def normalize_reservation_cutoff(value: Optional[str]) -> Optional[str]:
    """将输入的预约截止时间标准化为 HH:MM 格式。"""
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return None
        try:
            parsed = datetime.strptime(trimmed, "%H:%M")
        except ValueError:
            raise HTTPException(status_code=400, detail="预约时间格式应为HH:MM")
        return parsed.strftime("%H:%M")
    return None


async def handle_product_creation(
    staff: Dict[str, Any],
    *,
    name: str,
    category: str,
    price: float,
    stock: int,
    description: str,
    cost: float,
    owner_id: Optional[str],
    discount: Optional[str] = None,
    variants: Optional[str] = None,
    image: Optional[UploadFile],
    is_hot: bool = False,
    is_not_for_sale: bool = False,
    reservation_required: bool = False,
    reservation_cutoff: Optional[str] = None,
    reservation_note: Optional[str] = None
) -> Dict[str, Any]:
    new_file_path: Optional[str] = None
    try:
        assigned_owner_id = resolve_owner_id_for_staff(staff, owner_id)
        img_path = ""
        if image:
            img_path, new_file_path = await store_product_image(category, name, image)

        # 处理折扣
        discount_value = 10.0
        if discount is not None:
            try:
                discount_value = float(discount)
                if discount_value < 0.5 or discount_value > 10:
                    return error_response("折扣范围应为0.5~10折", 400)
            except Exception:
                return error_response("无效的折扣", 400)

        product_data = {
            "name": name,
            "category": category,
            "price": price,
            "stock": stock,
            "discount": discount_value,
            "description": description,
            "img_path": img_path,
            "cost": cost,
            "owner_id": assigned_owner_id,
            "is_hot": 1 if is_hot else 0,
            "is_not_for_sale": 1 if is_not_for_sale else 0,
            "reservation_required": 1 if reservation_required else 0,
            "reservation_cutoff": normalize_reservation_cutoff(reservation_cutoff),
            "reservation_note": (reservation_note or '').strip()[:120]
        }

        product_id = ProductDB.create_product(product_data)
        
        # 处理变体（规格）
        if variants:
            try:
                import json
                logger.info(f"收到 variants 数据: {variants}")
                variants_list = json.loads(variants)
                logger.info(f"解析后的 variants_list: {variants_list}")
                if isinstance(variants_list, list) and len(variants_list) > 0:
                    for variant in variants_list:
                        if isinstance(variant, dict) and 'name' in variant:
                            variant_id = VariantDB.create_variant(
                                product_id=product_id,
                                name=variant['name'],
                                stock=int(variant.get('stock', 0))
                            )
                            logger.info(f"成功创建变体: {variant_id}, 名称: {variant['name']}, 库存: {variant.get('stock', 0)}")
                else:
                    logger.warning(f"variants_list 不是有效的列表或为空: {variants_list}")
            except json.JSONDecodeError as e:
                logger.error(f"创建商品变体失败 - JSON解析错误: {e}, 原始数据: {variants}")
            except Exception as e:
                logger.error(f"创建商品变体失败: {e}, 类型: {type(e).__name__}", exc_info=True)
        
        # 返回创建的商品信息，方便前端直接使用
        created_product = ProductDB.get_product_by_id(product_id)
        return success_response("商品创建成功", {"product_id": product_id, "product": created_product})

    except HTTPException as exc:
        if new_file_path:
            try:
                os.remove(new_file_path)
            except Exception:
                pass
        return error_response(exc.detail, exc.status_code)
    except Exception as e:
        if new_file_path:
            try:
                os.remove(new_file_path)
            except Exception:
                pass
        logger.error(f"创建商品失败: {e}")
        return error_response("创建商品失败", 500)


async def handle_product_update(
    staff: Dict[str, Any],
    product_id: str,
    payload: "ProductUpdateRequest"
) -> Dict[str, Any]:
    try:
        existing_product = ensure_product_accessible(staff, product_id)
    except HTTPException as exc:
        return error_response(exc.detail, exc.status_code)

    update_data: Dict[str, Any] = {}

    if payload.name is not None:
        update_data['name'] = payload.name
    if payload.category is not None:
        update_data['category'] = payload.category
    if payload.price is not None:
        update_data['price'] = payload.price
    if payload.stock is not None:
        update_data['stock'] = payload.stock
    if payload.description is not None:
        update_data['description'] = payload.description
    if payload.discount is not None:
        try:
            discount_value = float(payload.discount)
            if discount_value < 0.5 or discount_value > 10:
                return error_response("折扣范围应为0.5~10折", 400)
            update_data['discount'] = discount_value
        except Exception:
            return error_response("无效的折扣", 400)
    if payload.is_active is not None:
        update_data['is_active'] = 1 if payload.is_active else 0
    if payload.is_hot is not None:
        update_data['is_hot'] = 1 if payload.is_hot else 0
    if payload.is_not_for_sale is not None:
        update_data['is_not_for_sale'] = 1 if payload.is_not_for_sale else 0
    if payload.cost is not None:
        if payload.cost < 0:
            return error_response("商品成本不能为负数", 400)
        update_data['cost'] = payload.cost
    if payload.reservation_required is not None:
        update_data['reservation_required'] = 1 if payload.reservation_required else 0
    if payload.reservation_cutoff is not None:
        update_data['reservation_cutoff'] = normalize_reservation_cutoff(payload.reservation_cutoff)
    if payload.reservation_note is not None:
        note_value = (payload.reservation_note or '').strip()
        update_data['reservation_note'] = note_value[:120]

    if staff.get('type') == 'agent':
        update_data['owner_id'] = staff.get('id')
    elif payload.owner_id is not None:
        try:
            resolved_owner = resolve_owner_id_for_staff(staff, payload.owner_id)
        except HTTPException as exc:
            return error_response(exc.detail, exc.status_code)
        update_data['owner_id'] = resolved_owner

    if not update_data:
        return error_response("没有提供更新数据", 400)

    try:
        success = ProductDB.update_product(product_id, update_data)
        if not success:
            return error_response("更新商品失败", 500)
    except Exception as e:
        logger.error(f"更新商品失败: {e}")
        return error_response("更新商品失败", 500)

    try:
        old_is_active = int(existing_product.get('is_active', 1) or 1)
    except Exception:
        old_is_active = 1
    new_is_active = update_data.get('is_active', old_is_active)
    if old_is_active == 1 and new_is_active == 0:
        try:
            removed = CartDB.remove_product_from_all_carts(product_id)
            logger.info(f"商品 {product_id} 已下架，已从 {removed} 个购物车中移除")
        except Exception as e:
            logger.warning(f"下架后移除购物车商品失败: {e}")

    return success_response("商品更新成功")


async def handle_product_image_update(
    staff: Dict[str, Any],
    product_id: str,
    image: Optional[UploadFile]
) -> Dict[str, Any]:
    try:
        existing = ensure_product_accessible(staff, product_id)
    except HTTPException as exc:
        return error_response(exc.detail, exc.status_code)

    if not image:
        return error_response("未上传图片", 400)

    old_img_path = existing.get("img_path", "")
    category = existing.get("category", "misc") or "misc"
    base_name = existing.get("name", "prod") or "prod"

    try:
        img_path, new_file_path = await store_product_image(category, base_name, image)
    except HTTPException as exc:
        return error_response(exc.detail, exc.status_code)

    try:
        ok = ProductDB.update_image_path(product_id, img_path)
    except Exception as e:
        ok = False
        logger.error(f"商品图片数据库更新失败: {e}")

    if not ok:
        try:
            os.remove(new_file_path)
        except Exception:
            pass
        return error_response("更新图片失败", 500)

    if old_img_path and str(old_img_path).strip():
        try:
            rel_path = str(old_img_path).lstrip('/\\')
            old_file_path = os.path.normpath(os.path.join(os.path.dirname(__file__), rel_path))
            items_root = os.path.normpath(items_dir)
            if old_file_path.startswith(items_root) and os.path.exists(old_file_path):
                os.remove(old_file_path)
                logger.info(f"成功删除原图片: {old_file_path}")
            else:
                logger.warning(f"跳过删除原图片（路径不安全或不存在）: {old_img_path} -> {old_file_path}")
        except Exception as e:
            logger.warning(f"删除原图片失败 {old_img_path}: {e}")

    return success_response("图片更新成功", {
        "img_path": img_path,
        "image_url": f"/items/{img_path.split('items/')[-1]}" if img_path else ""
    })


def build_product_listing_for_staff(
    staff: Dict[str, Any],
    scope: Dict[str, Any],
    *,
    query: Optional[str] = None,
    category: Optional[str] = None,
    include_inactive: bool = True
) -> Dict[str, Any]:
    owner_ids = scope.get('owner_ids')

    # 现在所有商品都有owner_id，不再需要include_unassigned
    include_unassigned = False

    if staff.get('type') != 'agent':
        if owner_ids is None:
            # 如果没有设置owner_ids，应该设置为空列表避免查询所有商品
            owner_ids = []

    if query:
        products = ProductDB.search_products(
            query,
            active_only=not include_inactive,
            owner_ids=owner_ids,
            include_unassigned=include_unassigned
        )
    elif category:
        products = ProductDB.get_products_by_category(
            category,
            owner_ids=owner_ids,
            include_unassigned=include_unassigned
        )
    else:
        products = ProductDB.get_all_products(
            owner_ids=owner_ids,
            include_unassigned=include_unassigned
        )

    def is_active(product: Dict[str, Any]) -> bool:
        try:
            return int(product.get('is_active', 1) or 1) == 1
        except Exception:
            return True

    if not include_inactive:
        products = [p for p in products if is_active(p)]

    product_ids = [p['id'] for p in products if p.get('id')]
    variant_map = VariantDB.get_for_products(product_ids)
    for p in products:
        variants = variant_map.get(p.get('id'), [])
        p['variants'] = variants
        p['has_variants'] = len(variants) > 0  # 设置 has_variants 字段
        p['is_not_for_sale'] = is_non_sellable(p)

    categories = sorted({p.get('category') for p in products if p.get('category')})
    active_count = sum(1 for p in products if is_active(p))
    inactive_count = len(products) - active_count
    total_stock = 0
    for p in products:
        try:
            if is_non_sellable(p):
                continue
            total_stock += max(int(p.get('stock', 0) or 0), 0)
        except Exception:
            continue

    return {
        "products": products,
        "stats": {
            "total": len(products),
            "active": active_count,
            "inactive": inactive_count,
            "total_stock": total_stock
        },
        "categories": categories,
        "scope": scope
    }


def resolve_owner_filter_for_staff(
    staff: Dict[str, Any],
    scope: Dict[str, Any],
    owner_param: Optional[str]
) -> Tuple[Optional[List[str]], bool, str]:
    """解析商品/分类统计的归属过滤"""
    if staff.get('type') == 'agent':
        return scope.get('owner_ids'), bool(scope.get('is_super_admin')), 'self'

    filter_value = (owner_param or '').strip() or 'self'
    lower = filter_value.lower()

    if lower == 'self':
        # 对于admin查看自己的商品，现在统一使用'admin'作为owner_id
        return ['admin'], False, 'self'

    if lower == 'all':
        return None, True, 'all'

    target = AdminDB.get_admin(filter_value, include_disabled=True, include_deleted=True)
    if not target or (target.get('role') or '').lower() != 'agent':
        raise HTTPException(status_code=400, detail="指定的代理不存在")

    return [filter_value], False, filter_value


def resolve_single_owner_for_staff(
    staff: Dict[str, Any],
    owner_param: Optional[str]
) -> Tuple[str, str]:
    """
    解析单一 owner_id，支持管理员在查询参数中指定代理。
    返回 (owner_id, normalized_filter)
    """
    scope = build_staff_scope(staff)
    owner_ids, _, normalized_filter = resolve_owner_filter_for_staff(staff, scope, owner_param)
    if normalized_filter == 'all':
        # 这些资源不支持一次性查询全部归属
        raise HTTPException(status_code=400, detail="不支持查询全部归属的数据范围")

    if owner_ids and len(owner_ids) > 0:
        return owner_ids[0], normalized_filter

    # 回退到当前身份的默认 owner
    fallback_owner = get_owner_id_for_staff(staff)
    if not fallback_owner:
        raise HTTPException(status_code=400, detail="无法解析归属范围")
    return fallback_owner, normalized_filter


def resolve_staff_order_scope(
    staff: Dict[str, Any],
    scope: Dict[str, Any],
    agent_param: Optional[str]
) -> Tuple[Optional[str], Optional[List[str]], Optional[List[str]], Optional[List[str]], Optional[List[str]], str]:
    selected_agent_id = scope.get('agent_id')
    selected_address_ids = scope.get('address_ids')
    selected_building_ids = scope.get('building_ids')
    exclude_address_ids: Optional[List[str]] = None
    exclude_building_ids: Optional[List[str]] = None

    if staff.get('type') == 'agent':
        return (
            staff.get('id'),
            selected_address_ids,
            selected_building_ids,
            None,
            None,
            'self'
        )

    filter_value = (agent_param or '').strip() or 'self'
    lower = filter_value.lower()

    if lower == 'all':
        return None, None, None, None, None, 'all'

    if lower == 'self':
        assignments = AgentAssignmentDB.list_agents_with_buildings(include_disabled=True)
        address_set: Set[str] = set()
        building_set: Set[str] = set()
        for entry in assignments:
            for record in entry.get('buildings') or []:
                addr_id = record.get('address_id')
                bld_id = record.get('building_id')
                if addr_id:
                    address_set.add(addr_id)
                if bld_id:
                    building_set.add(bld_id)
        return (
            None,
            None,
            None,
            list(address_set) if address_set else None,
            list(building_set) if building_set else None,
            'self'
        )

    target = AdminDB.get_admin(filter_value, include_disabled=True, include_deleted=True)
    if not target or (target.get('role') or '').lower() != 'agent':
        raise HTTPException(status_code=400, detail="指定的代理不存在")

    assignments = AgentAssignmentDB.get_buildings_for_agent(filter_value)
    address_ids = list({record.get('address_id') for record in assignments if record.get('address_id')}) or None
    building_ids = [record.get('building_id') for record in assignments if record.get('building_id')]
    
    # 如果代理已被删除且没有楼栋关联，从agent_deletions表中获取历史楼栋信息
    if target.get('deleted_at') and not assignments:
        deletion_records = AgentDeletionDB.list_active_records()
        for record in deletion_records:
            if record.get('agent_id') == filter_value:
                address_ids = record.get('address_ids') or None
                building_ids = record.get('building_ids') or None
                break

    return filter_value, address_ids, building_ids, None, None, filter_value

def compute_unified_order_status(order: Dict[str, Any]) -> str:
    """统一获取订单状态标签。"""
    ps = order.get("payment_status") if isinstance(order, dict) else None
    st = order.get("status") if isinstance(order, dict) else None
    if not ps and not st:
        return "未付款"
    if ps == "processing":
        return "待确认"
    if ps != "succeeded":
        return "未付款"
    if st == "shipped":
        return "配送中"
    if st == "delivered":
        return "已完成"
    return "待配送"


def resolve_order_timestamp_ms(order: Dict[str, Any]) -> Optional[float]:
    """优先使用已有时间戳，否则从字符串转换为毫秒。"""
    if not isinstance(order, dict):
        return None
    raw_ts = order.get("created_at_timestamp")
    if isinstance(raw_ts, (int, float)):
        return float(raw_ts) * 1000
    created_at_str = order.get("created_at")
    if created_at_str:
        try:
            return float(convert_sqlite_timestamp_to_unix(created_at_str, order.get("id"))) * 1000
        except Exception:
            return None
    return None


def build_agent_name_map() -> Dict[str, str]:
    """获取代理名称映射，包含禁用/已删除的代理方便展示历史数据。"""
    mapping: Dict[str, str] = {}
    agents = AdminDB.list_admins(role='agent', include_disabled=True, include_deleted=True)
    for agent in agents:
        agent_id = agent.get('id')
        if not agent_id:
            continue
        mapping[agent_id] = agent.get('name') or agent_id
    return mapping


def resolve_scope_label(selected_filter: Optional[str], staff: Dict[str, Any], agent_name_map: Dict[str, str]) -> str:
    """根据过滤条件生成归属范围标签。"""
    if staff.get('type') == 'agent':
        return staff.get('name') or staff.get('id') or '我的订单'
    lower = (selected_filter or 'self').lower()
    if lower == 'all':
        return '全部订单'
    if lower == 'self':
        return '管理员订单'
    return f"{agent_name_map.get(selected_filter, selected_filter)} 的订单"


def resolve_order_owner_label(order: Dict[str, Any], agent_name_map: Dict[str, str], staff: Dict[str, Any], is_admin_role: bool) -> str:
    agent_id = order.get("agent_id") if isinstance(order, dict) else None
    if is_admin_role:
        if agent_id:
            return agent_name_map.get(agent_id) or agent_id
        return staff.get('name') or staff.get('id') or '管理员'
    if agent_id:
        return agent_name_map.get(agent_id) or agent_id
    return staff.get('name') or staff.get('id') or '我的订单'


def build_export_row(order: Dict[str, Any], agent_name_map: Dict[str, str], staff: Dict[str, Any], is_admin_role: bool, tz_offset_minutes: Optional[int]) -> List[str]:
    shipping = order.get('shipping_info') if isinstance(order, dict) and isinstance(order.get('shipping_info'), dict) else {}
    owner_label = resolve_order_owner_label(order, agent_name_map, staff, is_admin_role)
    address_parts = [shipping.get('dormitory'), shipping.get('building')]
    base_address = ' '.join([part for part in address_parts if part]) or shipping.get('full_address') or ''
    detail_segments = [
        shipping.get('room'),
        shipping.get('address_detail'),
        shipping.get('detail'),
        shipping.get('extra')
    ]
    detail_address = ' '.join([seg for seg in detail_segments if seg]) or ''

    total_value = order.get('total_amount')
    total_text = f"{float(total_value):.2f}" if isinstance(total_value, (int, float)) else str(total_value or '')

    items = order.get('items') if isinstance(order, dict) and isinstance(order.get('items'), list) else []
    item_summary_parts: List[str] = []
    for item in items:
        if not item:
            continue
        markers: List[str] = []
        try:
            if item.get('is_auto_gift'):
                markers.append('赠品')
            if item.get('is_lottery'):
                markers.append('抽奖')
        except Exception:
            pass
        marker_text = f"[{'+'.join(markers)}]" if markers else ''
        base_name = (item.get('name') or item.get('product_name') or item.get('title') or '未命名商品') if isinstance(item, dict) else '未命名商品'
        variant = f"({item.get('variant_name')})" if isinstance(item, dict) and item.get('variant_name') else ''
        quantity = ''
        try:
            qty_val = int(item.get('quantity', 0))
            if qty_val:
                quantity = f"x{qty_val}"
        except Exception:
            quantity = ''
        item_summary_parts.append(' '.join(part for part in [marker_text, f"{base_name}{variant}".strip(), quantity] if part).strip())
    item_summary = "\n".join([part for part in item_summary_parts if part])

    created_ms = resolve_order_timestamp_ms(order)
    created_at_text = format_device_time_ms(created_ms, tz_offset_minutes) if created_ms is not None else ''
    unified_status = compute_unified_order_status(order)

    return [
        str(order.get('id') or ''),
        owner_label or '',
        order.get('student_id') or order.get('user_id') or '',
        shipping.get('phone') or '',
        base_address,
        detail_address,
        total_text,
        item_summary,
        unified_status,
        created_at_text
    ]

def write_export_workbook(rows: List[List[str]], file_path: str) -> None:
    """将行数据写入xlsx文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导出"
    header = ['订单号', '归属', '用户名', '电话', '地址', '详细地址', '订单金额', '订单信息', '订单状态', '创建时间']
    ws.append(header)
    for row in rows:
        ws.append(row)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        try:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        except ValueError:
            max_len = 0
        adjusted_width = min(max(max_len + 4, 12), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(file_path)

def prepare_export_scope(staff: Dict[str, Any], agent_filter_value: Optional[str]) -> Tuple[
    Optional[str], Optional[List[str]], Optional[List[str]], Optional[List[str]], Optional[List[str]], str, bool
]:
    """根据身份和筛选值计算导出范围及是否限制仅管理员订单。"""
    scope = build_staff_scope(staff)
    if staff.get('type') == 'agent':
        return (
            scope.get('agent_id'),
            scope.get('address_ids'),
            scope.get('building_ids'),
            None,
            None,
            'self',
            False
        )

    selected_agent_id, selected_address_ids, selected_building_ids, exclude_address_ids, exclude_building_ids, selected_filter = resolve_staff_order_scope(
        staff,
        scope,
        agent_filter_value
    )
    enforce_admin_only = bool(scope.get('filter_admin_orders')) if (selected_filter or '').lower() == 'self' else False
    return (
        selected_agent_id,
        selected_address_ids,
        selected_building_ids,
        exclude_address_ids,
        exclude_building_ids,
        selected_filter,
        enforce_admin_only
    )


def serialize_export_job(job: Dict[str, Any], staff_prefix: str) -> Dict[str, Any]:
    """转换导出记录为前端可用的数据结构。"""
    if not job:
        return {}
    download_url = None
    is_valid = False
    expires_at = job.get('expires_at')
    now = datetime.now()
    if job.get('status') == 'completed' and job.get('download_token'):
        try:
            if expires_at:
                expire_dt = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
                is_valid = expire_dt > now
            else:
                is_valid = True
        except Exception:
            is_valid = False
        if is_valid:
            download_url = f"{staff_prefix}/orders/export/download/{job.get('id')}?token={job.get('download_token')}"

    range_label = format_export_range_label(job.get('start_time_ms'), job.get('end_time_ms'), job.get('client_tz_offset'))

    return {
        "id": job.get('id'),
        "status": job.get('status'),
        "created_at": job.get('created_at'),
        "expires_at": expires_at,
        "exported_count": job.get('exported_count'),
        "total_count": job.get('total_count'),
        "range_label": range_label,
        "agent_filter": job.get('agent_filter'),
        "scope_label": job.get('scope_label'),
        "status_filter": job.get('status_filter'),
        "keyword": job.get('keyword'),
        "download_url": download_url,
        "filename": job.get('filename'),
        "message": job.get('message'),
        "is_active": is_valid
    }


async def handle_product_stock_update(
    staff: Dict[str, Any],
    product_id: str,
    stock_data: "StockUpdateRequest"
) -> Dict[str, Any]:
    try:
        ensure_product_accessible(staff, product_id)
    except HTTPException as exc:
        return error_response(exc.detail, exc.status_code)

    if stock_data.stock < 0:
        return error_response("库存不能为负数", 400)

    try:
        success = ProductDB.update_stock(product_id, stock_data.stock)
    except Exception as e:
        logger.error(f"更新库存失败: {e}")
        return error_response("更新库存失败", 500)

    if not success:
        return error_response("更新库存失败", 500)

    return success_response("库存更新成功", {"new_stock": stock_data.stock})


def serialize_agent_account(agent: Dict[str, Any], include_buildings: bool = True) -> Dict[str, Any]:
    """将数据库中的管理员记录转换为前端需要的结构"""
    data = {
        "id": agent.get('id'),
        "name": agent.get('name'),
        "role": agent.get('role'),
        "type": 'agent' if (agent.get('role') or '').lower() == 'agent' else 'admin',
        "created_at": agent.get('created_at'),
        "payment_qr_path": agent.get('payment_qr_path'),
        "is_active": False if str(agent.get('is_active', 1)).strip() in ('0', 'False', 'false') else True,
        "deleted_at": agent.get('deleted_at'),
        "is_deleted": bool(agent.get('deleted_at'))
    }
    if include_buildings:
        data["buildings"] = AgentAssignmentDB.get_buildings_for_agent(agent.get('id'))
    return data


def compute_registered_user_count(owner_ids: Optional[List[str]]) -> int:
    """
    根据归属范围统计注册用户数量。
    - owner_ids 为 None 时统计所有用户
    - owner_ids 包含 'admin' 时同样统计所有用户
    - 其余情况根据代理分配的地址/楼栋统计
    """
    try:
        if not owner_ids:
            return UserProfileDB.count_users_by_scope()

        agent_ids = [oid for oid in owner_ids if oid and oid != 'admin']
        if not agent_ids:
            return UserProfileDB.count_users_by_scope()

        address_ids: Set[str] = set()
        building_ids: Set[str] = set()
        for agent_id in agent_ids:
            assignments = AgentAssignmentDB.get_buildings_for_agent(agent_id)
            for record in assignments or []:
                addr = record.get('address_id')
                bld = record.get('building_id')
                if addr:
                    address_ids.add(addr)
                if bld:
                    building_ids.add(bld)

        agent_id_filter = agent_ids[0] if len(agent_ids) == 1 else None
        if agent_id_filter and not address_ids and not building_ids:
            return UserProfileDB.count_users_by_scope(agent_id=agent_id_filter)

        return UserProfileDB.count_users_by_scope(
            address_ids=list(address_ids),
            building_ids=list(building_ids),
            agent_id=agent_id_filter
        )
    except Exception as exc:
        logger.error(f"计算注册用户数量失败: {exc}")
        return 0


def validate_building_ids(building_ids: Optional[List[str]]) -> Tuple[List[str], List[str]]:
    valid: List[str] = []
    invalid: List[str] = []
    if not building_ids:
        return valid, invalid
    seen = set()
    for bid in building_ids:
        if not bid or bid in seen:
            continue
        seen.add(bid)
        building = BuildingDB.get_by_id(bid)
        if building:
            valid.append(bid)
        else:
            invalid.append(bid)
    return valid, invalid


def resolve_shopping_scope(request: Request, address_id: Optional[str] = None, building_id: Optional[str] = None) -> Dict[str, Optional[str]]:
    """根据请求参数和用户资料确定购物范围与归属代理"""
    resolved_address_id = address_id
    resolved_building_id = building_id
    agent_id: Optional[str] = None

    staff = get_current_staff_from_cookie(request)
    if staff and staff.get('type') == 'agent':
        staff_agent_id = staff.get('id')
        owner_ids = [staff_agent_id] if staff_agent_id else None
        return {
            "agent_id": staff_agent_id,
            "address_id": None,
            "building_id": None,
            "owner_ids": owner_ids
        }

    user = get_current_user_from_cookie(request)
    if user:
        logger.info(f"resolve_shopping_scope - 用户: {user['id']}")
        profile = UserProfileDB.get_shipping(user['id'])
        logger.info(f"resolve_shopping_scope - 用户配置: {profile}")
        if profile:
            if not resolved_address_id:
                resolved_address_id = profile.get('address_id') or profile.get('dormitory')
            if not resolved_building_id:
                resolved_building_id = profile.get('building_id')
        logger.info(f"resolve_shopping_scope - 解析后地址: {resolved_address_id}, 楼栋: {resolved_building_id}")
    else:
        logger.warning("resolve_shopping_scope - 未获取到用户信息")

    if resolved_address_id or resolved_building_id:
        validation = check_address_and_building(resolved_address_id, resolved_building_id)
        if not validation["is_valid"]:
            resolved_address_id = None
            resolved_building_id = None
    else:
        validation = check_address_and_building(None, None)

    # 修复Agent商品权限控制：
    # 1. 如果选择了具体楼栋，检查该楼栋是否分配给Agent
    # 2. 如果选择了地址但没有具体楼栋，检查该地址下是否有Agent分配
    # 3. 只有在明确有Agent分配的情况下，才限制显示Agent商品
    if resolved_building_id:
        assignment = AgentAssignmentDB.get_agent_for_building(resolved_building_id)
        if assignment and assignment.get('agent_id'):
            # 楼栋被分配给了Agent，显示该Agent的商品
            agent_id = assignment['agent_id']
            if not resolved_address_id:
                resolved_address_id = assignment.get('address_id')
    elif resolved_address_id:
        agents = AgentAssignmentDB.get_agent_ids_for_address(resolved_address_id)
        # 只有当地址下只有一个Agent时，才限制显示该Agent的商品
        # 如果地址下有多个Agent或没有Agent，则显示Admin商品
        if len(agents) == 1:
            agent_id = agents[0]
        # 如果len(agents) == 0 或 > 1，则agent_id保持为None，显示Admin商品

    # 修改owner_ids逻辑：
    # - 如果找到了唯一的agent_id，则只显示该Agent的商品
    # - 如果没有找到agent_id，则显示Admin的商品（统一使用'admin'）
    if agent_id:
        owner_ids = [agent_id]
    else:
        # 没有找到Agent，显示统一的Admin商品
        try:
            # 确认系统中有管理员存在
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM admins WHERE (role = 'super_admin' OR role = 'admin') AND is_active = 1")
                admin_count = cursor.fetchone()[0]
                if admin_count > 0:
                    owner_ids = ['admin']  # 使用统一的'admin'显示所有Admin商品
                else:
                    owner_ids = None  # 回退到显示未分配商品
        except Exception:
            owner_ids = None  # 出错时回退到显示未分配商品

    result = {
        "agent_id": agent_id,
        "address_id": resolved_address_id,
        "building_id": resolved_building_id,
        "owner_ids": owner_ids,
        "address_validation": validation
    }
    logger.info(f"resolve_shopping_scope - 最终结果: {result}")
    return result


def staff_can_access_product(staff: Dict[str, Any], product: Optional[Dict[str, Any]]) -> bool:
    if not product:
        return False
    if staff.get('type') == 'agent':
        return product.get('owner_id') == staff.get('id')
    return True


def staff_can_access_order(staff: Dict[str, Any], order: Optional[Dict[str, Any]], scope: Optional[Dict[str, Any]] = None) -> bool:
    if not order:
        return False
    scope = scope or build_staff_scope(staff)
    if scope.get('is_super_admin'):
        return True
    agent_id = scope.get('agent_id')
    if agent_id:
        if order.get('agent_id') == agent_id:
            return True
        buildings = scope.get('building_ids') or []
        addresses = scope.get('address_ids') or []
        if order.get('building_id') and order.get('building_id') in buildings:
            return True
        if order.get('address_id') and order.get('address_id') in addresses:
            return True
    return False

# CORS配置
cors_allow_origins = ["*"] if ALLOW_ALL_ORIGINS else ALLOWED_ORIGINS
cors_allow_credentials = not ALLOW_ALL_ORIGINS

if ALLOW_ALL_ORIGINS and not cors_allow_credentials:
    logger.warning("检测到通配符跨域设置，已禁用凭据共享以符合CORS规范。")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_allow_origins,
    allow_credentials=cors_allow_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition", "Content-Length", "Content-Type"],
)

# 静态文件服务
items_dir = os.path.join(os.path.dirname(__file__), "items")
os.makedirs(items_dir, exist_ok=True)

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
public_dir = os.path.join(project_root, "public")
os.makedirs(public_dir, exist_ok=True)
exports_dir = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(exports_dir, exist_ok=True)

class CachedStaticFiles(StaticFiles):
    def __init__(self, *args, max_age: int = STATIC_CACHE_MAX_AGE, **kwargs):
        super().__init__(*args, **kwargs)
        self._max_age = max_age

    async def get_response(self, path, scope):
        resp = await super().get_response(path, scope)
        if hasattr(resp, 'headers') and resp.status_code == 200:
            resp.headers["Cache-Control"] = f"public, max-age={self._max_age}, immutable"
            # CORS for static assets (images) to allow cross-origin caching/fetch
            try:
                origin = None
                for k, v in scope.get('headers', []):
                    if k.decode().lower() == 'origin':
                        origin = v.decode()
                        break
                allowed = STATIC_ALLOWED_ORIGINS
                if origin and origin in allowed:
                    resp.headers["Access-Control-Allow-Origin"] = origin
                    resp.headers["Vary"] = "Origin"
                elif ALLOW_ALL_ORIGINS:
                    resp.headers["Access-Control-Allow-Origin"] = "*"
                    resp.headers.pop("Vary", None)
                else:
                    resp.headers["Access-Control-Allow-Origin"] = "*"
                    resp.headers.pop("Vary", None)
                resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
                resp.headers["Access-Control-Allow-Headers"] = "*"
            except Exception:
                pass
        return resp

# Wrap static app with CORS to ensure ACAO header is set for images
_static = CachedStaticFiles(directory=items_dir, max_age=STATIC_CACHE_MAX_AGE)
_static_cors = CORSMiddleware(
    _static,
    allow_origins=cors_allow_origins,
    allow_methods=["GET", "OPTIONS"],
    allow_headers=["*"],
    allow_credentials=cors_allow_credentials,
    expose_headers=["Content-Length", "Content-Type"]
)
app.mount("/items", _static_cors, name="items")

# Mount public directory for payment QR codes and other dynamically generated static assets
_public_static = CachedStaticFiles(directory=public_dir, max_age=STATIC_CACHE_MAX_AGE)
_public_static_cors = CORSMiddleware(
    _public_static,
    allow_origins=cors_allow_origins,
    allow_methods=["GET", "OPTIONS"],
    allow_headers=["*"],
    allow_credentials=cors_allow_credentials,
    expose_headers=["Content-Length", "Content-Type"]
)
app.mount("/public", _public_static_cors, name="public")

# Also serve public files directly at root level (for compatibility with Next.js public directory behavior)
from fastapi import HTTPException
from starlette.staticfiles import StaticFiles
from starlette.responses import FileResponse
import mimetypes


# Pydantic模型
class LoginRequest(BaseModel):
    student_id: str
    password: str

class AdminLoginRequest(BaseModel):
    admin_id: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    nickname: Optional[str] = None  # 昵称，选填

class ProductCreate(BaseModel):
    name: str
    category: str
    price: float
    stock: int = 0
    description: str = ""

class CartUpdateRequest(BaseModel):
    action: str  # add, update, remove, clear
    product_id: Optional[str] = None
    quantity: Optional[int] = None
    variant_id: Optional[str] = None

class ChatMessage(BaseModel):
    role: str
    content: Optional[str] = None  # content 可以为 None（assistant 消息如果只有 tool_calls）
    tool_calls: Optional[List[Dict[str, Any]]] = None  # assistant 消息可能包含 tool_calls
    tool_call_id: Optional[str] = None  # tool 消息必须包含 tool_call_id

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    model: Optional[str] = None
    conversation_id: Optional[str] = None

class ChatThreadCreateRequest(BaseModel):
    title: Optional[str] = None

class ChatThreadUpdateRequest(BaseModel):
    title: Optional[str] = None

class ProductUpdateRequest(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    price: Optional[float] = None
    stock: Optional[int] = None
    discount: Optional[float] = None  # 折扣（以折为单位，10为不打折，0.5为五折）
    description: Optional[str] = None
    is_active: Optional[bool] = None
    is_hot: Optional[bool] = None
    is_not_for_sale: Optional[bool] = None
    cost: Optional[float] = None  # 商品成本
    owner_id: Optional[str] = None
    reservation_required: Optional[bool] = None
    reservation_cutoff: Optional[str] = None
    reservation_note: Optional[str] = None

class StockUpdateRequest(BaseModel):
    stock: int

class CategoryCreateRequest(BaseModel):
    name: str
    description: str = ""

class CategoryUpdateRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class ProductDeleteRequest(BaseModel):
    product_ids: List[str]

class BulkProductUpdateRequest(BaseModel):
    product_ids: List[str]
    discount: Optional[float] = None
    owner_id: Optional[str] = None
    is_active: Optional[bool] = None


class AgentCreateRequest(BaseModel):
    account: str
    password: str
    name: str
    building_ids: List[str] = []


class AgentUpdateRequest(BaseModel):
    password: Optional[str] = None
    name: Optional[str] = None
    building_ids: Optional[List[str]] = None
    is_active: Optional[bool] = None


class LocationUpdateRequest(BaseModel):
    address_id: str
    building_id: str

class OrderCreateRequest(BaseModel):
    shipping_info: Dict[str, str]
    payment_method: str = 'wechat'
    note: str = ''
    coupon_id: Optional[str] = None  # 选用的优惠券ID（可选）
    apply_coupon: Optional[bool] = True  # 是否应用优惠券（默认应用）
    reservation_requested: Optional[bool] = False  # 用户是否选择以预约的方式提交

class OrderStatusUpdateRequest(BaseModel):
    status: str


class PaymentStatusUpdateRequest(BaseModel):
    payment_status: str

class OrderExportRequest(BaseModel):
    start_time_ms: Optional[float] = None
    end_time_ms: Optional[float] = None
    status_filter: Optional[str] = None
    keyword: Optional[str] = None
    agent_filter: Optional[str] = None
    timezone_offset_minutes: Optional[int] = None

# 地址管理模型
class AddressCreateRequest(BaseModel):
    name: str
    enabled: bool = True
    sort_order: int = 0

class AddressUpdateRequest(BaseModel):
    name: Optional[str] = None
    enabled: Optional[bool] = None
    sort_order: Optional[int] = None

# 楼栋管理模型
class BuildingCreateRequest(BaseModel):
    address_id: str
    name: str
    enabled: bool = True
    sort_order: int = 0

class BuildingUpdateRequest(BaseModel):
    name: Optional[str] = None
    enabled: Optional[bool] = None
    sort_order: Optional[int] = None

class AddressReorderRequest(BaseModel):
    order: List[str]

class BuildingReorderRequest(BaseModel):
    address_id: str
    order: List[str]

# 优惠券模型
class CouponIssueRequest(BaseModel):
    student_id: str
    amount: float
    quantity: int = 1
    # ISO字符串或 'YYYY-MM-DD HH:MM:SS'，为空代表永久
    expires_at: Optional[str] = None

def log_model_configuration_snapshot() -> None:
    """记录环境变量与最终模型列表之间的差异，便于排查选择器缺少模型的问题。"""
    env_models_raw = os.getenv("MODEL", "")
    env_labels_raw = os.getenv("MODEL_NAME", "")
    supports_raw = os.getenv("SUPPORTS_THINKING", "")

    env_models = [item.strip() for item in env_models_raw.split(",") if item.strip()]
    env_labels = [item.strip() for item in env_labels_raw.split(",") if item.strip()]
    supports_flags = {item.strip().lower() for item in supports_raw.split(",") if item.strip()}

    configured_models = settings.model_order
    configured_names = [cfg.name for cfg in configured_models]

    if not configured_models:
        logger.error("模型选择器没有可用模型，请检查 MODEL/MODEL_NAME 环境变量。")
        return

    if env_models:
        duplicate_models = [name for name, count in Counter(env_models).items() if count > 1]
        if duplicate_models:
            logger.warning("MODEL 环境变量中存在重复模型: %s", duplicate_models)

        if len(env_models) != len(env_labels):
            logger.warning(
                "MODEL 与 MODEL_NAME 的数量不一致：MODEL=%d, MODEL_NAME=%d。多余的模型将不会出现在选择器中。",
                len(env_models),
                len(env_labels),
            )

        missing_models = [name for name in env_models if name not in configured_names]
        if missing_models:
            logger.warning(
                "以下模型在环境变量中配置但未被加载：%s。"
                "请确认模型名称与 MODEL_NAME 一一对应，并在修改 .env 后重新启动后端服务。",
                missing_models,
            )

    logger.info(
        "模型选择器当前可用模型：%s",
        [
            {
                "model": cfg.name,
                "label": cfg.label,
                "supports_thinking": cfg.supports_thinking,
            }
            for cfg in configured_models
        ],
    )

    logger.debug(
        "模型配置原始环境变量：MODEL=%r, MODEL_NAME=%r, SUPPORTS_THINKING=%r（解析后=%s）。最终加载模型=%s，supports_thinking=%s。",
        env_models_raw,
        env_labels_raw,
        supports_raw,
        sorted(supports_flags),
        configured_names,
        [cfg.supports_thinking for cfg in configured_models],
    )
    
    # 额外调试：检查 settings 对象的 ID 和 model_order 列表的 ID
    logger.debug(f"settings 对象 ID: {id(settings)}, model_order 列表 ID: {id(settings.model_order)}")
    logger.debug(f"get_settings() 缓存状态: {get_settings.cache_info()}")
    
    # 直接调用 get_settings() 检查是否与当前 settings 相同
    fresh_settings = get_settings()
    logger.debug(f"get_settings() 返回对象 ID: {id(fresh_settings)}, 是否为同一对象: {fresh_settings is settings}")
    logger.debug(f"get_settings().model_order 长度: {len(fresh_settings.model_order)}, 列表: {[cfg.name for cfg in fresh_settings.model_order]}")

    if env_models:
        stale_models = [name for name in configured_names if name not in env_models]
        if stale_models:
            logger.debug(
                "模型列表包含未在当前 MODEL 环境变量中的条目：%s。"
                "若该情况出乎意料，请清理配置缓存或确认运行环境中没有其他来源的默认模型。",
                stale_models,
            )

# 启动事件
async def run_startup_tasks() -> List[asyncio.Task]:
    """应用启动时初始化并启动后台任务，返回需要在关闭时清理的任务列表"""
    logger.info("正在启动宿舍智能小商城API...")
    
    # 初始化数据库（包含收款码数据清理和迁移）
    init_database()
    
    # 启动时清理无商品的空分类
    try:
        CategoryDB.cleanup_orphan_categories()
    except Exception as e:
        logger.warning(f"启动时清理空分类失败: {e}")
    
    # 修复旧系统遗留的owner_id为None的商品，分配给默认admin
    try:
        fix_legacy_product_ownership()
    except Exception as e:
        logger.warning(f"修复旧商品归属失败: {e}")
    
    # 迁移现有admin商品到统一的'admin'owner_id
    try:
        migrate_admin_products_to_unified_owner()
    except Exception as e:
        logger.warning(f"迁移admin商品归属失败: {e}")
    
    # 修复旧系统遗留的配置数据owner_id为None的问题
    try:
        fix_legacy_config_ownership()
    except Exception as e:
        logger.warning(f"修复旧配置数据归属失败: {e}")
    
    # 启动定时清理任务
    maintenance_tasks: List[asyncio.Task] = []
    maintenance_tasks.append(asyncio.create_task(periodic_cleanup(), name="periodic_cleanup"))
    # 每分钟清理一次过期未付款订单
    maintenance_tasks.append(asyncio.create_task(expired_unpaid_cleanup(), name="expired_unpaid_cleanup"))
    
    log_model_configuration_snapshot()
    logger.info("宿舍智能小商城API启动完成")
    return maintenance_tasks

async def periodic_cleanup():
    """定时清理任务"""
    while True:
        try:
            # 每天凌晨3点清理过期聊天记录
            await asyncio.sleep(24 * 60 * 60)  # 24小时
            cleanup_old_chat_logs()
            # 顺带清理无商品的空分类
            try:
                CategoryDB.cleanup_orphan_categories()
            except Exception as e:
                logger.warning(f"定时清理空分类失败: {e}")
            try:
                removed_exports = OrderExportDB.cleanup_expired_files(exports_dir)
                if removed_exports:
                    logger.info(f"清理过期导出文件 {removed_exports} 个")
            except Exception as e:
                logger.warning(f"清理导出文件失败: {e}")
        except Exception as e:
            logger.error(f"定时清理任务失败: {e}")

async def expired_unpaid_cleanup():
    """清理超过15分钟未付款订单"""
    while True:
        try:
            await asyncio.sleep(60)
            from database import OrderDB
            deleted = OrderDB.purge_expired_unpaid_orders(15)
            if deleted:
                logger.info(f"清理过期未付款订单: 删除 {deleted} 笔")
        except Exception as e:
            logger.error(f"清理过期未付款订单任务失败: {e}")

# ==================== 认证路由 ====================

@app.post("/auth/login")
async def login(request: LoginRequest, response: Response):
    """用户登录"""
    try:
        try:
            staff_result = AuthManager.login_admin(request.student_id, request.password)
        except AuthError as exc:
            return error_response(exc.message, exc.status_code)
        if staff_result:
            set_auth_cookie(response, staff_result["access_token"])
            return success_response("登录成功", staff_result)

        result = await AuthManager.login_user(request.student_id, request.password)
        if not result:
            return error_response("账号或密码错误", 401)

        # 设置Cookie
        set_auth_cookie(response, result["access_token"])

        return success_response("登录成功", result)
    
    except Exception as e:
        logger.error(f"登录失败: {e}")
        return error_response("登录失败，请稍后重试", 500)

@app.post("/auth/admin-login")
async def admin_login(request: AdminLoginRequest, response: Response):
    """管理员登录"""
    try:
        try:
            result = AuthManager.login_admin(request.admin_id, request.password)
        except AuthError as exc:
            return error_response(exc.message, exc.status_code)
        if not result:
            return error_response("账号或密码错误", 401)

        # 设置Cookie
        set_auth_cookie(response, result["access_token"])

        return success_response("管理员登录成功", result)
    
    except Exception as e:
        logger.error(f"管理员登录失败: {e}")
        return error_response("管理员登录失败，请稍后重试", 500)

@app.post("/auth/logout")
async def logout(response: Response):
    """用户登出"""
    clear_auth_cookie(response)
    return success_response("登出成功")

@app.get("/auth/me")
async def get_current_user_info(request: Request):
    """获取当前用户信息"""
    # 首先尝试作为用户获取
    user = get_current_user_from_cookie(request)
    if user:
        return success_response("获取用户信息成功", user)

    # 如果不是用户，尝试作为管理员获取
    from auth import get_current_staff_from_cookie
    admin = get_current_staff_from_cookie(request)
    if admin:
        return success_response("获取工作人员信息成功", admin)

    return error_response("未登录", 401)

@app.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    """刷新令牌"""
    # 首先尝试作为用户获取
    user = get_current_user_from_cookie(request)
    if user:
        # 生成新的用户令牌
        token_data = {
            "sub": user["id"],
            "type": "user",
            "name": user["name"]
        }
        new_token = AuthManager.create_access_token(token_data)
        set_auth_cookie(response, new_token)
        return success_response("令牌刷新成功", {"access_token": new_token})
    
    # 如果不是用户，尝试作为管理员获取
    from auth import get_current_admin_from_cookie
    admin = get_current_admin_from_cookie(request)
    if admin:
        # 生成新的管理员令牌
        token_data = {
            "sub": admin["id"],
            "type": "admin",
            "name": admin["name"],
            "role": admin["role"]
        }
        new_token = AuthManager.create_access_token(token_data)
        set_auth_cookie(response, new_token)
        return success_response("管理员令牌刷新成功", {"access_token": new_token})
    
    return error_response("令牌无效", 401)

# ==================== 注册相关接口 ====================

@app.get("/auth/registration-status")
async def get_registration_status():
    """获取注册功能是否启用"""
    try:
        # 默认关闭注册功能，管理员可手动开启
        enabled = SettingsDB.get('registration_enabled', 'false').lower() == 'true'
        reservation_enabled = SettingsDB.get('shop_reservation_enabled', 'false') == 'true'
        return success_response("获取注册状态成功", {
            "enabled": enabled,
            "reservation_enabled": reservation_enabled
        })
    except Exception as e:
        logger.error(f"获取注册状态失败: {e}")
        return error_response("获取注册状态失败", 500)

@app.post("/auth/register")
async def register_user(request: RegisterRequest, response: Response):
    """用户注册"""
    try:
        # 检查注册功能是否启用（默认关闭）
        enabled = SettingsDB.get('registration_enabled', 'false').lower() == 'true'
        if not enabled:
            return error_response("注册功能未启用", 403)
        
        # 验证用户名和密码
        username = request.username.strip()
        password = request.password.strip()
        
        # 用户名验证：至少2个字符
        if len(username) < 2:
            return error_response("用户名至少需要2个字符", 400)
        
        # 密码验证：需包含数字和字母
        import re
        if len(password) < 6:
            return error_response("密码至少需要6个字符", 400)
        
        has_letter = bool(re.search(r'[a-zA-Z]', password))
        has_digit = bool(re.search(r'\d', password))
        
        if not (has_letter and has_digit):
            return error_response("密码必须包含数字和字母", 400)
        
        # 检查用户名是否已存在（包括普通用户、管理员和代理）
        existing_user = UserDB.get_user(username)
        if existing_user:
            return error_response("用户名已存在", 400)
        
        # 检查管理员表中是否存在相同用户名
        existing_admin = AdminDB.get_admin(username)
        if existing_admin:
            return error_response("用户名已存在", 400)
        
        # 创建用户 - 如果有昵称则使用昵称，否则使用用户名作为姓名
        display_name = request.nickname.strip() if request.nickname and request.nickname.strip() else username
        success = UserDB.create_user(username, password, display_name, id_status=2)
        if not success:
            return error_response("注册失败，请稍后重试", 500)
        
        # 自动登录
        result = await AuthManager.login_user(username, password)
        if result:
            set_auth_cookie(response, result["access_token"])
            return success_response("注册成功，已自动登录", result)
        else:
            return error_response("注册成功但自动登录失败，请手动登录", 500)
            
    except Exception as e:
        logger.error(f"用户注册失败: {e}")
        return error_response("注册失败，请稍后重试", 500)

@app.post("/admin/registration-settings")
async def update_registration_settings(request: Request):
    """管理员更新注册/预约设置"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        params = request.query_params or {}
        enabled_param = params.get('enabled')
        reservation_param = params.get('reservation_enabled')
        payload: Dict[str, Any] = {}

        content_type = request.headers.get('content-type', '').lower()
        if 'application/json' in content_type:
            try:
                payload = await request.json()
            except Exception:
                payload = {}

        def resolve_bool(value: Any) -> Optional[bool]:
            if value is None:
                return None
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return bool(value)
            return is_truthy(str(value))

        enabled_value = resolve_bool(payload.get('enabled')) if 'enabled' in payload else resolve_bool(enabled_param)
        reservation_value = resolve_bool(payload.get('reservation_enabled')) if 'reservation_enabled' in payload else resolve_bool(reservation_param)

        if enabled_value is not None:
            SettingsDB.set('registration_enabled', 'true' if enabled_value else 'false')
        if reservation_value is not None:
            SettingsDB.set('shop_reservation_enabled', 'true' if reservation_value else 'false')

        current_enabled = SettingsDB.get('registration_enabled', 'false').lower() == 'true'
        current_reservation = SettingsDB.get('shop_reservation_enabled', 'false') == 'true'

        return success_response("注册设置更新成功", {
            "enabled": current_enabled,
            "reservation_enabled": current_reservation
        })
    except Exception as e:
        logger.error(f"更新注册设置失败: {e}")
        return error_response("更新注册设置失败", 500)

@app.get("/admin/shop-settings")
async def get_shop_settings(request: Request):
    """获取商城设置"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        show_inactive = SettingsDB.get('show_inactive_in_shop', 'false') == 'true'
        return success_response("获取商城设置成功", {"show_inactive_in_shop": show_inactive})
    except Exception as e:
        logger.error(f"获取商城设置失败: {e}")
        return error_response("获取商城设置失败", 500)

@app.put("/admin/shop-settings")
async def update_shop_settings(request: Request):
    """更新商城设置"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        body = await request.json()
        show_inactive = body.get('show_inactive_in_shop', False)
        
        SettingsDB.set('show_inactive_in_shop', 'true' if show_inactive else 'false')
        return success_response("商城设置更新成功", {"show_inactive_in_shop": show_inactive})
    except Exception as e:
        logger.error(f"更新商城设置失败: {e}")
        return error_response("更新商城设置失败", 500)

# ==================== 学号搜索（管理员） ====================

@app.get("/admin/students/search")
async def admin_search_students(request: Request, q: str = "", limit: int = 20):
    """
    按学号、用户姓名、配送名模糊搜索
    - 管理员可以搜索所有用户
    - 代理只能搜索配送地址在其管辖区域内的用户
    """
    staff = get_current_staff_required_from_cookie(request)
    try:
        like = f"%{q.strip()}%" if q else "%"
        scope = build_staff_scope(staff)
        address_ids = [aid for aid in (scope.get('address_ids') or []) if aid]
        building_ids = [bid for bid in (scope.get('building_ids') or []) if bid]

        from database import get_db_connection
        with get_db_connection() as conn:
            cur = conn.cursor()
            # 支持同时搜索：学号(u.id)、用户姓名(u.name)、配送名(up.name)
            params: List[Any] = [like, like, like]
            search_condition = "(u.id LIKE ? OR u.name LIKE ? OR up.name LIKE ?)"
            filters: List[str] = [search_condition]

            if staff.get('type') == 'agent':
                # 代理只能搜索其管辖区域内有配送地址的用户
                if not address_ids and not building_ids:
                    return success_response("搜索成功", {"students": []})
                coverage_parts: List[str] = []
                if address_ids:
                    placeholders = ','.join('?' * len(address_ids))
                    coverage_parts.append(f"up.address_id IN ({placeholders})")
                    params.extend(address_ids)
                if building_ids:
                    placeholders = ','.join('?' * len(building_ids))
                    coverage_parts.append(f"up.building_id IN ({placeholders})")
                    params.extend(building_ids)
                filters.append('(' + ' OR '.join(coverage_parts) + ')')
                filters.append("((up.address_id IS NOT NULL AND TRIM(up.address_id) != '') OR (up.building_id IS NOT NULL AND TRIM(up.building_id) != ''))")

            query = f'''
                SELECT DISTINCT
                    u.id AS student_id,
                    u.name AS user_name,
                    up.name AS profile_name,
                    COALESCE(NULLIF(TRIM(u.name), ''), NULLIF(TRIM(up.name), ''), u.id) AS display_name
                FROM users u
                LEFT JOIN user_profiles up
                  ON (up.user_id = u.user_id OR (up.user_id IS NULL AND up.student_id = u.id))
                WHERE {' AND '.join(filters)}
                ORDER BY u.id ASC
                LIMIT ?
            '''
            params.append(max(1, min(limit, 50)))
            cur.execute(query, tuple(params))
            rows = cur.fetchall() or []

        # 返回时包含学号、用户姓名、配送名，便于前端展示
        items = []
        for row in rows:
            items.append({
                "id": row["student_id"],
                "name": row["display_name"],
                "user_name": row["user_name"],  # 用户姓名（注册时设置）
                "profile_name": row["profile_name"]  # 配送名
            })
        return success_response("搜索成功", {"students": items})
    except Exception as e:
        logger.error(f"搜索学号失败: {e}")
        return error_response("搜索失败", 500)

# ==================== 优惠券（用户/管理员） ====================

@app.get("/coupons/my")
async def my_coupons(request: Request):
    """
    获取当前用户可用的优惠券列表
    重要：只返回用户当前配送地址对应代理发放的优惠券
    - 如果用户在代理1的区域，只能看到代理1发放的优惠券
    - 如果用户切换到代理2的区域，则只能看到代理2发放的优惠券
    - 同一代理管辖的不同区域之间，优惠券互通
    """
    user = get_current_user_required_from_cookie(request)
    try:
        # 获取用户当前配送地址对应的代理
        scope = resolve_shopping_scope(request)
        owner_id = get_owner_id_from_scope(scope)
        
        # restrict_owner=True 确保只返回当前代理发放的优惠券
        coupons = CouponDB.get_active_for_student(
            user["id"],
            owner_id=owner_id,
            restrict_owner=True
        ) or []
        return success_response("获取优惠券成功", {"coupons": coupons})
    except Exception as e:
        logger.error(f"获取优惠券失败: {e}")
        return error_response("获取优惠券失败", 500)

@app.get("/admin/coupons")
async def admin_list_coupons(request: Request, student_id: Optional[str] = None, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        items = CouponDB.list_all(student_id, owner_id=owner_id)  # student_id参数名保持不变，但内部已支持user_id
        return success_response("获取优惠券列表成功", {"coupons": items})
    except Exception as e:
        logger.error(f"管理员获取优惠券失败: {e}")
        return error_response("获取优惠券失败", 500)


@app.get("/agent/coupons")
async def agent_list_coupons(request: Request, student_id: Optional[str] = None):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        items = CouponDB.list_all(student_id, owner_id=owner_id)  # student_id参数名保持不变，但内部已支持user_id
        return success_response("获取优惠券列表成功", {"coupons": items})
    except Exception as e:
        logger.error(f"代理获取优惠券失败: {e}")
        return error_response("获取优惠券失败", 500)

@app.post("/admin/coupons/issue")
async def admin_issue_coupons(payload: CouponIssueRequest, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        amt = float(payload.amount)
        if amt <= 0:
            return error_response("金额必须大于0", 400)
        qty = int(payload.quantity or 1)
        if qty <= 0 or qty > 200:
            return error_response("发放数量需为 1-200", 400)
        # 规范化时间格式
        expires_at = None
        if payload.expires_at:
            try:
                from datetime import datetime as _dt
                try:
                    dt = _dt.fromisoformat(payload.expires_at)
                except Exception:
                    dt = _dt.strptime(payload.expires_at, "%Y-%m-%d %H:%M:%S")
                expires_at = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return error_response("无效的过期时间格式", 400)
        ids = CouponDB.issue_coupons(payload.student_id, amt, qty, expires_at, owner_id=owner_id)  # student_id会自动解析为user_id
        if not ids:
            return error_response("发放失败，学号不存在或其他错误", 400)
        return success_response("发放成功", {"issued": len(ids), "coupon_ids": ids})
    except Exception as e:
        logger.error(f"发放优惠券失败: {e}")
        return error_response("发放优惠券失败", 500)


@app.post("/agent/coupons/issue")
async def agent_issue_coupons(payload: CouponIssueRequest, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        amt = float(payload.amount)
        if amt <= 0:
            return error_response("金额必须大于0", 400)
        qty = int(payload.quantity or 1)
        if qty <= 0 or qty > 200:
            return error_response("发放数量需为 1-200", 400)
        expires_at = None
        if payload.expires_at:
            try:
                from datetime import datetime as _dt
                try:
                    dt = _dt.fromisoformat(payload.expires_at)
                except Exception:
                    dt = _dt.strptime(payload.expires_at, "%Y-%m-%d %H:%M:%S")
                expires_at = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return error_response("无效的过期时间格式", 400)
        ids = CouponDB.issue_coupons(payload.student_id, amt, qty, expires_at, owner_id=owner_id)  # student_id会自动解析为user_id
        if not ids:
            return error_response("发放失败，学号不存在或其他错误", 400)
        return success_response("发放成功", {"issued": len(ids), "coupon_ids": ids})
    except Exception as e:
        logger.error(f"代理发放优惠券失败: {e}")
        return error_response("发放优惠券失败", 500)

@app.patch("/admin/coupons/{coupon_id}/revoke")
async def admin_revoke_coupon(coupon_id: str, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        ok = CouponDB.revoke(coupon_id, owner_id)
        if not ok:
            return error_response("撤回失败或已撤回/不存在", 400)
        return success_response("已撤回")
    except Exception as e:
        logger.error(f"撤回优惠券失败: {e}")
        return error_response("撤回失败", 500)


@app.patch("/agent/coupons/{coupon_id}/revoke")
async def agent_revoke_coupon(coupon_id: str, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        ok = CouponDB.revoke(coupon_id, owner_id)
        if not ok:
            return error_response("撤回失败或已撤回/不存在", 400)
        return success_response("已撤回")
    except Exception as e:
        logger.error(f"代理撤回优惠券失败: {e}")
        return error_response("撤回失败", 500)

@app.delete("/admin/coupons/{coupon_id}")
async def admin_delete_coupon(coupon_id: str, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        ok = CouponDB.permanently_delete_coupon(coupon_id, owner_id)
        if not ok:
            return error_response("删除失败，可能优惠券不存在或未撤回", 400)
        return success_response("已删除")
    except Exception as e:
        logger.error(f"删除优惠券失败: {e}")
        return error_response("删除失败", 500)

@app.delete("/agent/coupons/{coupon_id}")
async def agent_delete_coupon(coupon_id: str, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        ok = CouponDB.permanently_delete_coupon(coupon_id, owner_id)
        if not ok:
            return error_response("删除失败，可能优惠券不存在或未撤回", 400)
        return success_response("已删除")
    except Exception as e:
        logger.error(f"代理删除优惠券失败: {e}")
        return error_response("删除失败", 500)

# ==================== 商品路由 ====================

@app.get("/products")
async def get_products(request: Request, category: Optional[str] = None, address_id: Optional[str] = None, building_id: Optional[str] = None, hot_only: Optional[str] = None):
    """获取商品列表"""
    try:
        scope = resolve_shopping_scope(request, address_id, building_id)
        owner_ids = scope["owner_ids"]
        
        # 修复Agent商品权限控制：
        # 现在所有商品都有owner_id，所以统一使用owner_ids过滤，不再依赖include_unassigned
        include_unassigned = False

        # 检查是否在商城中显示下架商品
        show_inactive = SettingsDB.get('show_inactive_in_shop', 'false') == 'true'
        
        hot_filter = is_truthy(hot_only)
        if category:
            products = ProductDB.get_products_by_category(
                category,
                owner_ids=owner_ids,
                include_unassigned=include_unassigned,
                hot_only=hot_filter
            )
        else:
            products = ProductDB.get_all_products(
                owner_ids=owner_ids,
                include_unassigned=include_unassigned,
                hot_only=hot_filter
            )
        
        # 根据设置过滤下架商品
        if not show_inactive:
            products = [p for p in products if p.get('is_active', 1) != 0]
        
        # 补充规格信息
        product_ids = [p["id"] for p in products]
        variants_map = VariantDB.get_for_products(product_ids)
        for p in products:
            vts = variants_map.get(p["id"], [])
            p["variants"] = vts
            p["has_variants"] = len(vts) > 0
            p["is_not_for_sale"] = is_non_sellable(p)
            if p["has_variants"]:
                p["total_variant_stock"] = sum(v.get("stock", 0) for v in vts)
            if p["is_not_for_sale"]:
                p["stock_display"] = "∞"
        return success_response("获取商品列表成功", {"products": products, "scope": scope})
    
    except Exception as e:
        logger.error(f"获取商品失败: {e}")
        return error_response("获取商品失败", 500)

@app.get("/products/search")
async def search_products(request: Request, q: str, address_id: Optional[str] = None, building_id: Optional[str] = None):
    """搜索商品"""
    try:
        scope = resolve_shopping_scope(request, address_id, building_id)
        owner_ids = scope["owner_ids"]
        
        # 修复Agent商品权限控制：现在所有商品都有owner_id，统一使用owner_ids过滤
        include_unassigned = False
            
        # 检查是否在商城中显示下架商品
        show_inactive = SettingsDB.get('show_inactive_in_shop', 'false') == 'true'
        
        products = ProductDB.search_products(q, owner_ids=owner_ids, include_unassigned=include_unassigned)
        
        # 根据设置过滤下架商品
        if not show_inactive:
            products = [p for p in products if p.get('is_active', 1) != 0]
        
        product_ids = [p["id"] for p in products]
        variants_map = VariantDB.get_for_products(product_ids)
        for p in products:
            vts = variants_map.get(p["id"], [])
            p["variants"] = vts
            p["has_variants"] = len(vts) > 0
            p["is_not_for_sale"] = is_non_sellable(p)
            if p["has_variants"]:
                p["total_variant_stock"] = sum(v.get("stock", 0) for v in vts)
            if p["is_not_for_sale"]:
                p["stock_display"] = "∞"
        return success_response("搜索成功", {"products": products, "query": q, "scope": scope})
    
    except Exception as e:
        logger.error(f"搜索商品失败: {e}")
        return error_response("搜索失败", 500)

@app.get("/products/categories")
async def get_categories(request: Request, address_id: Optional[str] = None, building_id: Optional[str] = None):
    """获取商品分类（只返回有商品的分类）"""
    try:
        scope = resolve_shopping_scope(request, address_id, building_id)
        owner_ids = scope["owner_ids"]
        
        # 修复Agent商品权限控制：现在所有商品都有owner_id，统一使用owner_ids过滤
        include_unassigned = False
        
        # 检查是否在商城中显示下架商品
        show_inactive = SettingsDB.get('show_inactive_in_shop', 'false') == 'true'
            
        # 返回前自动清理空分类
        try:
            CategoryDB.cleanup_orphan_categories()
        except Exception:
            pass
            
        if show_inactive:
            # 显示下架商品时，正常获取分类
            categories = CategoryDB.get_categories_with_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        else:
            # 不显示下架商品时，只获取有上架商品的分类
            categories = CategoryDB.get_categories_with_active_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
            
        return success_response("获取分类成功", {"categories": categories, "scope": scope})
    
    except Exception as e:
        logger.error(f"获取分类失败: {e}")
        return error_response("获取分类失败", 500)

# ==================== 地址（宿舍区/自提点等）路由 ====================

@app.get("/addresses")
async def get_enabled_addresses():
    """获取启用且有启用楼栋的地址列表"""
    try:
        # 获取有启用楼栋的地址
        addrs = AddressDB.get_enabled_addresses_with_buildings()
        
        # 如果没有找到有楼栋的地址，记录日志但返回空列表
        if not addrs:
            # 获取所有启用的地址
            all_enabled_addrs = AddressDB.get_enabled_addresses()
            if all_enabled_addrs:
                # 有启用的地址但没有楼栋，不返回任何地址
                # 这样用户界面就不会显示没有楼栋的园区
                logger.info(f"发现 {len(all_enabled_addrs)} 个启用地址但无可用楼栋，不向用户显示")
            else:
                # 完全没有地址
                logger.info("没有找到任何启用的地址")
            addrs = []
            
        return success_response("获取地址成功", {"addresses": addrs})
    except Exception as e:
        logger.error(f"获取地址失败: {e}")
        return error_response("获取地址失败", 500)

@app.get("/admin/addresses")
async def admin_get_addresses(request: Request):
    """获取全部地址（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        addrs = AddressDB.get_all_addresses(include_disabled=True)
        # 若为空，则自动创建默认 桃园/六舍，便于管理员直接看到并管理
        if not addrs:
            try:
                addr_id = AddressDB.create_address("桃园", True, 0)
                if addr_id:
                    # 创建默认楼栋 六舍
                    try:
                        from database import BuildingDB
                        BuildingDB.create_building(addr_id, "六舍", True, 0)
                    except Exception:
                        pass
                    addrs = AddressDB.get_all_addresses(include_disabled=True)
            except Exception:
                pass
        return success_response("获取地址成功", {"addresses": addrs})
    except Exception as e:
        logger.error(f"管理员获取地址失败: {e}")
        return error_response("获取地址失败", 500)

@app.post("/admin/addresses")
async def admin_create_address(payload: AddressCreateRequest, request: Request):
    """创建地址（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        if AddressDB.get_by_name(payload.name):
            return error_response("地址名称已存在", 400)
        addr_id = AddressDB.create_address(payload.name, payload.enabled, payload.sort_order)
        if not addr_id:
            return error_response("创建地址失败，名称可能冲突", 400)
        return success_response("地址创建成功", {"address_id": addr_id})
    except Exception as e:
        logger.error(f"创建地址失败: {e}")
        return error_response("创建地址失败", 500)

@app.put("/admin/addresses/{address_id}")
async def admin_update_address(address_id: str, payload: AddressUpdateRequest, request: Request):
    """更新地址（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        existing = AddressDB.get_by_id(address_id)
        if not existing:
            return error_response("地址不存在", 404)
        # 名称冲突检查
        if payload.name and payload.name != existing.get("name"):
            if AddressDB.get_by_name(payload.name):
                return error_response("地址名称已存在", 400)
        agent_ids_to_expire: Optional[List[str]] = None
        if payload.enabled is not None:
            try:
                was_enabled = 1 if int(existing.get('enabled', 1) or 1) == 1 else 0
            except Exception:
                was_enabled = 1
            will_enable = 1 if payload.enabled else 0
            if was_enabled == 1 and will_enable == 0:
                agent_ids_to_expire = AgentAssignmentDB.get_agent_ids_for_address(address_id)
        ok = AddressDB.update_address(address_id, payload.name, payload.enabled, payload.sort_order)
        if not ok:
            return error_response("更新地址失败", 400)
        if agent_ids_to_expire:
            expire_agent_tokens_for_address(address_id, agent_ids_to_expire)
        return success_response("地址更新成功")
    except Exception as e:
        logger.error(f"更新地址失败: {e}")
        return error_response("更新地址失败", 500)

@app.delete("/admin/addresses/{address_id}")
async def admin_delete_address(address_id: str, request: Request):
    """删除地址（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        existing = AddressDB.get_by_id(address_id)
        if not existing:
            return error_response("地址不存在", 404)
        agent_ids_to_expire = AgentAssignmentDB.get_agent_ids_for_address(address_id)
        # 允许级联删除：先删除楼栋，再删除地址（由 AddressDB 实现）
        ok = AddressDB.delete_address(address_id)
        if not ok:
            return error_response("删除地址失败", 400)
        if agent_ids_to_expire:
            expire_agent_tokens_for_address(address_id, agent_ids_to_expire)
        return success_response("地址删除成功")
    except Exception as e:
        logger.error(f"删除地址失败: {e}")
        return error_response("删除地址失败", 500)

@app.post("/admin/addresses/reorder")
async def admin_reorder_addresses(payload: AddressReorderRequest, request: Request):
    """批量重排地址顺序（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        if not payload.order or not isinstance(payload.order, list):
            return error_response("无效排序数据", 400)
        ok = AddressDB.reorder(payload.order)
        if not ok:
            return error_response("重排失败", 400)
        return success_response("重排成功")
    except Exception as e:
        logger.error(f"地址重排失败: {e}")
        return error_response("地址重排失败", 500)

# ==================== 楼栋路由 ====================

@app.get("/buildings")
async def get_enabled_buildings(address_id: Optional[str] = None, address_name: Optional[str] = None):
    """根据地址获取启用的楼栋，若为空则回退默认“六舍”"""
    try:
        addr_id = address_id
        if not addr_id and address_name:
            addr = AddressDB.get_by_name(address_name)
            addr_id = addr.get('id') if addr else None

        buildings = []
        if addr_id:
            buildings = BuildingDB.get_enabled_buildings(addr_id)

        if not buildings:
            buildings = [{
                "id": "bld_default_6she",
                "address_id": addr_id or "addr_default_taoyuan",
                "name": "六舍",
                "enabled": 1,
                "sort_order": 0,
                "created_at": None,
                "updated_at": None
            }]
        return success_response("获取楼栋成功", {"buildings": buildings})
    except Exception as e:
        logger.error(f"获取楼栋失败: {e}")
        return error_response("获取楼栋失败", 500)

@app.get("/admin/buildings")
async def admin_get_buildings(request: Request, address_id: Optional[str] = None):
    """获取楼栋（可按地址过滤）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        blds = BuildingDB.get_all_buildings(address_id=address_id, include_disabled=True)
        return success_response("获取楼栋成功", {"buildings": blds})
    except Exception as e:
        logger.error(f"管理员获取楼栋失败: {e}")
        return error_response("获取楼栋失败", 500)

@app.post("/admin/buildings")
async def admin_create_building(payload: BuildingCreateRequest, request: Request):
    """创建楼栋（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        addr = AddressDB.get_by_id(payload.address_id)
        if not addr:
            return error_response("所属地址不存在", 400)
        if BuildingDB.get_by_name_in_address(payload.address_id, payload.name):
            return error_response("该地址下楼栋名称已存在", 400)
        bld_id = BuildingDB.create_building(payload.address_id, payload.name, payload.enabled, payload.sort_order)
        if not bld_id:
            return error_response("创建楼栋失败，名称冲突", 400)
        return success_response("楼栋创建成功", {"building_id": bld_id})
    except Exception as e:
        logger.error(f"创建楼栋失败: {e}")
        return error_response("创建楼栋失败", 500)

@app.put("/admin/buildings/{building_id}")
async def admin_update_building(building_id: str, payload: BuildingUpdateRequest, request: Request):
    """更新楼栋（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        existing = BuildingDB.get_by_id(building_id)
        if not existing:
            return error_response("楼栋不存在", 404)
        # 名称冲突校验（同地址下唯一）
        if payload.name and payload.name != existing.get('name'):
            if BuildingDB.get_by_name_in_address(existing.get('address_id'), payload.name):
                return error_response("该地址下楼栋名称已存在", 400)
        ok = BuildingDB.update_building(building_id, payload.name, payload.enabled, payload.sort_order)
        if not ok:
            return error_response("更新楼栋失败", 400)
        return success_response("楼栋更新成功")
    except Exception as e:
        logger.error(f"更新楼栋失败: {e}")
        return error_response("更新楼栋失败", 500)

@app.delete("/admin/buildings/{building_id}")
async def admin_delete_building(building_id: str, request: Request):
    """删除楼栋（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        existing = BuildingDB.get_by_id(building_id)
        if not existing:
            return error_response("楼栋不存在", 404)
        ok = BuildingDB.delete_building(building_id)
        if not ok:
            return error_response("删除楼栋失败", 400)
        return success_response("楼栋删除成功")
    except Exception as e:
        logger.error(f"删除楼栋失败: {e}")
        return error_response("删除楼栋失败", 500)

@app.post("/admin/buildings/reorder")
async def admin_reorder_buildings(payload: BuildingReorderRequest, request: Request):
    """对某地址下的楼栋批量重排（管理员）"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        if not payload.order or not isinstance(payload.order, list):
            return error_response("无效排序数据", 400)
        # 校验地址存在
        addr = AddressDB.get_by_id(payload.address_id)
        if not addr:
            return error_response("地址不存在", 404)
        ok = BuildingDB.reorder(payload.address_id, payload.order)
        if not ok:
            return error_response("重排失败", 400)
        return success_response("重排成功")
    except Exception as e:
        logger.error(f"楼栋重排失败: {e}")
        return error_response("楼栋重排失败", 500)

# ==================== 代理管理 ====================

@app.get("/admin/agents")
async def admin_list_agents(request: Request, include_inactive: bool = False):
    staff = get_current_super_admin_required_from_cookie(request)
    include_disabled = str(include_inactive).lower() in ("1", "true", "yes")
    include_deleted_param = request.query_params.get('include_deleted')
    include_deleted = include_disabled or (
        include_deleted_param is not None and str(include_deleted_param).lower() in ("1", "true", "yes")
    )
    try:
        agents = AdminDB.list_admins(
            role='agent',
            include_disabled=include_disabled,
            include_deleted=False
        )
        data = [serialize_agent_account(agent) for agent in agents]
        deleted_agents: List[Dict[str, Any]] = []
        if include_deleted:
            for record in AgentDeletionDB.list_active_records():
                agent_id = record.get('agent_id')
                # 检查该已删除代理是否有订单
                has_orders = False
                if agent_id:
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT COUNT(*) FROM orders WHERE agent_id = ?", (agent_id,))
                        order_count = cursor.fetchone()[0]
                        has_orders = order_count > 0
                
                # 只有有订单的已删除代理才添加到列表中
                if has_orders:
                    deleted_at_raw = record.get('deleted_at')
                    deleted_at_timestamp = None
                    if deleted_at_raw:
                        try:
                            deleted_at_timestamp = convert_sqlite_timestamp_to_unix(deleted_at_raw, agent_id)
                        except Exception as e:
                            logger.warning(f"转换删除时间失败 {agent_id}: {e}")
                    
                    deleted_agents.append({
                        "id": agent_id,
                        "name": record.get('agent_name') or agent_id,
                        "deleted_at": deleted_at_timestamp,
                        "address_ids": record.get('address_ids') or [],
                        "building_ids": record.get('building_ids') or [],
                        "is_deleted": True
                    })
        return success_response("获取代理列表成功", {"agents": data, "deleted_agents": deleted_agents})
    except Exception as e:
        logger.error(f"获取代理列表失败: {e}")
        return error_response("获取代理列表失败", 500)


@app.post("/admin/agents")
async def admin_create_agent(payload: AgentCreateRequest, request: Request):
    staff = get_current_super_admin_required_from_cookie(request)
    try:
        account = payload.account.strip()
        if not account:
            return error_response("账号不能为空", 400)
        if not payload.password or len(payload.password) < 3:
            return error_response("密码至少3位", 400)
        name = payload.name.strip() if payload.name else payload.account
        created = AdminDB.create_admin(account, payload.password, name, role='agent')
        if not created:
            return error_response("账号已存在", 400)

        valid_buildings, invalid_buildings = validate_building_ids(payload.building_ids)
        inherited_orders_count = 0
        if valid_buildings:
            AgentAssignmentDB.set_agent_buildings(account, valid_buildings)
            new_assignments = AgentAssignmentDB.get_buildings_for_agent(account)
            if new_assignments:
                # 继承已删除代理的订单数据
                address_ids = [item.get('address_id') for item in new_assignments]
                building_ids = [item.get('building_id') for item in new_assignments]
                inherited_orders_count = AgentDeletionDB.inherit_deleted_agent_orders(
                    address_ids, 
                    building_ids, 
                    account, 
                    name
                )
                if inherited_orders_count > 0:
                    logger.info(f"新代理 {account} 继承了 {inherited_orders_count} 个订单")
        else:
            new_assignments = []

        agent = AdminDB.get_admin(account, include_disabled=True, include_deleted=True)
        data = serialize_agent_account(agent)
        data['invalid_buildings'] = invalid_buildings
        data['inherited_orders_count'] = inherited_orders_count
        
        # 如果继承了数据，在消息中提示
        message = "代理创建成功"
        if inherited_orders_count > 0:
            message = f"代理创建成功，已自动继承相同区域已删除代理的所有数据（订单 {inherited_orders_count} 个及商品、配置、收款码等）"
        
        return success_response(message, {"agent": data})
    except Exception as e:
        logger.error(f"创建代理失败: {e}")
        return error_response("创建代理失败", 500)


@app.put("/admin/agents/{agent_id}")
async def admin_update_agent(agent_id: str, payload: AgentUpdateRequest, request: Request):
    staff = get_current_super_admin_required_from_cookie(request)
    try:
        agent = AdminDB.get_admin(agent_id, include_disabled=True, include_deleted=True)
        if not agent or (agent.get('role') or '').lower() != 'agent':
            return error_response("代理不存在", 404)
        if agent.get('deleted_at'):
            return error_response("该代理已删除，无法编辑", 400)

        def normalize_active(value: Any) -> bool:
            try:
                return int(value) == 1
            except Exception:
                return str(value).strip().lower() in ('1', 'true', 'yes', 'on')

        original_active = normalize_active(agent.get('is_active', 1))
        needs_token_reset = False

        update_fields: Dict[str, Any] = {}
        updated_name = agent.get('name')
        if payload.password:
            if len(payload.password) < 3:
                return error_response("密码至少3位", 400)
            update_fields['password'] = payload.password
            needs_token_reset = True
        if payload.name:
            updated_name = payload.name.strip()
            update_fields['name'] = updated_name
        if payload.is_active is not None:
            new_active = normalize_active(payload.is_active)
            update_fields['is_active'] = 1 if new_active else 0
            if new_active != original_active:
                needs_token_reset = True

        if update_fields:
            updated = AdminDB.update_admin(agent_id, **update_fields)
            if not updated:
                return error_response("更新代理信息失败", 400)

        invalid_buildings: List[str] = []
        inherited_orders_count = 0
        if payload.building_ids is not None:
            valid_buildings, invalid_buildings = validate_building_ids(payload.building_ids)
            assignments_ok = AgentAssignmentDB.set_agent_buildings(agent_id, valid_buildings)
            if not assignments_ok:
                return error_response("更新代理负责楼栋失败", 500)
            fresh_assignments = AgentAssignmentDB.get_buildings_for_agent(agent_id)
            if fresh_assignments:
                # 继承已删除代理的订单数据
                address_ids = [item.get('address_id') for item in fresh_assignments]
                building_ids = [item.get('building_id') for item in fresh_assignments]
                inherited_orders_count = AgentDeletionDB.inherit_deleted_agent_orders(
                    address_ids, 
                    building_ids, 
                    agent_id, 
                    updated_name or agent_id
                )
                if inherited_orders_count > 0:
                    logger.info(f"代理 {agent_id} 更新楼栋后继承了 {inherited_orders_count} 个订单")

        if needs_token_reset:
            AdminDB.bump_token_version(agent_id)

        refreshed = AdminDB.get_admin(agent_id, include_disabled=True, include_deleted=True)
        data = serialize_agent_account(refreshed)
        if payload.building_ids is not None:
            data['invalid_buildings'] = invalid_buildings
            data['inherited_orders_count'] = inherited_orders_count
        
        # 如果继承了数据，在消息中提示
        message = "代理更新成功"
        if inherited_orders_count > 0:
            message = f"代理更新成功，已自动继承相同区域已删除代理的所有数据（订单 {inherited_orders_count} 个及商品、配置、收款码等）"
        
        return success_response(message, {"agent": data})
    except Exception as e:
        logger.error(f"更新代理失败: {e}")
        return error_response("更新代理失败", 500)


@app.delete("/admin/agents/{agent_id}")
async def admin_delete_agent(agent_id: str, request: Request):
    staff = get_current_super_admin_required_from_cookie(request)
    try:
        if agent_id in AdminDB.SAFE_SUPER_ADMINS:
            return error_response("禁止删除系统管理员", 400)
        agent = AdminDB.get_admin(agent_id, include_disabled=True, include_deleted=True)
        if not agent or (agent.get('role') or '').lower() != 'agent':
            return error_response("代理不存在", 404)
        assignments = AgentAssignmentDB.get_buildings_for_agent(agent_id)
        deleted = AdminDB.soft_delete_admin(agent_id)
        if not deleted:
            return error_response("停用代理失败", 400)
        address_ids = [item.get('address_id') for item in assignments or []]
        building_ids = [item.get('building_id') for item in assignments or []]
        if not AgentDeletionDB.record_deletion(
            agent_id,
            agent.get('name') or agent_id,
            address_ids,
            building_ids
        ):
            logger.warning(f"记录代理删除信息失败: {agent_id}")
        if not AgentAssignmentDB.set_agent_buildings(agent_id, []):
            logger.warning(f"清空代理 {agent_id} 的楼栋关联失败")
        return success_response("代理已删除")
    except Exception as e:
        logger.error(f"删除代理失败: {e}")
        return error_response("删除代理失败", 500)


# ==================== 收款码管理路由 ====================

# 收款码数据模型
class PaymentQrCreateRequest(BaseModel):
    name: str

class PaymentQrUpdateRequest(BaseModel):
    name: Optional[str] = None

class PaymentQrStatusRequest(BaseModel):
    is_enabled: bool

@app.get("/admin/payment-qrs")
async def admin_get_payment_qrs(request: Request, owner_id: Optional[str] = None):
    """管理员获取收款码列表，支持切换归属"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        target_owner_id, normalized = resolve_single_owner_for_staff(staff, owner_id)
        owner_type = 'agent' if (normalized not in ('self', None) or staff.get('type') == 'agent') else 'admin'
        qrs = PaymentQrDB.get_payment_qrs(target_owner_id, owner_type, include_disabled=True)
        return success_response("获取收款码列表成功", {"payment_qrs": qrs})
    except Exception as e:
        logger.error(f"获取管理员收款码列表失败: {e}")
        return error_response("获取收款码列表失败", 500)

@app.get("/agent/payment-qrs")
async def agent_get_payment_qrs(request: Request):
    """代理获取自己的收款码列表"""
    staff = get_current_staff_required_from_cookie(request)
    if staff.get('role') != 'agent':
        return error_response("权限不足", 403)
    try:
        qrs = PaymentQrDB.get_payment_qrs(staff['id'], 'agent', include_disabled=True)
        return success_response("获取收款码列表成功", {"payment_qrs": qrs})
    except Exception as e:
        logger.error(f"获取代理收款码列表失败: {e}")
        return error_response("获取收款码列表失败", 500)

@app.post("/admin/payment-qrs")
async def admin_create_payment_qr(
    request: Request,
    name: str = Form(...),
    file: UploadFile = File(...),
    owner_id: Optional[str] = None
):
    """管理员创建收款码"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        if not file or not file.filename:
            return error_response("请上传图片文件", 400)

        target_owner_id, normalized = resolve_single_owner_for_staff(staff, owner_id)
        owner_type = 'agent' if (normalized not in ('self', None) or staff.get('type') == 'agent') else 'admin'

        # 验证文件类型
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_extensions:
            return error_response("只支持图片格式：jpg, jpeg, png, gif, webp", 400)
        
        # 生成文件名
        import time
        timestamp = int(time.time() * 1000)
        safe_name = re.sub(r'[^0-9A-Za-z\u4e00-\u9fa5_-]+', '_', name)
        filename = f"payment_qr_{target_owner_id}_{timestamp}_{safe_name}{ext}"
        target_path = os.path.join(public_dir, filename)

        # 保存文件
        content = await file.read()
        with open(target_path, 'wb') as f:
            f.write(content)

        web_path = f"/public/{filename}"
        
        # 创建收款码记录
        qr_id = PaymentQrDB.create_payment_qr(target_owner_id, owner_type, name, web_path)
        qr = PaymentQrDB.get_payment_qr(qr_id)
        
        return success_response("收款码创建成功", {"payment_qr": qr})
    except Exception as e:
        logger.error(f"创建管理员收款码失败: {e}")
        return error_response("创建收款码失败", 500)

@app.post("/agent/payment-qrs")
async def agent_create_payment_qr(
    request: Request,
    name: str = Form(...),
    file: UploadFile = File(...)
):
    """代理创建收款码"""
    staff = get_current_staff_required_from_cookie(request)
    if staff.get('role') != 'agent':
        return error_response("权限不足", 403)
    try:
        if not file or not file.filename:
            return error_response("请上传图片文件", 400)
        
        # 验证文件类型
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_extensions:
            return error_response("只支持图片格式：jpg, jpeg, png, gif, webp", 400)
        
        # 生成文件名
        import time
        timestamp = int(time.time() * 1000)
        safe_name = re.sub(r'[^0-9A-Za-z\u4e00-\u9fa5_-]+', '_', name)
        filename = f"payment_qr_{staff['id']}_{timestamp}_{safe_name}{ext}"
        target_path = os.path.join(public_dir, filename)
        
        # 保存文件
        content = await file.read()
        with open(target_path, 'wb') as f:
            f.write(content)
        
        web_path = f"/public/{filename}"
        
        # 创建收款码记录
        qr_id = PaymentQrDB.create_payment_qr(staff['id'], 'agent', name, web_path)
        qr = PaymentQrDB.get_payment_qr(qr_id)
        
        return success_response("收款码创建成功", {"payment_qr": qr})
    except Exception as e:
        logger.error(f"创建代理收款码失败: {e}")
        return error_response("创建收款码失败", 500)

@app.put("/admin/payment-qrs/{qr_id}")
async def admin_update_payment_qr(qr_id: str, payload: PaymentQrUpdateRequest, request: Request, owner_id: Optional[str] = None):
    """管理员更新收款码"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        target_owner_id, normalized = resolve_single_owner_for_staff(staff, owner_id)
        owner_type = 'agent' if (normalized not in ('self', None) or staff.get('type') == 'agent') else 'admin'
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != target_owner_id or qr['owner_type'] != owner_type:
            return error_response("收款码不存在或无权限", 404)
        
        # 更新
        if payload.name:
            PaymentQrDB.update_payment_qr(qr_id, name=payload.name)
        
        updated_qr = PaymentQrDB.get_payment_qr(qr_id)
        return success_response("收款码更新成功", {"payment_qr": updated_qr})
    except Exception as e:
        logger.error(f"更新管理员收款码失败: {e}")
        return error_response("更新收款码失败", 500)

@app.put("/agent/payment-qrs/{qr_id}")
async def agent_update_payment_qr(qr_id: str, payload: PaymentQrUpdateRequest, request: Request):
    """代理更新收款码"""
    staff = get_current_staff_required_from_cookie(request)
    if staff.get('role') != 'agent':
        return error_response("权限不足", 403)
    try:
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != staff['id'] or qr['owner_type'] != 'agent':
            return error_response("收款码不存在或无权限", 404)
        
        # 更新
        if payload.name:
            PaymentQrDB.update_payment_qr(qr_id, name=payload.name)
        
        updated_qr = PaymentQrDB.get_payment_qr(qr_id)
        return success_response("收款码更新成功", {"payment_qr": updated_qr})
    except Exception as e:
        logger.error(f"更新代理收款码失败: {e}")
        return error_response("更新收款码失败", 500)

@app.patch("/admin/payment-qrs/{qr_id}/status")
async def admin_update_payment_qr_status(qr_id: str, payload: PaymentQrStatusRequest, request: Request, owner_id: Optional[str] = None):
    """管理员更新收款码启用状态"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        target_owner_id, normalized = resolve_single_owner_for_staff(staff, owner_id)
        owner_type = 'agent' if (normalized not in ('self', None) or staff.get('type') == 'agent') else 'admin'
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != target_owner_id or qr['owner_type'] != owner_type:
            return error_response("收款码不存在或无权限", 404)
        
        # 如果要禁用，确保至少有一个启用的收款码
        if not payload.is_enabled:
            enabled_qrs = PaymentQrDB.get_enabled_payment_qrs(target_owner_id, owner_type)
            if len(enabled_qrs) <= 1 and qr['is_enabled'] == 1:
                return error_response("至少需要保留一个启用的收款码", 400)
        
        # 更新状态
        PaymentQrDB.update_payment_qr_status(qr_id, payload.is_enabled)
        updated_qr = PaymentQrDB.get_payment_qr(qr_id)
        return success_response("收款码状态更新成功", {"payment_qr": updated_qr})
    except Exception as e:
        logger.error(f"更新管理员收款码状态失败: {e}")
        return error_response("更新收款码状态失败", 500)

@app.patch("/agent/payment-qrs/{qr_id}/status")
async def agent_update_payment_qr_status(qr_id: str, payload: PaymentQrStatusRequest, request: Request):
    """代理更新收款码启用状态"""
    staff = get_current_staff_required_from_cookie(request)
    if staff.get('role') != 'agent':
        return error_response("权限不足", 403)
    try:
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != staff['id'] or qr['owner_type'] != 'agent':
            return error_response("收款码不存在或无权限", 404)
        
        # 如果要禁用，确保至少有一个启用的收款码
        if not payload.is_enabled:
            enabled_qrs = PaymentQrDB.get_enabled_payment_qrs(staff['id'], 'agent')
            if len(enabled_qrs) <= 1 and qr['is_enabled'] == 1:
                return error_response("至少需要保留一个启用的收款码", 400)
        
        # 更新状态
        PaymentQrDB.update_payment_qr_status(qr_id, payload.is_enabled)
        updated_qr = PaymentQrDB.get_payment_qr(qr_id)
        return success_response("收款码状态更新成功", {"payment_qr": updated_qr})
    except Exception as e:
        logger.error(f"更新代理收款码状态失败: {e}")
        return error_response("更新收款码状态失败", 500)

@app.delete("/admin/payment-qrs/{qr_id}")
async def admin_delete_payment_qr(qr_id: str, request: Request, owner_id: Optional[str] = None):
    """管理员删除收款码"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        target_owner_id, normalized = resolve_single_owner_for_staff(staff, owner_id)
        owner_type = 'agent' if (normalized not in ('self', None) or staff.get('type') == 'agent') else 'admin'
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != target_owner_id or qr['owner_type'] != owner_type:
            return error_response("收款码不存在或无权限", 404)
        
        # 注释掉最后一个收款码的限制，允许删除唯一收款码
        # all_qrs = PaymentQrDB.get_payment_qrs('admin', 'admin', include_disabled=True)
        # if len(all_qrs) <= 1:
        #     return error_response("至少需要保留一个收款码", 400)
        
        # 删除文件
        try:
            if qr['image_path'] and qr['image_path'].startswith('/'):
                file_path = os.path.join(public_dir, qr['image_path'][1:])
                if os.path.exists(file_path):
                    os.remove(file_path)
        except Exception as e:
            logger.warning(f"删除收款码文件失败: {e}")
        
        # 删除记录
        PaymentQrDB.delete_payment_qr(qr_id)
        
        # 如果删除的是启用的收款码，确保还有其他启用的
        PaymentQrDB.ensure_at_least_one_enabled(target_owner_id, owner_type)
        
        return success_response("收款码删除成功")
    except Exception as e:
        logger.error(f"删除管理员收款码失败: {e}")
        return error_response("删除收款码失败", 500)

@app.delete("/agent/payment-qrs/{qr_id}")
async def agent_delete_payment_qr(qr_id: str, request: Request):
    """代理删除收款码"""
    staff = get_current_staff_required_from_cookie(request)
    if staff.get('role') != 'agent':
        return error_response("权限不足", 403)
    try:
        # 验证收款码所有权
        qr = PaymentQrDB.get_payment_qr(qr_id)
        if not qr or qr['owner_id'] != staff['id'] or qr['owner_type'] != 'agent':
            return error_response("收款码不存在或无权限", 404)
        
        # 注释掉最后一个收款码的限制，允许删除唯一收款码
        # all_qrs = PaymentQrDB.get_payment_qrs(staff['id'], 'agent', include_disabled=True)
        # if len(all_qrs) <= 1:
        #     return error_response("至少需要保留一个收款码", 400)
        
        # 删除文件
        try:
            if qr['image_path'] and qr['image_path'].startswith('/'):
                file_path = os.path.join(public_dir, qr['image_path'][1:])
                if os.path.exists(file_path):
                    os.remove(file_path)
        except Exception as e:
            logger.warning(f"删除收款码文件失败: {e}")
        
        # 删除记录
        PaymentQrDB.delete_payment_qr(qr_id)
        
        # 如果删除的是启用的收款码，确保还有其他启用的
        PaymentQrDB.ensure_at_least_one_enabled(staff['id'], 'agent')
        
        return success_response("收款码删除成功")
    except Exception as e:
        logger.error(f"删除代理收款码失败: {e}")
        return error_response("删除收款码失败", 500)

# 获取支付收款码（不需要订单ID）
@app.get("/payment-qr")
async def get_payment_qr(address_id: str = None, building_id: str = None, request: Request = None):
    """根据地址信息获取对应的收款码"""
    user = get_current_user_from_cookie(request)
    if not user:
        return error_response("未登录", 401)

    try:
        query_address_id = address_id
        query_building_id = building_id
        if not query_address_id or not query_building_id:
            profile = UserProfileDB.get_shipping(user['id'])
            if profile:
                query_address_id = query_address_id or profile.get('address_id')
                query_building_id = query_building_id or profile.get('building_id')

        validation = check_address_and_building(query_address_id, query_building_id)
        if not validation.get('is_valid'):
            reason = validation.get('reason')
            if reason in ('missing_address', 'missing_building'):
                message = validation.get('message') or '请先选择配送地址'
            else:
                message = '地址不存在或未启用，请联系管理员'
            return error_response(message, 400)

        address_id = validation.get('address', {}).get('id') if validation.get('address') else query_address_id
        building_id = validation.get('building', {}).get('id') if validation.get('building') else query_building_id

        # 确定收款码所有者
        qr_owner_id = None
        qr_owner_type = None

        # 如果提供了地址信息，尝试查找对应的代理
        if building_id:
            from database import AgentAssignmentDB
            assignment_map = AgentAssignmentDB.get_assignment_map_for_buildings([building_id])
            agent_id = assignment_map.get(building_id)
            
            if agent_id:
                # 有代理，使用代理的收款码
                qr_owner_id = agent_id
                qr_owner_type = 'agent'
            else:
                # 没有代理，使用管理员收款码
                qr_owner_id = 'admin'
                qr_owner_type = 'admin'
        else:
            # 没有地址信息，使用管理员收款码
            qr_owner_id = 'admin'
            qr_owner_type = 'admin'
        
        # 获取随机启用的收款码
        qr = PaymentQrDB.get_random_enabled_qr(qr_owner_id, qr_owner_type)
        
        if not qr:
            # 如果没有找到收款码，尝试从旧系统获取
            if qr_owner_type == 'agent':
                agent = AdminDB.get_admin(qr_owner_id)
                if agent and agent.get('payment_qr_path'):
                    return success_response("获取收款码成功", {
                        "payment_qr": {
                            "image_path": agent['payment_qr_path'],
                            "name": f"{agent.get('name', qr_owner_id)}的收款码",
                            "owner_type": qr_owner_type
                        }
                    })
            
            # 如果还是没有，返回默认标识
            return success_response("获取收款码成功", {
                "payment_qr": {
                    "name": "无收款码",
                    "owner_type": "default"
                }
            })
        
        return success_response("获取收款码成功", {"payment_qr": qr})
        
    except Exception as e:
        logger.error(f"获取收款码失败: {e}")
        return error_response("获取收款码失败", 500)

# 获取订单收款码（保留兼容性）
@app.get("/orders/{order_id}/payment-qr")
async def get_order_payment_qr(order_id: str, request: Request):
    """获取订单对应的收款码"""
    user = get_current_user_from_cookie(request)
    if not user:
        return error_response("未登录", 401)
    
    try:
        # 获取订单信息
        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        
        # 验证订单所有权
        if order['student_id'] != user['id']:
            return error_response("无权限访问该订单", 403)
        
        # 转发到新的收款码API
        return await get_payment_qr(
            address_id=order.get('address_id'),
            building_id=order.get('building_id'),
            request=request
        )
        
    except Exception as e:
        logger.error(f"获取订单收款码失败: {e}")
        return error_response("获取收款码失败", 500)

# ==================== 购物车路由 ====================

@app.get("/cart")
async def get_cart(request: Request):
    """获取购物车"""
    # 验证用户登录状态
    user = get_current_user_required_from_cookie(request)

    try:
        logger.info(f"获取购物车请求 - 用户ID: {user['id']}, 用户信息: {user}")
        
        # 检查用户是否能被解析
        user_ref = UserDB.resolve_user_reference(user['id'])
        logger.info(f"获取购物车 - 用户解析结果: {user_ref}")

        scope = resolve_shopping_scope(request)
        owner_ids = scope["owner_ids"]
        owner_scope_id = get_owner_id_from_scope(scope)
        address_validation = scope.get('address_validation') or check_address_and_building(None, None)

        # 修复Agent商品权限控制：现在所有商品都有owner_id，统一使用owner_ids过滤
        include_unassigned = False

        cart_data = CartDB.get_cart(user["id"])
        if not cart_data:
            logger.info(f"用户 {user['id']} 没有购物车数据，返回空购物车")
            return success_response("获取购物车成功", {
                "items": [], 
                "total_quantity": 0, 
                "total_price": 0.0,
                "scope": scope,
                "lottery_threshold": LotteryConfigDB.get_threshold(owner_scope_id),
                "lottery_enabled": LotteryConfigDB.get_enabled(owner_scope_id),
                "address_validation": address_validation
            })

        # 获取购物车中的商品信息
        items_dict = cart_data["items"]
        logger.info(f"购物车原始数据: {items_dict}")

        # 获取所有商品信息
        all_products = ProductDB.get_all_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        product_dict = {p["id"]: p for p in all_products}
        
        # 构建前端需要的购物车商品列表
        cart_items = []
        total_quantity = 0  # 仅统计上架商品
        total_price = 0.0   # 仅统计上架商品（折扣后小计之和）
        
        SEP = '@@'
        for key, quantity in items_dict.items():
            product_id = key
            variant_id = None
            if isinstance(key, str) and SEP in key:
                product_id, variant_id = key.split(SEP, 1)
            if product_id in product_dict:
                product = product_dict[product_id]
                is_active = 1 if int(product.get("is_active", 1) or 1) == 1 else 0
                non_sellable = is_non_sellable(product)
                # 应用折扣（以折为单位，10为不打折）
                zhe = float(product.get("discount", 10.0) or 10.0)
                unit_price = round(float(product["price"]) * (zhe / 10.0), 2)
                subtotal = unit_price * quantity
                if non_sellable:
                    subtotal = 0.0

                # 仅将上架商品计入总数量与总价
                if is_active == 1:
                    total_quantity += quantity
                    if not non_sellable:
                        total_price += subtotal

                item = {
                    "product_id": product_id,
                    "name": product["name"],
                    "unit_price": round(unit_price, 2),
                    "quantity": quantity,
                    "subtotal": round(subtotal, 2),
                    "stock": product["stock"] if not non_sellable else "∞",
                    "category": product.get("category", ""),
                    "img_path": product.get("img_path", ""),
                    "is_active": is_active,
                    "is_not_for_sale": non_sellable
                }
                try:
                    requires_reservation = int(product.get("reservation_required", 0) or 0) == 1
                except Exception:
                    requires_reservation = bool(product.get("reservation_required"))
                if requires_reservation:
                    item["reservation_required"] = True
                    cutoff_val = product.get("reservation_cutoff")
                    if cutoff_val:
                        try:
                            item["reservation_cutoff"] = normalize_reservation_cutoff(str(cutoff_val))
                        except HTTPException:
                            item["reservation_cutoff"] = None
                    note_val = (product.get("reservation_note") or '').strip()
                    if note_val:
                        item["reservation_note"] = note_val[:120]
                if variant_id:
                    variant = VariantDB.get_by_id(variant_id)
                    if variant:
                        item["variant_id"] = variant_id
                        item["variant_name"] = variant.get("name")
                        item["stock"] = variant.get("stock", 0)
                if non_sellable:
                    item["stock"] = "∞"
                cart_items.append(item)
                
        logger.info(f"处理后的购物车数据 - 商品数: {len(cart_items)}, 总数量: {total_quantity}, 总价: {total_price}")
        
        # 获取配送费配置
        delivery_scope = scope
        owner_id = get_owner_id_from_scope(delivery_scope)
        delivery_config = DeliverySettingsDB.get_delivery_config(owner_id)
        active_cart_items = [item for item in cart_items if item.get('is_active', 1) == 1 and (item.get('quantity') or 0) > 0]
        has_reservation_items = any(item.get('reservation_required') for item in active_cart_items)
        all_items_reservation_required = bool(active_cart_items) and all(item.get('reservation_required') for item in active_cart_items)
        non_sellable_only = bool(active_cart_items) and all(item.get('is_not_for_sale') for item in active_cart_items)

        # 运费计算：购物车为空不收取，基础配送费或免配送费门槛任意一个为0则免费，否则达到门槛免费，否则收取基础配送费
        shipping_fee = 0.0 if total_quantity == 0 or delivery_config['delivery_fee'] == 0 or delivery_config['free_delivery_threshold'] == 0 or total_price >= delivery_config['free_delivery_threshold'] or non_sellable_only else delivery_config['delivery_fee']
        cart_result = {
            "items": cart_items,
            "total_quantity": total_quantity,
            "total_price": round(total_price, 2),
            "shipping_fee": round(shipping_fee, 2),
            "payable_total": round(total_price + shipping_fee, 2),
            "delivery_fee": delivery_config['delivery_fee'],
            "free_delivery_threshold": delivery_config['free_delivery_threshold'],
            "lottery_threshold": LotteryConfigDB.get_threshold(owner_scope_id),
            "lottery_enabled": LotteryConfigDB.get_enabled(owner_scope_id),
            "address_validation": address_validation,
            "has_reservation_items": has_reservation_items,
            "all_reservation_items": all_items_reservation_required
        }

        cart_result["scope"] = delivery_scope
        return success_response("获取购物车成功", cart_result)
    
    except Exception as e:
        logger.error(f"获取购物车失败: {e}")
        return error_response("获取购物车失败", 500)

@app.post("/cart/update")
async def update_cart(
    cart_request: CartUpdateRequest,
    request: Request
):
    """更新购物车"""
    # 验证用户登录状态
    user = get_current_user_required_from_cookie(request)
    
    try:
        # 添加详细的日志记录
        logger.info(f"购物车更新请求 - 用户ID: {user['id']}, 用户信息: {user}, 动作: {cart_request.action}, 商品ID: {cart_request.product_id}, 数量: {cart_request.quantity}")
        
        # 检查用户是否能被解析
        user_ref = UserDB.resolve_user_reference(user['id'])
        logger.info(f"用户解析结果: {user_ref}")

        scope = resolve_shopping_scope(request)
        owner_ids = scope["owner_ids"]
        include_unassigned = False if owner_ids else True
        logger.info(f"购物车更新 - 权限范围: {scope}")

        accessible_products = ProductDB.get_all_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        product_dict = {p["id"]: p for p in accessible_products}
        logger.info(f"购物车更新 - 可访问商品数量: {len(accessible_products)}")

        current_cart = CartDB.get_cart(user["id"])
        items = current_cart["items"] if current_cart else {}
        logger.info(f"当前购物车内容: {items}")

        if cart_request.action == "clear":
            items = {}
        elif cart_request.action == "remove" and cart_request.product_id:
            key = cart_request.product_id
            if cart_request.variant_id:
                key = f"{key}@@{cart_request.variant_id}"
            items.pop(key, None)
        elif cart_request.action in ["add", "update"] and cart_request.product_id and cart_request.quantity is not None:
            # 验证商品是否存在并获取商品信息
            product = product_dict.get(cart_request.product_id)
            logger.info(f"商品权限检查 - 商品ID: {cart_request.product_id}, 找到商品: {product is not None}")
            
            if product:
                logger.info(f"商品详情 - ID: {product.get('id')}, 名称: {product.get('name')}, 上架状态: {product.get('is_active')}, 拥有者: {product.get('owner_id')}")

            if not product:
                logger.error(f"商品无权访问或不存在: {cart_request.product_id}, 可访问商品列表: {list(product_dict.keys())[:10]}...")
                return error_response("商品不在当前地址的可售范围内", 403)

            # 禁止对下架商品进行添加或正数更新（允许通过数量<=0来移除）
            try:
                is_active = 1 if int(product.get("is_active", 1) or 1) == 1 else 0
            except Exception:
                is_active = 1
            if is_active != 1 and cart_request.quantity and cart_request.quantity > 0:
                return error_response("该商品已下架，无法添加或更新数量", 400)
            
            # 处理规格键与库存
            key = cart_request.product_id
            non_sellable = is_non_sellable(product)
            limit_stock = None if non_sellable else product["stock"]
            if cart_request.variant_id:
                key = f"{key}@@{cart_request.variant_id}"
                v = VariantDB.get_by_id(cart_request.variant_id)
                if not v or v.get('product_id') != cart_request.product_id:
                    return error_response("规格不存在", 400)
                limit_stock = None if non_sellable else int(v.get('stock', 0))

            if cart_request.action == "add":
                if cart_request.quantity <= 0:
                    logger.error(f"无效的数量: {cart_request.quantity}")
                    return error_response("数量必须大于0", 400)
                
                # 库存验证
                current_quantity = items.get(key, 0)
                new_quantity = current_quantity + cart_request.quantity
                if limit_stock is not None and new_quantity > limit_stock:
                    logger.error(f"库存不足 - 商品: {cart_request.product_id}, 规格: {cart_request.variant_id or '-'}, 当前购物车数量: {current_quantity}, 尝试添加: {cart_request.quantity}, 库存: {limit_stock}")
                    return error_response(f"库存不足，当前库存: {limit_stock}，购物车中已有: {current_quantity}", 400)
                items[key] = new_quantity
                logger.info(f"添加商品后的购物车: {items}")
            else:  # update
                if cart_request.quantity > 0:
                    # 更新时也需要验证库存
                    if limit_stock is not None and cart_request.quantity > limit_stock:
                        logger.error(f"更新数量超过库存 - 商品: {cart_request.product_id}, 规格: {cart_request.variant_id or '-'}, 尝试设置: {cart_request.quantity}, 库存: {limit_stock}")
                        return error_response(f"数量超过库存，最大可设置: {limit_stock}", 400)
                    items[key] = cart_request.quantity
                else:
                    items.pop(key, None)
        else:
            # 如果没有匹配任何条件，记录错误日志
            logger.error(f"购物车更新条件不匹配 - 动作: {cart_request.action}, 商品ID: {cart_request.product_id}, 数量: {cart_request.quantity}")
            return error_response("无效的购物车更新请求", 400)
        
        # 额外清理一次：移除购物车内的已下架商品，防止残留
        cleaned = {}
        for k, v in items.items():
            pid = k.split('@@', 1)[0] if isinstance(k, str) else k
            p = product_dict.get(pid)
            logger.info(f"清理检查 - 商品ID: {pid}, 数量: {v}, 商品信息: {p is not None}")
            
            if p is None:
                logger.warning(f"商品 {pid} 不在可访问商品列表中，将被过滤")
                continue
                
            try:
                active = 1 if int(p.get("is_active", 1) or 1) == 1 else 0
            except Exception:
                active = 1
            
            logger.info(f"商品 {pid} - 上架状态: {active}, 数量: {v}")
            
            if active == 1 and v > 0:
                cleaned[k] = v
            else:
                logger.info(f"商品 {pid} 被过滤 - 上架状态: {active}, 数量: {v}")

        logger.info(f"清理前购物车内容: {items}")
        logger.info(f"清理后购物车内容: {cleaned}")

        # 更新数据库
        update_result = CartDB.update_cart(user["id"], cleaned)
        logger.info(f"数据库更新结果: {update_result}, 保存到数据库的内容: {cleaned}")
        
        return success_response("购物车更新成功", {"action": cart_request.action, "items": cleaned, "scope": scope})
    
    except Exception as e:
        logger.error(f"更新购物车失败: {e}")
        return error_response("更新购物车失败", 500)

# ==================== 商店状态（打烊） ====================

@app.get("/shop/status")
async def get_shop_status():
    """获取店铺开关状态"""
    try:
        is_open = SettingsDB.get('shop_is_open', '1') != '0'
        note = SettingsDB.get('shop_closed_note', '')
        allow_reservation = SettingsDB.get('shop_reservation_enabled', 'false') == 'true'
        return success_response("获取店铺状态成功", {
            "is_open": is_open,
            "note": note,
            "allow_reservation": allow_reservation
        })
    except Exception as e:
        logger.error(f"获取店铺状态失败: {e}")
        return error_response("获取店铺状态失败", 500)

class ShopStatusUpdate(BaseModel):
    is_open: bool
    note: Optional[str] = None

@app.patch("/admin/shop/status")
async def update_shop_status(payload: ShopStatusUpdate, request: Request):
    """更新店铺开关（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        SettingsDB.set('shop_is_open', '1' if payload.is_open else '0')
        if payload.note is not None:
            SettingsDB.set('shop_closed_note', payload.note)
        return success_response("店铺状态已更新", {"is_open": payload.is_open})
    except Exception as e:
        logger.error(f"更新店铺状态失败: {e}")
        return error_response("更新店铺状态失败", 500)

# ==================== 代理独立状态管理 ====================

class AgentStatusUpdateRequest(BaseModel):
    is_open: bool
    closed_note: Optional[str] = ''
    allow_reservation: Optional[bool] = False

@app.get("/agent/status")
async def get_agent_status(request: Request):
    """获取代理的营业状态"""
    agent = get_current_agent_from_cookie(request)
    if not agent:
        return error_response("需要代理权限", 403)
    
    try:
        status = AgentStatusDB.get_agent_status(agent['id'])
        return success_response("获取代理状态成功", {
            "is_open": bool(status.get('is_open', 1)),
            "closed_note": status.get('closed_note', ''),
            "allow_reservation": bool(status.get('allow_reservation', 0))
        })
    except Exception as e:
        logger.error(f"获取代理状态失败: {e}")
        return error_response("获取代理状态失败", 500)

@app.patch("/agent/status")
async def update_agent_status(payload: AgentStatusUpdateRequest, request: Request):
    """更新代理的营业状态"""
    agent = get_current_agent_from_cookie(request)
    if not agent:
        return error_response("需要代理权限", 403)
    
    try:
        success = AgentStatusDB.update_agent_status(
            agent['id'], 
            payload.is_open, 
            payload.closed_note or '',
            bool(payload.allow_reservation)
        )
        if success:
            return success_response("代理状态已更新", {
                "is_open": payload.is_open,
                "closed_note": payload.closed_note or '',
                "allow_reservation": bool(payload.allow_reservation)
            })
        else:
            return error_response("更新代理状态失败", 500)
    except Exception as e:
        logger.error(f"更新代理状态失败: {e}")
        return error_response("更新代理状态失败", 500)

@app.get("/shop/agent-status")
async def get_user_agent_status(
    request: Request,
    address_id: Optional[str] = None, 
    building_id: Optional[str] = None
):
    """获取用户所属代理的营业状态"""
    try:
        # 使用现有的 resolve_shopping_scope 函数获取用户所属代理
        scope = resolve_shopping_scope(request, address_id, building_id)
        agent_id = scope.get('agent_id')
        
        if not agent_id:
            # 没有对应代理，返回全局店铺状态
            is_open = SettingsDB.get('shop_is_open', '1') != '0'
            note = SettingsDB.get('shop_closed_note', '')
            allow_reservation = SettingsDB.get('shop_reservation_enabled', 'false') == 'true'
            return success_response("获取店铺状态成功", {
                "is_open": is_open,
                "note": note,
                "is_agent": False,
                "allow_reservation": allow_reservation
            })

        # 有对应代理，返回代理状态
        status = AgentStatusDB.get_agent_status(agent_id)
        return success_response("获取代理状态成功", {
            "is_open": bool(status.get('is_open', 1)),
            "note": status.get('closed_note', ''),
            "is_agent": True,
            "agent_id": agent_id,
            "allow_reservation": bool(status.get('allow_reservation', 0))
        })
    except Exception as e:
        logger.error(f"获取用户代理状态失败: {e}")
        return error_response("获取状态失败", 500)

# ==================== 规格管理（管理员） ====================

class VariantCreate(BaseModel):
    name: str
    stock: int

class VariantUpdate(BaseModel):
    name: Optional[str] = None
    stock: Optional[int] = None

@app.get("/admin/products/{product_id}/variants")
async def list_variants(product_id: str, request: Request):
    staff = get_current_staff_required_from_cookie(request)
    try:
        product = ProductDB.get_product_by_id(product_id)
        if not product:
            return error_response("商品不存在", 404)
        if not staff_can_access_product(staff, product):
            return error_response("无权访问该商品", 403)
        return success_response("获取规格成功", {"variants": VariantDB.get_by_product(product_id)})
    except Exception as e:
        logger.error(f"获取规格失败: {e}")
        return error_response("获取规格失败", 500)


@app.get("/agent/products/{product_id}/variants")
async def agent_list_variants(product_id: str, request: Request):
    require_agent_with_scope(request)
    return await list_variants(product_id, request)


@app.post("/admin/products/{product_id}/variants")
async def create_variant(product_id: str, payload: VariantCreate, request: Request):
    staff = get_current_staff_required_from_cookie(request)
    try:
        product = ProductDB.get_product_by_id(product_id)
        if not product:
            return error_response("商品不存在", 404)
        if not staff_can_access_product(staff, product):
            return error_response("无权访问该商品", 403)
        vid = VariantDB.create_variant(product_id, payload.name, payload.stock)
        return success_response("规格创建成功", {"variant_id": vid})
    except Exception as e:
        logger.error(f"规格创建失败: {e}")
        return error_response("规格创建失败", 500)


@app.post("/agent/products/{product_id}/variants")
async def agent_create_variant(product_id: str, payload: VariantCreate, request: Request):
    require_agent_with_scope(request)
    return await create_variant(product_id, payload, request)


@app.put("/admin/variants/{variant_id}")
async def update_variant(variant_id: str, payload: VariantUpdate, request: Request):
    staff = get_current_staff_required_from_cookie(request)
    try:
        variant = VariantDB.get_by_id(variant_id)
        if not variant:
            return error_response("规格不存在", 404)
        product = ProductDB.get_product_by_id(variant.get('product_id'))
        if not staff_can_access_product(staff, product):
            return error_response("无权访问该商品", 403)
        ok = VariantDB.update_variant(variant_id, payload.name, payload.stock)
        if not ok:
            return error_response("无有效更新项", 400)
        return success_response("规格已更新")
    except Exception as e:
        logger.error(f"规格更新失败: {e}")
        return error_response("规格更新失败", 500)


@app.put("/agent/variants/{variant_id}")
async def agent_update_variant(variant_id: str, payload: VariantUpdate, request: Request):
    require_agent_with_scope(request)
    return await update_variant(variant_id, payload, request)


@app.delete("/admin/variants/{variant_id}")
async def delete_variant(variant_id: str, request: Request):
    staff = get_current_staff_required_from_cookie(request)
    try:
        variant = VariantDB.get_by_id(variant_id)
        if not variant:
            return error_response("规格不存在", 404)
        product = ProductDB.get_product_by_id(variant.get('product_id'))
        if not staff_can_access_product(staff, product):
            return error_response("无权访问该商品", 403)
        ok = VariantDB.delete_variant(variant_id)
        if not ok:
            return error_response("规格不存在", 404)
        return success_response("规格已删除")
    except Exception as e:
        logger.error(f"规格删除失败: {e}")
        return error_response("规格删除失败", 500)


@app.delete("/agent/variants/{variant_id}")
async def agent_delete_variant(variant_id: str, request: Request):
    require_agent_with_scope(request)
    return await delete_variant(variant_id, request)


# ==================== 管理员路由 ====================

@app.post("/admin/products")
async def create_product(
    request: Request,
    name: str = Form(...),
    category: str = Form(...),
    price: float = Form(...),
    stock: int = Form(0),
    description: str = Form(""),
    cost: float = Form(0.0),
    owner_id: Optional[str] = Form(None),
    discount: Optional[str] = Form(None),
    variants: Optional[str] = Form(None),
    is_hot: Optional[str] = Form(None),
    is_not_for_sale: Optional[str] = Form(None),
    reservation_required: Optional[str] = Form(None),
    reservation_cutoff: Optional[str] = Form(None),
    reservation_note: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None)
):
    """管理员创建商品"""
    # 验证管理员权限
    staff = get_current_staff_required_from_cookie(request)
    return await handle_product_creation(
        staff,
        name=name,
        category=category,
        price=price,
        stock=stock,
        description=description,
        cost=cost,
        owner_id=owner_id,
        discount=discount,
        variants=variants,
        image=image,
        is_hot=is_truthy(is_hot),
        is_not_for_sale=is_truthy(is_not_for_sale),
        reservation_required=is_truthy(reservation_required) if reservation_required is not None else False,
        reservation_cutoff=reservation_cutoff,
        reservation_note=reservation_note
    )


@app.post("/agent/products")
async def agent_create_product(
    request: Request,
    name: str = Form(...),
    category: str = Form(...),
    price: float = Form(...),
    stock: int = Form(0),
    description: str = Form(""),
    cost: float = Form(0.0),
    discount: Optional[str] = Form(None),
    variants: Optional[str] = Form(None),
    is_hot: Optional[str] = Form(None),
    is_not_for_sale: Optional[str] = Form(None),
    reservation_required: Optional[str] = Form(None),
    reservation_cutoff: Optional[str] = Form(None),
    reservation_note: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None)
):
    agent, _ = require_agent_with_scope(request)
    # Agent创建的商品应该明确设置为该Agent的ID作为owner_id
    agent_owner_id = get_owner_id_for_staff(agent)
    return await handle_product_creation(
        agent,
        name=name,
        category=category,
        price=price,
        stock=stock,
        description=description,
        cost=cost,
        owner_id=agent_owner_id,
        discount=discount,
        variants=variants,
        image=image,
        is_hot=is_truthy(is_hot),
        is_not_for_sale=is_truthy(is_not_for_sale),
        reservation_required=is_truthy(reservation_required) if reservation_required is not None else False,
        reservation_cutoff=reservation_cutoff,
        reservation_note=reservation_note
    )


@app.get("/admin/products")
async def admin_list_products(
    request: Request,
    q: Optional[str] = None,
    category: Optional[str] = None,
    include_inactive: Optional[bool] = True,
    owner_id: Optional[str] = None
):
    staff = get_current_staff_required_from_cookie(request)

    query = q.strip() if isinstance(q, str) and q.strip() else None
    category_filter = category.strip() if isinstance(category, str) and category.strip() else None

    if include_inactive is None:
        include_inactive_flag = True
    elif isinstance(include_inactive, str):
        include_inactive_flag = include_inactive.strip().lower() not in ('false', '0', 'no')
    else:
        include_inactive_flag = bool(include_inactive)

    scope = build_staff_scope(staff)
    owner_ids, include_unassigned, _ = resolve_owner_filter_for_staff(staff, scope, owner_id)
    scope_override = dict(scope)
    scope_override['owner_ids'] = owner_ids
    scope_override['is_super_admin'] = include_unassigned

    data = build_product_listing_for_staff(
        staff,
        scope_override,
        query=query,
        category=category_filter,
        include_inactive=include_inactive_flag
    )
    return success_response("获取商品列表成功", data)


@app.get("/agent/categories")
async def agent_get_categories(request: Request):
    """获取代理的所有分类"""
    try:
        agent, scope = require_agent_with_scope(request)
        owner_ids = scope.get('owner_ids')
        
        # 代理不需要include_unassigned，因为所有商品都有owner_id
        include_unassigned = False
        
        # 返回前自动清理空分类
        try:
            CategoryDB.cleanup_orphan_categories()
        except Exception:
            pass
            
        # 获取所有有商品的分类（包括下架商品）
        categories = CategoryDB.get_categories_with_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        
        return success_response("获取分类成功", {"categories": categories})
    
    except Exception as e:
        logger.error(f"获取代理分类失败: {e}")
        return error_response("获取分类失败", 500)

@app.get("/agent/products")
async def agent_list_products(
    request: Request,
    q: Optional[str] = None,
    category: Optional[str] = None,
    include_inactive: bool = True
):
    agent, scope = require_agent_with_scope(request)
    query = q.strip() if isinstance(q, str) else None
    category_filter = category.strip() if isinstance(category, str) and category.strip() else None
    data = build_product_listing_for_staff(
        agent,
        scope,
        query=query,
        category=category_filter,
        include_inactive=include_inactive
    )
    return success_response("获取商品列表成功", data)

@app.get("/admin/stats")
async def get_admin_stats(request: Request, owner_id: Optional[str] = None):
    """获取管理统计信息"""
    # 验证管理员权限
    staff = get_current_staff_required_from_cookie(request)

    try:
        scope = build_staff_scope(staff)
        owner_ids, include_unassigned, normalized_filter = resolve_owner_filter_for_staff(staff, scope, owner_id)

        # 复用仪表盘数据获取总订单/总销售额/净利润
        dashboard_summary = OrderDB.get_dashboard_stats(
            period='week',
            agent_id=scope.get('agent_id') if normalized_filter != 'admin' else None,
            address_ids=scope.get('address_ids'),
            building_ids=scope.get('building_ids'),
            filter_admin_orders=scope.get('filter_admin_orders', False)
        )

        products = ProductDB.get_all_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        try:
            CategoryDB.cleanup_orphan_categories()
        except Exception:
            pass
        categories = CategoryDB.get_categories_with_products(
            owner_ids=owner_ids,
            include_unassigned=include_unassigned
        )
        # 注册人数：管理员概览需要所有注册用户数量，不受区域过滤
        users_count = compute_registered_user_count(None)

        for p in products:
            p['is_not_for_sale'] = is_non_sellable(p)

        total_stock = 0
        for p in products:
            try:
                if is_non_sellable(p):
                    continue
                total_stock += max(int(p.get('stock', 0) or 0), 0)
            except Exception:
                continue

        stats = {
            "total_products": len(products),
            "categories": len(categories),
            "total_stock": total_stock,
            "recent_products": products[:5],  # 最近5个商品
            "users_count": users_count,
            "total_orders": dashboard_summary.get("total_orders", 0),
            "total_revenue": dashboard_summary.get("total_revenue", 0.0),
            "total_profit": dashboard_summary.get("profit_stats", {}).get("total_profit", 0.0),
            "scope": scope,
            "owner_filter": normalized_filter
        }

        return success_response("获取统计信息成功", stats)
    
    except Exception as e:
        logger.error(f"获取统计信息失败: {e}")
        return error_response("获取统计信息失败", 500)

@app.get("/admin/users/count")
async def get_users_count(request: Request, owner_id: Optional[str] = None, agent_id: Optional[str] = None):
    """根据当前工作人员权限范围获取注册人数
    - 如果提供 agent_id 参数，使用订单范围逻辑（支持排除代理地址/楼栋）
    - 如果只提供 owner_id 参数，使用商品范围逻辑（用于商品管理页面）
    """
    staff = get_current_staff_required_from_cookie(request)
    try:
        scope = build_staff_scope(staff)
        
        # 如果提供了 agent_id，使用订单范围逻辑
        if agent_id is not None:
            (
                selected_agent_id,
                selected_address_ids,
                selected_building_ids,
                exclude_address_ids,
                exclude_building_ids,
                normalized_filter
            ) = resolve_staff_order_scope(staff, scope, agent_id)
            
            # 根据订单范围统计用户数
            count = UserProfileDB.count_users_by_scope(
                agent_id=selected_agent_id,
                address_ids=selected_address_ids,
                building_ids=selected_building_ids,
                exclude_address_ids=exclude_address_ids,
                exclude_building_ids=exclude_building_ids
            )
            
            return success_response("获取注册人数成功", {
                "count": count,
                "agent_filter": normalized_filter
            })
        else:
            # 使用原来的商品范围逻辑
            owner_ids, _, normalized_filter = resolve_owner_filter_for_staff(staff, scope, owner_id)
            count = compute_registered_user_count(owner_ids)
            return success_response("获取注册人数成功", {
                "count": count,
                "owner_filter": normalized_filter
            })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取注册人数失败: {e}")
        return error_response("获取注册人数失败", 500)

@app.get("/admin/products/{product_id}")
async def get_product_details(product_id: str, request: Request):
    """获取商品详情"""
    # 验证管理员权限
    staff = get_current_staff_required_from_cookie(request)

    try:
        product = ProductDB.get_product_by_id(product_id)
        if not product:
            return error_response("商品不存在", 404)
        if not staff_can_access_product(staff, product):
            return error_response("无权访问该商品", 403)

        # 补充规格信息（与 /products 路由保持一致）
        variants = VariantDB.get_by_product(product_id)
        product["variants"] = variants
        product["has_variants"] = len(variants) > 0
        if product["has_variants"]:
            product["total_variant_stock"] = sum(v.get("stock", 0) for v in variants)

        return success_response("获取商品详情成功", {"product": product})
    
    except Exception as e:
        logger.error(f"获取商品详情失败: {e}")
        return error_response("获取商品详情失败", 500)


@app.get("/agent/products/{product_id}")
async def agent_get_product_details(product_id: str, request: Request):
    require_agent_with_scope(request)
    return await get_product_details(product_id, request)

@app.put("/admin/products/{product_id}")
async def update_product(
    product_id: str,
    product_data: "ProductUpdateRequest",
    request: Request
):
    """更新商品信息"""
    staff = get_current_staff_required_from_cookie(request)
    return await handle_product_update(staff, product_id, product_data)


@app.put("/agent/products/{product_id}")
async def agent_update_product(
    product_id: str,
    product_data: "ProductUpdateRequest",
    request: Request
):
    agent, _ = require_agent_with_scope(request)
    return await handle_product_update(agent, product_id, product_data)

@app.put("/admin/products/0")
async def bulk_update_products(payload: BulkProductUpdateRequest, request: Request):
    """批量更新商品（目前支持批量折扣）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        if not payload.product_ids:
            return error_response("未提供商品ID", 400)

        update_fields: Dict[str, Any] = {}
        # 仅支持折扣批量更新
        if payload.discount is not None:
            try:
                d = float(payload.discount)
                if d < 0.5 or d > 10:
                    return error_response("折扣范围应为0.5~10折", 400)
                update_fields['discount'] = d
            except Exception:
                return error_response("无效的折扣", 400)

        if not update_fields:
            return error_response("没有可更新的字段", 400)

        updated = 0
        not_found: List[str] = []
        for pid in payload.product_ids:
            p = ProductDB.get_product_by_id(pid)
            if not p:
                not_found.append(pid)
                continue
            ok = ProductDB.update_product(pid, update_fields)
            if ok:
                updated += 1
        return success_response("批量更新完成", {"updated": updated, "not_found": not_found})
    except Exception as e:
        logger.error(f"批量更新商品失败: {e}")
        return error_response("批量更新商品失败", 500)

# 兼容：支持直接 PUT /admin/products 进行批量更新（避免某些环境对路径"/0"的特殊处理）
@app.put("/admin/products")
async def bulk_update_products_alt(payload: BulkProductUpdateRequest, request: Request):
    return await bulk_update_products(payload, request)

@app.patch("/admin/products/{product_id}/stock")
async def update_product_stock(
    product_id: str,
    stock_data: "StockUpdateRequest",
    request: Request
):
    """更新商品库存"""
    staff = get_current_staff_required_from_cookie(request)
    return await handle_product_stock_update(staff, product_id, stock_data)


@app.patch("/agent/products/{product_id}/stock")
async def agent_update_product_stock(
    product_id: str,
    stock_data: "StockUpdateRequest",
    request: Request
):
    agent, _ = require_agent_with_scope(request)
    return await handle_product_stock_update(agent, product_id, stock_data)

@app.delete("/admin/products/{product_id}")
async def delete_products(
    product_id: str, 
    request: Request,
    delete_request: Optional[ProductDeleteRequest] = None
):
    """删除商品（支持单个或批量）"""
    staff = get_current_staff_required_from_cookie(request)

    try:
        if delete_request and delete_request.product_ids:
            product_ids = delete_request.product_ids
            if len(product_ids) > 100:
                return error_response("批量删除数量不能超过100件商品", 400)

            allowed_ids: List[str] = []
            for pid in product_ids:
                product = ProductDB.get_product_by_id(pid)
                if product and staff_can_access_product(staff, product):
                    allowed_ids.append(pid)

            if not allowed_ids:
                return error_response("无权删除指定商品", 403)

            logger.info(f"工作人员 {staff['id']} 请求批量删除商品: {allowed_ids}")
            result = ProductDB.batch_delete_products(allowed_ids)

            if not result.get("success"):
                logger.error(f"批量删除失败: {result.get('message')}")
                return error_response(result.get("message", "批量删除失败"), 400)

            try:
                for pid in result.get("deleted_ids", []) or []:
                    try:
                        removed = CartDB.remove_product_from_all_carts(pid)
                        logger.info(f"商品 {pid} 批量删除后，已从 {removed} 个购物车中移除")
                    except Exception as er:
                        logger.warning(f"批量删除后移除购物车商品失败 {pid}: {er}")
            except Exception as e:
                logger.warning(f"批量移除购物车商品异常: {e}")

            deleted_img_paths = result.get("deleted_img_paths", [])
            for img_path in deleted_img_paths:
                try:
                    img_file_path = os.path.join(os.path.dirname(__file__), img_path)
                    if os.path.exists(img_file_path):
                        os.remove(img_file_path)
                        logger.info(f"成功删除商品图片: {img_file_path}")
                except Exception as e:
                    logger.warning(f"删除商品图片失败 {img_path}: {e}")

            return success_response(result.get("message", "批量删除商品成功"), {
                "deleted_count": result.get("deleted_count", 0),
                "deleted_ids": result.get("deleted_ids", []),
                "not_found_ids": result.get("not_found_ids", [])
            })

        # 单个删除
        existing_product = ProductDB.get_product_by_id(product_id)
        if not existing_product:
            return error_response("商品不存在", 404)
        if not staff_can_access_product(staff, existing_product):
            return error_response("无权删除该商品", 403)

        img_path = existing_product.get("img_path", "")
        success = ProductDB.delete_product(product_id)
        if not success:
            return error_response("删除商品失败", 500)

        try:
            removed = CartDB.remove_product_from_all_carts(product_id)
            logger.info(f"商品 {product_id} 删除后，已从 {removed} 个购物车中移除")
        except Exception as e:
            logger.warning(f"删除后移除购物车商品失败: {e}")

        if img_path and img_path.strip():
            try:
                img_file_path = os.path.join(os.path.dirname(__file__), img_path)
                if os.path.exists(img_file_path):
                    os.remove(img_file_path)
                    logger.info(f"成功删除商品图片: {img_file_path}")
            except Exception as e:
                logger.warning(f"删除商品图片失败 {img_path}: {e}")

        return success_response("商品删除成功")

    except Exception as e:
        logger.error(f"删除商品失败: {e}")
        return error_response("删除商品失败", 500)


@app.delete("/agent/products/{product_id}")
async def agent_delete_products(
    product_id: str,
    request: Request,
    delete_request: Optional[ProductDeleteRequest] = None
):
    require_agent_with_scope(request)
    return await delete_products(product_id, request, delete_request)

@app.post("/admin/products/{product_id}/image")
async def update_product_image(
    product_id: str,
    request: Request,
    image: Optional[UploadFile] = File(None)
):
    """更新商品图片（仅图片）"""
    staff = get_current_staff_required_from_cookie(request)
    return await handle_product_image_update(staff, product_id, image)


@app.post("/agent/products/{product_id}/image")
async def agent_update_product_image(
    product_id: str,
    request: Request,
    image: Optional[UploadFile] = File(None)
):
    agent, _ = require_agent_with_scope(request)
    return await handle_product_image_update(agent, product_id, image)

# ==================== 分类管理路由 ====================

@app.get("/admin/categories")
async def get_admin_categories(request: Request, owner_id: Optional[str] = None):
    """获取所有分类（管理员）"""
    # 验证管理员权限
    admin = get_current_admin_required_from_cookie(request)
    
    try:
        scope = build_staff_scope(admin)
        owner_ids, include_unassigned, _ = resolve_owner_filter_for_staff(admin, scope, owner_id)

        try:
            CategoryDB.cleanup_orphan_categories()
        except Exception:
            pass
        categories = CategoryDB.get_categories_with_products(
            owner_ids=owner_ids,
            include_unassigned=include_unassigned
        )
        return success_response("获取分类成功", {"categories": categories})
    
    except Exception as e:
        logger.error(f"获取分类失败: {e}")
        return error_response("获取分类失败", 500)

@app.post("/admin/categories")
async def create_category(
    category_data: CategoryCreateRequest,
    request: Request
):
    """创建新分类"""
    # 验证管理员权限
    admin = get_current_admin_required_from_cookie(request)
    
    try:
        # 检查分类名称是否已存在
        existing_category = CategoryDB.get_category_by_name(category_data.name)
        if existing_category:
            return error_response("分类名称已存在", 400)
        
        category_id = CategoryDB.create_category(category_data.name, category_data.description)
        if not category_id:
            return error_response("创建分类失败", 500)
        
        return success_response("分类创建成功", {"category_id": category_id})
    
    except Exception as e:
        logger.error(f"创建分类失败: {e}")
        return error_response("创建分类失败", 500)

@app.put("/admin/categories/{category_id}")
async def update_category(
    category_id: str,
    category_data: CategoryUpdateRequest,
    request: Request
):
    """更新分类"""
    # 验证管理员权限
    admin = get_current_admin_required_from_cookie(request)
    
    try:
        # 检查分类是否存在
        existing_category = CategoryDB.get_category_by_id(category_id)
        if not existing_category:
            return error_response("分类不存在", 404)
        
        # 如果要更新名称，检查新名称是否已存在
        if category_data.name and category_data.name != existing_category['name']:
            name_exists = CategoryDB.get_category_by_name(category_data.name)
            if name_exists:
                return error_response("分类名称已存在", 400)
        
        success = CategoryDB.update_category(category_id, category_data.name, category_data.description)
        if not success:
            return error_response("更新分类失败", 500)
        
        return success_response("分类更新成功")
    
    except Exception as e:
        logger.error(f"更新分类失败: {e}")
        return error_response("更新分类失败", 500)

@app.delete("/admin/categories/{category_id}")
async def delete_category(category_id: str, request: Request):
    """删除分类"""
    # 验证管理员权限
    admin = get_current_admin_required_from_cookie(request)
    
    try:
        # 检查分类是否存在
        existing_category = CategoryDB.get_category_by_id(category_id)
        if not existing_category:
            return error_response("分类不存在", 404)
        
        success = CategoryDB.delete_category(category_id)
        if not success:
            return error_response("删除失败，该分类下还有商品", 400)
        
        return success_response("分类删除成功")
    
    except Exception as e:
        logger.error(f"删除分类失败: {e}")
        return error_response("删除分类失败", 500)

# ==================== 订单路由 ====================

@app.post("/orders")
async def create_order(
    order_request: OrderCreateRequest,
    request: Request
):
    """创建订单"""
    # 验证用户登录状态
    user = get_current_user_required_from_cookie(request)

    try:
        shipping_info = dict(order_request.shipping_info or {})
        scope = resolve_shopping_scope(
            request,
            address_id=shipping_info.get('address_id'),
            building_id=shipping_info.get('building_id')
        )

        validation = scope.get('address_validation') or check_address_and_building(
            shipping_info.get('address_id'),
            shipping_info.get('building_id')
        )
        if not validation.get('is_valid'):
            reason = validation.get('reason')
            if reason in ('missing_address', 'missing_building'):
                message = validation.get('message') or '请先选择收货地址'
            else:
                message = validation.get('message') or '地址不存在或未启用，请联系管理员'
            return error_response(message, 400)

        # 检查代理/店铺打烊与预约状态
        agent_id = scope.get('agent_id')
        reservation_due_to_closure = False
        closure_note = ''
        allow_reservation_when_closed = False
        closure_requires_reservation_only = False
        closure_prefix = '店铺已暂停营业。'

        if agent_id:
            status = AgentStatusDB.get_agent_status(agent_id)
            agent_open = bool(status.get('is_open', 1))
            allow_reservation_when_closed = bool(status.get('allow_reservation', 0))
            if not agent_open:
                closure_note = status.get('closed_note', '')
                closure_prefix = '当前区域代理已暂停营业。'
                reservation_due_to_closure = True
                if not allow_reservation_when_closed:
                    closure_requires_reservation_only = True
        else:
            is_open = SettingsDB.get('shop_is_open', '1') != '0'
            allow_reservation_when_closed = SettingsDB.get('shop_reservation_enabled', 'false') == 'true'
            if not is_open:
                closure_note = SettingsDB.get('shop_closed_note', '')
                reservation_due_to_closure = True
                if not allow_reservation_when_closed:
                    closure_requires_reservation_only = True

        # 获取用户购物车
        cart_data = CartDB.get_cart(user["id"])
        if not cart_data or not cart_data["items"]:
            return error_response("购物车为空，无法创建订单", 400)

        owner_ids = scope["owner_ids"]
        include_unassigned = False if owner_ids else True

        # 同步回写标准化后的地址信息
        shipping_info['address_id'] = scope.get('address_id')
        shipping_info['building_id'] = scope.get('building_id')
        if scope.get('agent_id'):
            shipping_info['agent_id'] = scope['agent_id']

        # 获取购物车中的商品信息
        items_dict = cart_data["items"]
        all_products = ProductDB.get_all_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
        product_dict = {p["id"]: p for p in all_products}
        
        # 构建订单商品列表并计算总金额（应用折扣）
        order_items = []
        total_amount = 0.0
        items_require_reservation = False
        cart_item_count = 0
        all_cart_items_reservation_only = True

        SEP = '@@'
        for key, quantity in items_dict.items():
            product_id = key
            variant_id = None
            if isinstance(key, str) and SEP in key:
                product_id, variant_id = key.split(SEP, 1)
            if product_id in product_dict:
                product = product_dict[product_id]
                # 忽略下架商品
                if int(product.get("is_active", 1) or 1) != 1:
                    continue
                non_sellable = is_non_sellable(product)
                
                # 折扣后单价
                zhe = float(product.get("discount", 10.0) or 10.0)
                unit_price = round(float(product["price"]) * (zhe / 10.0), 2)
                
                # 库存检查：有规格则检查规格库存，否则检查商品库存
                if variant_id:
                    from database import VariantDB
                    variant = VariantDB.get_by_id(variant_id)
                    if not variant or variant.get('product_id') != product_id:
                        return error_response("规格不存在", 400)
                    if not non_sellable and quantity > int(variant.get('stock', 0)):
                        return error_response(f"商品 {product['name']}（{variant.get('name')}）库存不足", 400)
                else:
                    if not non_sellable and quantity > product.get("stock", 0):
                        return error_response(f"商品 {product['name']} 库存不足", 400)

                subtotal = unit_price * quantity
                if non_sellable:
                    subtotal = 0.0
                else:
                    total_amount += subtotal
                cart_item_count += 1
                
                requires_reservation = False
                try:
                    requires_reservation = int(product.get("reservation_required", 0) or 0) == 1
                except Exception:
                    requires_reservation = bool(product.get("reservation_required"))
                cutoff_value = product.get("reservation_cutoff")
                note_value = (product.get("reservation_note") or '').strip()

                item = {
                    "product_id": product_id,
                    "name": product["name"],
                    "unit_price": round(unit_price, 2),
                    "quantity": quantity,
                    "subtotal": round(subtotal, 2),
                    "category": product.get("category", ""),
                    "img_path": product.get("img_path", ""),
                    "is_not_for_sale": non_sellable
                }
                if requires_reservation:
                    items_require_reservation = True
                    item["is_reservation"] = True
                    if cutoff_value:
                        try:
                            item["reservation_cutoff"] = normalize_reservation_cutoff(str(cutoff_value))
                        except HTTPException:
                            item["reservation_cutoff"] = None
                    if note_value:
                        item["reservation_note"] = note_value[:120]
                else:
                    all_cart_items_reservation_only = False
                if variant_id:
                    item["variant_id"] = variant_id
                    item["variant_name"] = variant.get("name")
                order_items.append(item)

        # 保存商品金额小计（不含运费），用于满额判断
        items_subtotal = round(total_amount, 2)
        all_items_are_reservation = cart_item_count > 0 and all_cart_items_reservation_only

        # 打烊期间的预约逻辑
        if reservation_due_to_closure:
            if allow_reservation_when_closed:
                # 管理面板开启了预约：所有商品都可以预约购买，无需检查
                pass
            else:
                # 管理面板未开启预约：仅允许标记为预约的商品购买
                if not all_items_are_reservation:
                    fallback = closure_note.strip() or '暂不支持下单'
                    return error_response(f"{closure_prefix}当前仅支持预约商品下单，请移除非预约商品后再试。{fallback}", 400)

        # 新的满额赠品系统：支持多层次门槛配置
        owner_scope_id = get_owner_id_from_scope(scope)
        lottery_threshold = LotteryConfigDB.get_threshold(owner_scope_id)

        try:
            applicable_thresholds = GiftThresholdDB.get_applicable_thresholds(items_subtotal, owner_scope_id)
            logger.info(f"订单金额 {items_subtotal} 元，适用门槛: {[t.get('threshold_amount') for t in applicable_thresholds]}")
            
            for threshold in applicable_thresholds:
                threshold_id = threshold.get('id')
                threshold_amount = threshold.get('threshold_amount', 0)
                gift_products = threshold.get('gift_products', 0) == 1
                gift_coupon = threshold.get('gift_coupon', 0) == 1
                coupon_amount = threshold.get('coupon_amount', 0)
                applicable_times = threshold.get('applicable_times', 0)
                
                # 处理商品赠品
                if gift_products and applicable_times > 0:
                    try:
                        selected_gifts = GiftThresholdDB.pick_gifts_for_threshold(threshold_id, owner_scope_id, applicable_times)
                        for gift in selected_gifts:
                            try:
                                gift_quantity = int(gift.get('quantity', 1))
                            except Exception:
                                gift_quantity = 1
                            if gift_quantity <= 0:
                                continue
                            gift_item = {
                                "product_id": gift.get('product_id'),
                                "name": gift.get('display_name') or gift.get('product_name') or '满额赠品',
                                "unit_price": 0.0,
                                "quantity": gift_quantity,
                                "subtotal": 0.0,
                                "category": gift.get('category') or '满额赠品',
                                "img_path": gift.get('img_path') or '',
                                "is_auto_gift": True,
                                "auto_gift_item_id": gift.get('threshold_item_id'),
                                "auto_gift_product_name": gift.get('product_name'),
                                "auto_gift_variant_name": gift.get('variant_name'),
                                "gift_threshold_id": threshold_id,
                                "gift_threshold_amount": threshold_amount
                            }
                            if gift.get('variant_id'):
                                gift_item['variant_id'] = gift.get('variant_id')
                                if gift.get('variant_name'):
                                    gift_item['variant_name'] = gift.get('variant_name')
                            order_items.append(gift_item)
                    except Exception as e:
                        logger.warning(f"生成满额赠品失败 (门槛{threshold_amount}): {e}")
                
                # 处理优惠券赠品（在订单创建时不发放，记录信息供支付成功后发放）
                if gift_coupon and coupon_amount > 0 and applicable_times > 0:
                    logger.info(f"记录满额优惠券待发放：{applicable_times} 张 {coupon_amount} 元（门槛{threshold_amount}）")
        except Exception as e:
            logger.warning(f"处理满额赠品配置失败: {e}")

        # 若已达满门槛且抽奖已启用，则自动附加可用抽奖奖品（不计入总价）
        rewards_attached_ids: List[str] = []
        lottery_enabled = LotteryConfigDB.get_enabled(owner_scope_id)
        if lottery_enabled and items_subtotal >= lottery_threshold:
            try:
                rewards = RewardDB.get_eligible_rewards(
                    user["id"],
                    owner_scope_id,
                    restrict_owner=True
                ) or []
                for r in rewards:
                    qty = int(r.get("prize_quantity") or 1)
                    prize_name = r.get("prize_name") or "抽奖奖品"
                    prize_pid = r.get("prize_product_id") or f"prize_{int(datetime.now().timestamp())}"
                    prize_variant_id = r.get("prize_variant_id")
                    prize_variant_name = r.get("prize_variant_name")
                    prize_product_name = r.get("prize_product_name") or prize_name
                    prize_img_path = r.get("prize_img_path") or ""
                    try:
                        recorded_value = float(r.get("prize_unit_price") or 0.0)
                    except Exception:
                        recorded_value = 0.0
                    lottery_item = {
                        "product_id": prize_pid,
                        "name": prize_name,
                        "unit_price": 0.0,
                        "quantity": qty,
                        "subtotal": 0.0,
                        "category": "抽奖",
                        "is_lottery": True,
                        "img_path": prize_img_path,
                        "image_url": prize_img_path,
                        "lottery_display_name": prize_name,
                        "lottery_product_id": prize_pid,
                        "lottery_product_name": prize_product_name,
                        "lottery_variant_id": prize_variant_id,
                        "lottery_variant_name": prize_variant_name,
                        "lottery_unit_price": recorded_value,
                        "lottery_group_id": r.get("prize_group_id"),
                        "lottery_reward_id": r.get("id")
                    }
                    if prize_variant_id:
                        lottery_item["variant_id"] = prize_variant_id
                        if prize_variant_name:
                            lottery_item["variant_name"] = prize_variant_name
                    order_items.append(lottery_item)
                    rewards_attached_ids.append(r.get("id"))
            except Exception as e:
                logger.warning(f"附加抽奖奖品失败: {e}")

        user_confirms_reservation = bool(order_request.reservation_requested)
        if items_require_reservation and not user_confirms_reservation:
            if reservation_due_to_closure:
                tip = closure_note or '当前打烊，仅支持预约购买'
                return error_response(f"{tip}（请确认预约购买后再试）", 400)
            return error_response("该商品需要预约购买，请确认预约方式后再提交订单", 400)

        reservation_reasons: List[str] = []
        if items_require_reservation:
            reservation_reasons.append('商品预约')
        if reservation_due_to_closure:
            reservation_reasons.append('店铺打烊预约')

        is_reservation_order = len(reservation_reasons) > 0
        if is_reservation_order:
            shipping_info['reservation'] = True
            shipping_info['reservation_reasons'] = reservation_reasons
            if reservation_due_to_closure:
                shipping_info['reservation_due_to_closure'] = True
                if closure_note:
                    shipping_info['reservation_closure_note'] = closure_note
            if items_require_reservation:
                shipping_info['reservation_items'] = True

        # 获取配送费配置并计算运费（仅基于上架商品金额）
        scope = resolve_shopping_scope(request)
        owner_id = get_owner_id_from_scope(scope)
        delivery_config = DeliverySettingsDB.get_delivery_config(owner_id)
        
        # 运费规则：基础配送费或免配送费门槛任意一个为0则免费，否则达到门槛免费，否则收取基础配送费（商品金额为0时不收取）
        shipping_fee = 0.0 if delivery_config['delivery_fee'] == 0 or delivery_config['free_delivery_threshold'] == 0 or items_subtotal >= delivery_config['free_delivery_threshold'] else (delivery_config['delivery_fee'] if items_subtotal > 0 else 0.0)

        # 处理优惠券（每单最多1张；仅当商品金额严格大于券额时可用）
        discount_amount = 0.0
        used_coupon_id = None
        if order_request.apply_coupon and order_request.coupon_id:
            try:
                coupon = CouponDB.check_valid_for_student(order_request.coupon_id, user["id"], owner_scope_id)  # 校验归属、状态、未过期
                if coupon:
                    try:
                        amt = float(coupon.get('amount') or 0)
                    except Exception:
                        amt = 0.0
                    if items_subtotal > amt and amt > 0:
                        discount_amount = round(amt, 2)
                        used_coupon_id = coupon.get('id')
            except Exception as e:
                logger.warning(f"校验优惠券失败: {e}")

        # 订单总金额 = 商品小计 - 优惠 + 运费（不为负）
        total_amount = round(max(0.0, items_subtotal - discount_amount) + shipping_fee, 2)

        if cart_item_count == 0:
            return error_response("购物车中没有可结算的上架商品", 400)

        # 创建订单（暂不扣减库存，等待支付成功）
        order_id = OrderDB.create_order(
            user_identifier=user["id"],  # 使用学号，数据库会自动解析为user_id
            total_amount=round(total_amount, 2),
            shipping_info=shipping_info,
            items=order_items,
            payment_method=order_request.payment_method,
            note=order_request.note,
            discount_amount=discount_amount,
            coupon_id=used_coupon_id,
            address_id=scope.get('address_id'),
            building_id=scope.get('building_id'),
            agent_id=scope.get('agent_id'),
            is_reservation=is_reservation_order,
            reservation_reason='; '.join(reservation_reasons) if reservation_reasons else None
        )
        # 锁定优惠券，防止未付款或待确认期间被重复使用
        if used_coupon_id and discount_amount > 0:
            try:
                CouponDB.lock_for_order(used_coupon_id, order_id)
            except Exception as e:
                logger.warning(f"锁定优惠券失败: {e}")
        # 标记已附加的奖品为已消费（绑定到本订单）
        if rewards_attached_ids:
            try:
                RewardDB.consume_rewards(
                    user["id"],
                    rewards_attached_ids,
                    order_id,
                    owner_scope_id
                )
            except Exception as e:
                logger.warning(f"标记抽奖奖品消费失败: {e}")
        
        # 立即更新用户的最新收货信息，确保下次自动填写使用最新数据
        try:
            shipping_profile = dict(shipping_info)
            UserProfileDB.upsert_shipping(user["id"], shipping_profile)
            logger.info(f"已更新用户 {user['id']} 的最新收货信息")
        except Exception as e:
            logger.warning(f"更新用户收货信息失败: {e}")
        
        # 注意：库存扣减移到支付成功后处理
        # 购物车清空也移到支付成功后处理
        
        return success_response("订单创建成功", {"order_id": order_id, "total_amount": round(total_amount, 2), "discount_amount": round(discount_amount, 2), "coupon_id": used_coupon_id})
    
    except Exception as e:
        logger.error(f"创建订单失败: {e}")
        return error_response("创建订单失败", 500)

def _enrich_order_items_with_images(order: Dict) -> Dict:
    """为订单商品补充图片信息（主要用于抽奖/赠品商品的旧订单兼容）"""
    items = order.get("items") or []
    if not items:
        return order
    
    # 收集需要查询图片的商品ID
    product_ids_to_fetch = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        # 只处理缺少图片的抽奖或赠品商品
        if (item.get("is_lottery") or item.get("is_auto_gift")):
            if not item.get("img_path") and not item.get("image_url"):
                # 优先使用 lottery_product_id，其次使用 product_id
                pid = item.get("lottery_product_id") or item.get("product_id")
                if pid and not pid.startswith("prize_"):  # 排除生成的临时ID
                    product_ids_to_fetch.add(pid)
    
    if not product_ids_to_fetch:
        return order
    
    # 批量查询商品图片
    product_images = {}
    for pid in product_ids_to_fetch:
        try:
            product = ProductDB.get_product_by_id(pid)
            if product and product.get("img_path"):
                product_images[pid] = product["img_path"]
        except Exception:
            pass
    
    # 更新订单项的图片
    for item in items:
        if not isinstance(item, dict):
            continue
        if (item.get("is_lottery") or item.get("is_auto_gift")):
            if not item.get("img_path") and not item.get("image_url"):
                pid = item.get("lottery_product_id") or item.get("product_id")
                if pid and pid in product_images:
                    item["img_path"] = product_images[pid]
                    item["image_url"] = product_images[pid]
    
    return order

@app.get("/orders/my")
async def get_my_orders(request: Request):
    """获取用户的订单列表"""
    # 验证用户登录状态
    user = get_current_user_required_from_cookie(request)
    
    try:
        orders = OrderDB.get_orders_by_student(user["id"])
        
        # 将创建时间转换为时间戳（秒），正确处理时区问题
        for order in orders:
            if order.get("created_at"):
                order["created_at_timestamp"] = convert_sqlite_timestamp_to_unix(order["created_at"], order["id"])
            # 为订单商品补充图片信息
            _enrich_order_items_with_images(order)
        
        return success_response("获取订单列表成功", {"orders": orders})
    
    except Exception as e:
        logger.error(f"获取订单列表失败: {e}")
        return error_response("获取订单列表失败", 500)

@app.get("/orders/{order_id}")
async def get_order_detail(order_id: str, request: Request):
    """获取订单详情"""
    # 验证用户登录状态
    user = get_current_user_required_from_cookie(request)
    
    try:
        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        
        # 检查权限（只有订单创建者或管理员可以查看）
        if order["student_id"] != user["id"] and user.get("type") != "admin":
            return error_response("无权查看此订单", 403)
        
        # 添加时间戳转换（与订单列表接口保持一致）
        if order.get("created_at"):
            order["created_at_timestamp"] = convert_sqlite_timestamp_to_unix(order["created_at"], order["id"])
        
        # 为订单商品补充图片信息
        _enrich_order_items_with_images(order)
        
        return success_response("获取订单详情成功", {"order": order})
    
    except Exception as e:
        logger.error(f"获取订单详情失败: {e}")
        return error_response("获取订单详情失败", 500)

# ==================== 管理员订单路由 ====================

@app.get("/admin/orders")
async def get_all_orders(
    request: Request,
    limit: Optional[int] = 20,
    offset: Optional[int] = 0,
    order_id: Optional[str] = None,
    agent_id: Optional[str] = None,
    keyword: Optional[str] = None
):
    """获取订单（管理员）——支持分页与按订单ID精确搜索。
    默认每次最多返回20条，通过翻页继续获取，避免一次拿全表。
    """
    staff = get_current_staff_required_from_cookie(request)

    try:
        # 后端兜底保护，强制限制最大单次返回数量
        try:
            limit_val = int(limit or 20)
        except Exception:
            limit_val = 20
        if limit_val <= 0:
            limit_val = 20
        if limit_val > 100:
            limit_val = 100
        try:
            offset_val = int(offset or 0)
        except Exception:
            offset_val = 0
        if offset_val < 0:
            offset_val = 0

        scope = build_staff_scope(staff)
        (
            selected_agent_id,
            selected_address_ids,
            selected_building_ids,
            exclude_address_ids,
            exclude_building_ids,
            selected_filter
        ) = resolve_staff_order_scope(staff, scope, agent_id)

        page_data = OrderDB.get_orders_paginated(
            order_id=order_id,
            keyword=keyword,
            limit=limit_val,
            offset=offset_val,
            agent_id=selected_agent_id,
            address_ids=selected_address_ids,
            building_ids=selected_building_ids,
            exclude_address_ids=exclude_address_ids,
            exclude_building_ids=exclude_building_ids
        )
        orders = page_data.get("orders", [])
        total = int(page_data.get("total", 0))

        # 为管理员订单列表也添加时间戳转换
        for order in orders:
            if order.get("created_at"):
                order["created_at_timestamp"] = convert_sqlite_timestamp_to_unix(order["created_at"], order["id"])

        stats = OrderDB.get_order_stats(
            agent_id=selected_agent_id,
            address_ids=selected_address_ids,
            building_ids=selected_building_ids,
            exclude_address_ids=exclude_address_ids,
            exclude_building_ids=exclude_building_ids
        )
        has_more = (offset_val + len(orders)) < total
        return success_response("获取订单列表成功", {
            "orders": orders,
            "stats": stats,
            "total": total,
            "limit": limit_val,
            "offset": offset_val,
            "has_more": has_more,
            "scope": scope,
            "selected_agent_id": selected_agent_id,
            "selected_agent_filter": selected_filter or 'self'
        })
    
    except Exception as e:
        logger.error(f"获取订单列表失败: {e}")
        return error_response("获取订单列表失败", 500)


@app.get("/agent/orders")
async def get_agent_orders(
    request: Request,
    limit: Optional[int] = 20,
    offset: Optional[int] = 0,
    order_id: Optional[str] = None,
    keyword: Optional[str] = None
):
    """获取订单列表（代理）"""
    _agent, scope = require_agent_with_scope(request)

    try:
        try:
            limit_val = int(limit or 20)
        except Exception:
            limit_val = 20
        if limit_val <= 0:
            limit_val = 20
        if limit_val > 100:
            limit_val = 100

        try:
            offset_val = int(offset or 0)
        except Exception:
            offset_val = 0
        if offset_val < 0:
            offset_val = 0

        page_data = OrderDB.get_orders_paginated(
            order_id=order_id,
            keyword=keyword,
            limit=limit_val,
            offset=offset_val,
            agent_id=scope.get('agent_id'),
            address_ids=scope.get('address_ids'),
            building_ids=scope.get('building_ids')
        )

        orders = page_data.get("orders", [])
        total = int(page_data.get("total", 0))

        for order in orders:
            if order.get("created_at"):
                order["created_at_timestamp"] = convert_sqlite_timestamp_to_unix(order["created_at"], order["id"])

        stats = OrderDB.get_order_stats(
            agent_id=scope.get('agent_id'),
            address_ids=scope.get('address_ids'),
            building_ids=scope.get('building_ids')
        )

        has_more = (offset_val + len(orders)) < total

        return success_response("获取订单列表成功", {
            "orders": orders,
            "stats": stats,
            "total": total,
            "limit": limit_val,
            "offset": offset_val,
            "has_more": has_more,
            "scope": scope
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"代理获取订单列表失败: {e}")
        return error_response("获取订单列表失败", 500)


async def create_export_job_for_staff(staff: Dict[str, Any], payload: OrderExportRequest, staff_prefix: str):
    """创建导出任务，返回基础信息和最新历史。"""
    start_ms = payload.start_time_ms
    end_ms = payload.end_time_ms
    if start_ms is not None and end_ms is not None and start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms
    tz_offset = payload.timezone_offset_minutes
    keyword = (payload.keyword or '').strip() or None
    status_filter = (payload.status_filter or '').strip()
    unified_status = status_filter if status_filter and status_filter != '全部' else None

    (
        selected_agent_id,
        selected_address_ids,
        selected_building_ids,
        exclude_address_ids,
        exclude_building_ids,
        selected_filter,
        filter_admin_orders
    ) = prepare_export_scope(staff, payload.agent_filter if staff.get('type') == 'admin' else 'self')

    page_data = OrderDB.get_orders_paginated(
        keyword=keyword,
        limit=1,
        offset=0,
        agent_id=selected_agent_id,
        address_ids=selected_address_ids,
        building_ids=selected_building_ids,
        exclude_address_ids=exclude_address_ids,
        exclude_building_ids=exclude_building_ids,
        start_time_ms=start_ms,
        end_time_ms=end_ms,
        unified_status=unified_status,
        filter_admin_orders=filter_admin_orders
    )
    total = int(page_data.get("total") or 0)
    if total <= 0:
        return error_response("当前筛选条件下没有可导出的订单", 400)

    agent_name_map = build_agent_name_map()
    scope_label = resolve_scope_label(selected_filter, staff, agent_name_map)
    owner_id = get_owner_id_for_staff(staff) or staff.get('id')
    filename = build_export_filename(start_ms, end_ms)

    job = OrderExportDB.create_job(
        owner_id=owner_id,
        role=staff.get('type'),
        agent_filter=selected_filter,
        keyword=keyword,
        status_filter=unified_status,
        start_time_ms=start_ms,
        end_time_ms=end_ms,
        scope_label=scope_label,
        filename=filename,
        total_count=total,
        client_tz_offset=tz_offset
    )

    history_rows = OrderExportDB.list_jobs_for_owner(owner_id, limit=12)
    history = [serialize_export_job(entry, staff_prefix) for entry in history_rows]

    return success_response("导出任务已创建", {
        "job_id": job["id"],
        "stream_path": f"{staff_prefix}/orders/export/stream/{job['id']}",
        "history": history,
        "expires_at": job.get("expires_at"),
        "total": total,
        "filename": job.get("filename"),
        "range_label": format_export_range_label(start_ms, end_ms, tz_offset),
        "scope_label": scope_label
    })


async def stream_export_for_staff(request: Request, staff: Dict[str, Any], staff_prefix: str, job_id: str):
    """SSE流式导出订单并推送进度。"""
    owner_id = get_owner_id_for_staff(staff) or staff.get('id')
    job = OrderExportDB.get_job(job_id)
    if not job or job.get('owner_id') != owner_id:
        raise HTTPException(status_code=404, detail="导出任务不存在")

    agent_name_map = build_agent_name_map()
    selected_filter_value = job.get('agent_filter') or 'self'
    (
        selected_agent_id,
        selected_address_ids,
        selected_building_ids,
        exclude_address_ids,
        exclude_building_ids,
        _resolved_filter,
        filter_admin_orders
    ) = prepare_export_scope(staff, selected_filter_value)

    unified_status = job.get('status_filter') or None
    keyword = job.get('keyword') or None
    start_ms = job.get('start_time_ms')
    end_ms = job.get('end_time_ms')
    tz_offset = job.get('client_tz_offset')
    is_admin_role = staff.get('type') == 'admin'
    safe_filename = os.path.basename(job.get('filename') or "") or f"{job_id}.xlsx"
    file_path = os.path.abspath(os.path.join(exports_dir, safe_filename))
    safe_root = os.path.abspath(exports_dir)
    if not file_path.startswith(safe_root):
        file_path = os.path.join(safe_root, f"{job_id}.xlsx")

    def is_expired(expires_at: Optional[str]) -> bool:
        if not expires_at:
            return False
        try:
            expire_dt = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
            return expire_dt <= datetime.now()
        except Exception:
            return False

    async def event_generator():
        nonlocal job
        try:
            if is_expired(job.get('expires_at')):
                OrderExportDB.update_job(job_id, status='expired', message='导出链接已过期')
                refreshed = OrderExportDB.get_job(job_id)
                history_rows = OrderExportDB.list_jobs_for_owner(owner_id, limit=12)
                history = [serialize_export_job(entry, staff_prefix) for entry in history_rows]
                yield {"data": json.dumps({
                    "status": "expired",
                    "message": "导出链接已过期，请重新生成",
                    "history": history,
                    "range_label": format_export_range_label(start_ms, end_ms, tz_offset)
                })}
                return

            if job.get('status') == 'completed' and os.path.exists(file_path):
                history_rows = OrderExportDB.list_jobs_for_owner(owner_id, limit=12)
                history = [serialize_export_job(entry, staff_prefix) for entry in history_rows]
                final_job = serialize_export_job(job, staff_prefix)
                final_job.update({
                    "status": "completed",
                    "progress": 100,
                    "stage": "已完成",
                    "exported": job.get('exported_count'),
                    "total": job.get('total_count'),
                    "history": history
                })
                yield {"data": json.dumps(final_job)}
                return

            OrderExportDB.update_job(job_id, status='running', message='正在准备导出')
            yield {"data": json.dumps({
                "status": "running",
                "stage": "准备导出",
                "progress": 5,
                "total": job.get('total_count'),
                "range_label": format_export_range_label(start_ms, end_ms, tz_offset)
            })}

            exported_rows: List[List[str]] = []
            exported_count = 0
            total_count = 0
            offset = 0
            batch_size = 200
            progress = 5

            while True:
                page_data = OrderDB.get_orders_paginated(
                    keyword=keyword,
                    limit=batch_size,
                    offset=offset,
                    agent_id=selected_agent_id,
                    address_ids=selected_address_ids,
                    building_ids=selected_building_ids,
                    exclude_address_ids=exclude_address_ids,
                    exclude_building_ids=exclude_building_ids,
                    start_time_ms=start_ms,
                    end_time_ms=end_ms,
                    unified_status=unified_status,
                    filter_admin_orders=filter_admin_orders,
                    allow_large_limit=True
                )
                orders_batch = page_data.get("orders") or []
                if offset == 0:
                    total_count = int(page_data.get("total") or 0)
                    OrderExportDB.update_job(job_id, total_count=total_count)

                if not orders_batch:
                    break

                for order in orders_batch:
                    exported_rows.append(build_export_row(order, agent_name_map, staff, is_admin_role, tz_offset))

                exported_count = len(exported_rows)
                OrderExportDB.update_job(job_id, exported_count=exported_count)
                progress = 10 if total_count == 0 else min(96, max(10, int(exported_count / max(total_count, 1) * 85)))
                yield {"data": json.dumps({
                    "status": "running",
                    "stage": "正在解析数据",
                    "progress": progress,
                    "exported": exported_count,
                    "total": total_count,
                    "message": f"正在解析... {exported_count}/{total_count or '未知'}"
                })}

                offset += len(orders_batch)
                if len(orders_batch) < batch_size:
                    break

            if exported_count == 0:
                OrderExportDB.update_job(job_id, status='failed', message='当前筛选无数据')
                yield {"data": json.dumps({
                    "status": "failed",
                    "message": "当前筛选条件下没有可导出的订单"
                })}
                return

            OrderExportDB.update_job(job_id, exported_count=exported_count, message='正在生成文件')
            yield {"data": json.dumps({
                "status": "running",
                "stage": "生成文件",
                "progress": min(98, max(progress, 90)),
                "exported": exported_count,
                "total": total_count
            })}

            await asyncio.to_thread(write_export_workbook, exported_rows, file_path)

            OrderExportDB.update_job(
                job_id,
                status='completed',
                exported_count=exported_count,
                total_count=total_count,
                file_path=file_path,
                message='导出完成',
                filename=os.path.basename(file_path)
            )
            job = OrderExportDB.get_job(job_id)
            history_rows = OrderExportDB.list_jobs_for_owner(owner_id, limit=12)
            history = [serialize_export_job(entry, staff_prefix) for entry in history_rows]
            final_job = serialize_export_job(job, staff_prefix)
            final_job.update({
                "status": "completed",
                "progress": 100,
                "stage": "已完成",
                "exported": exported_count,
                "total": total_count,
                "history": history
            })
            yield {"data": json.dumps(final_job)}
        except Exception as e:
            logger.error(f"导出订单失败({job_id}): {e}")
            OrderExportDB.update_job(job_id, status='failed', message=str(e))
            yield {"data": json.dumps({
                "status": "failed",
                "message": str(e) or "导出失败"
            })}

    return EventSourceResponse(event_generator(), ping=15000)


async def download_export_for_staff(staff: Dict[str, Any], job_id: str, token: Optional[str]):
    """下载导出文件并校验有效期。"""
    owner_id = get_owner_id_for_staff(staff) or staff.get('id')
    job = OrderExportDB.get_job(job_id)
    if not job or job.get('owner_id') != owner_id:
        raise HTTPException(status_code=404, detail="导出记录不存在")
    if token is None or token != job.get('download_token'):
        raise HTTPException(status_code=403, detail="下载链接已失效，请重新导出")
    if job.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="文件尚未生成，请稍后重试")
    expires_at = job.get('expires_at')
    if expires_at:
        try:
            expire_dt = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
            if expire_dt <= datetime.now():
                OrderExportDB.update_job(job_id, status='expired', message='导出链接已过期')
                raise HTTPException(status_code=410, detail="导出链接已过期，请重新导出")
        except ValueError:
            pass

    file_path = job.get('file_path')
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="导出文件不存在，请重新导出")

    filename = os.path.basename(job.get('filename') or file_path) or f"{job_id}.xlsx"
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


async def get_export_history_for_staff(staff: Dict[str, Any], staff_prefix: str):
    owner_id = get_owner_id_for_staff(staff) or staff.get('id')
    rows = OrderExportDB.list_jobs_for_owner(owner_id, limit=12)
    history = [serialize_export_job(row, staff_prefix) for row in rows]
    return success_response("获取导出记录成功", {"history": history})


@app.post("/admin/orders/export")
async def admin_create_order_export(request: Request, payload: OrderExportRequest):
    staff = get_current_staff_required_from_cookie(request)
    return await create_export_job_for_staff(staff, payload, '/admin')


@app.post("/agent/orders/export")
async def agent_create_order_export(request: Request, payload: OrderExportRequest):
    agent, _ = require_agent_with_scope(request)
    return await create_export_job_for_staff(agent, payload, '/agent')


@app.get("/admin/orders/export/stream/{job_id}")
async def admin_stream_order_export(job_id: str, request: Request):
    staff = get_current_staff_required_from_cookie(request)
    return await stream_export_for_staff(request, staff, '/admin', job_id)


@app.get("/agent/orders/export/stream/{job_id}")
async def agent_stream_order_export(job_id: str, request: Request):
    agent, _ = require_agent_with_scope(request)
    return await stream_export_for_staff(request, agent, '/agent', job_id)


@app.get("/admin/orders/export/download/{job_id}")
async def admin_download_order_export(job_id: str, request: Request, token: Optional[str] = None):
    staff = get_current_staff_required_from_cookie(request)
    return await download_export_for_staff(staff, job_id, token)


@app.get("/agent/orders/export/download/{job_id}")
async def agent_download_order_export(job_id: str, request: Request, token: Optional[str] = None):
    agent, _ = require_agent_with_scope(request)
    return await download_export_for_staff(agent, job_id, token)


@app.get("/admin/orders/export/history")
async def admin_export_history(request: Request):
    staff = get_current_staff_required_from_cookie(request)
    return await get_export_history_for_staff(staff, '/admin')


@app.get("/agent/orders/export/history")
async def agent_export_history(request: Request):
    agent, _ = require_agent_with_scope(request)
    return await get_export_history_for_staff(agent, '/agent')

@app.get("/admin/lottery-config")
async def admin_get_lottery_config(request: Request, owner_id: Optional[str] = None):
    """读取抽奖配置（管理员）。"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        prizes = LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True)
        config = LotteryConfigDB.get_config(owner_id)
        return success_response("获取抽奖配置成功", {
            "prizes": prizes,
            "threshold_amount": config["threshold_amount"],
            "is_enabled": config["is_enabled"]
        })
    except Exception as e:
        logger.error(f"读取抽奖配置失败: {e}")
        return error_response("读取抽奖配置失败", 500)


@app.get("/agent/lottery-config")
async def agent_get_lottery_config(request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        prizes = LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True)
        config = LotteryConfigDB.get_config(owner_id)
        return success_response("获取抽奖配置成功", {
            "prizes": prizes,
            "threshold_amount": config["threshold_amount"],
            "is_enabled": config["is_enabled"]
        })
    except Exception as e:
        logger.error(f"代理读取抽奖配置失败: {e}")
        return error_response("读取抽奖配置失败", 500)


class LotteryPrizeItemInput(BaseModel):
    id: Optional[str] = None
    product_id: str
    variant_id: Optional[str] = None


class LotteryPrizeInput(BaseModel):
    id: Optional[str] = None
    display_name: str
    weight: float
    is_active: Optional[bool] = True
    items: List[LotteryPrizeItemInput] = []


class LotteryConfigUpdateRequest(BaseModel):
    prizes: List[LotteryPrizeInput] = []
    threshold_amount: Optional[float] = None


class LotteryThresholdUpdateRequest(BaseModel):
    threshold_amount: float


class LotteryEnabledUpdateRequest(BaseModel):
    is_enabled: bool


class AutoGiftItemInput(BaseModel):
    product_id: str
    variant_id: Optional[str] = None


class AutoGiftUpdateRequest(BaseModel):
    items: List[AutoGiftItemInput] = []


def normalize_per_order_limit(value: Optional[Any]) -> Optional[int]:
    """将传入的每单赠品上限标准化为正整数或None"""
    if value is None:
        return None
    try:
        numeric = int(value)
    except (ValueError, TypeError):
        return None
    return numeric if numeric > 0 else None


# 满额门槛配置模型
class GiftThresholdCreate(BaseModel):
    threshold_amount: float
    gift_products: bool = False
    gift_coupon: bool = False
    coupon_amount: float = 0.0
    per_order_limit: Optional[int] = None
    items: List[AutoGiftItemInput] = []


class GiftThresholdUpdate(BaseModel):
    threshold_amount: Optional[float] = None
    gift_products: Optional[bool] = None
    gift_coupon: Optional[bool] = None
    coupon_amount: Optional[float] = None
    per_order_limit: Optional[int] = None
    is_active: Optional[bool] = None
    items: Optional[List[AutoGiftItemInput]] = None


# 配送费设置模型
class DeliverySettingsCreate(BaseModel):
    delivery_fee: float = 1.0
    free_delivery_threshold: float = 10.0


def _persist_lottery_prize_from_payload(
    prize: LotteryPrizeInput,
    owner_id: Optional[str],
    override_id: Optional[str] = None
) -> str:
    display_name = (prize.display_name or '').strip()
    if not display_name:
        raise ValueError("奖项名称不能为空")
    try:
        weight_value = float(prize.weight)
    except Exception:
        raise ValueError("奖项权重必须为数字")
    is_active = True if prize.is_active is None else bool(prize.is_active)
    items_payload: List[Dict[str, Any]] = []
    for item in prize.items or []:
        if not item.product_id:
            continue
        items_payload.append({
            'id': item.id,
            'product_id': item.product_id,
            'variant_id': item.variant_id
        })
    return LotteryDB.upsert_prize(
        override_id or prize.id,
        display_name,
        weight_value,
        is_active,
        items_payload,
        owner_id
    )


def _search_inventory_for_selector(
    term: Optional[str],
    staff: Optional[Dict[str, Any]] = None,
    owner_override: Optional[str] = None
) -> List[Dict[str, Any]]:
    scope = build_staff_scope(staff) if staff else None
    owner_ids = scope.get('owner_ids') if scope else None
    if owner_override:
        owner_ids = [owner_override]
    # 现在所有商品都有owner_id，不再需要include_unassigned
    include_unassigned = False
    try:
        # 修改：搜索时包含下架商品，以便在选择器中显示下架状态
        if term:
            products = ProductDB.search_products(term, active_only=False, owner_ids=owner_ids, include_unassigned=include_unassigned)
        else:
            products = ProductDB.get_all_products(owner_ids=owner_ids, include_unassigned=include_unassigned)
    except Exception as e:
        logger.error(f"搜索商品失败: {e}")
        return []

    filtered: List[Dict[str, Any]] = []
    try:
        product_ids = [p['id'] for p in products]
    except Exception:
        product_ids = []

    variant_map: Dict[str, List[Dict[str, Any]]] = {}
    if product_ids:
        try:
            variant_map = VariantDB.get_for_products(product_ids)
        except Exception as e:
            logger.warning(f"获取规格失败: {e}")
            variant_map = {}

    for product in products:
        # 修改：正确处理 is_active 字段，不再跳过下架商品
        raw_is_active = product.get('is_active')
        if raw_is_active is None:
            is_active = True
        else:
            try:
                is_active = int(raw_is_active) == 1
            except Exception:
                is_active = True

        try:
            base_price = float(product.get('price') or 0)
        except Exception:
            base_price = 0.0
        try:
            discount = float(product.get('discount', 10.0) or 10.0)
        except Exception:
            discount = 10.0
        retail_price = round(base_price * (discount / 10.0), 2)

        variants = variant_map.get(product.get('id')) or []
        if variants:
            for variant in variants:
                try:
                    stock = int(variant.get('stock') or 0)
                except Exception:
                    stock = 0
                # 修改：available 需要同时考虑库存和上架状态
                available = is_active and stock > 0
                filtered.append({
                    'product_id': product.get('id'),
                    'product_name': product.get('name'),
                    'variant_id': variant.get('id'),
                    'variant_name': variant.get('name'),
                    'stock': stock,
                    'retail_price': retail_price,
                    'img_path': product.get('img_path'),
                    'category': product.get('category'),
                    'is_active': is_active,
                    'available': available
                })
        else:
            try:
                stock = int(product.get('stock') or 0)
            except Exception:
                stock = 0
            # 修改：available 需要同时考虑库存和上架状态
            available = is_active and stock > 0
            filtered.append({
                'product_id': product.get('id'),
                'product_name': product.get('name'),
                'variant_id': None,
                'variant_name': None,
                'stock': stock,
                'retail_price': retail_price,
                'img_path': product.get('img_path'),
                'category': product.get('category'),
                'is_active': is_active,
                'available': available
            })

    # 修改：排序时上架商品优先，然后按名称排序
    filtered.sort(key=lambda x: (0 if x.get('is_active') else 1, x.get('product_name') or '', x.get('variant_name') or ''))
    return filtered[:100]


@app.put("/admin/lottery-config")
async def admin_update_lottery_config(payload: LotteryConfigUpdateRequest, request: Request, owner_id: Optional[str] = None):
    """批量更新抽奖配置，完全覆盖现有奖项。"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        prizes_payload = payload.prizes or []
        saved_ids: List[str] = []
        for prize in prizes_payload:
            saved_id = _persist_lottery_prize_from_payload(prize, owner_id)
            saved_ids.append(saved_id)
        LotteryDB.delete_prizes_not_in(saved_ids, owner_id)
        refreshed = LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True)
        if payload.threshold_amount is not None:
            LotteryConfigDB.set_threshold(owner_id, payload.threshold_amount)
        threshold_value = LotteryConfigDB.get_threshold(owner_id)
        return success_response("抽奖配置已更新", {
            "prizes": refreshed,
            "threshold_amount": threshold_value
        })
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"更新抽奖配置失败: {e}")
        return error_response("更新抽奖配置失败", 500)


@app.put("/agent/lottery-config")
async def agent_update_lottery_config(payload: LotteryConfigUpdateRequest, request: Request):
    """代理批量更新抽奖配置，完全覆盖自身奖项。"""
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        prizes_payload = payload.prizes or []
        saved_ids: List[str] = []
        for prize in prizes_payload:
            saved_id = _persist_lottery_prize_from_payload(prize, owner_id)
            saved_ids.append(saved_id)
        LotteryDB.delete_prizes_not_in(saved_ids, owner_id)
        refreshed = LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True)
        if payload.threshold_amount is not None:
            LotteryConfigDB.set_threshold(owner_id, payload.threshold_amount)
        threshold_value = LotteryConfigDB.get_threshold(owner_id)
        return success_response("抽奖配置已更新", {
            "prizes": refreshed,
            "threshold_amount": threshold_value
        })
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"代理更新抽奖配置失败: {e}")
        return error_response("更新抽奖配置失败", 500)


@app.patch("/admin/lottery-config/threshold")
async def admin_update_lottery_threshold(payload: LotteryThresholdUpdateRequest, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        value = LotteryConfigDB.set_threshold(owner_id, payload.threshold_amount)
        return success_response("抽奖门槛已更新", {"threshold_amount": value})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"更新抽奖门槛失败: {e}")
        return error_response("更新抽奖门槛失败", 500)


@app.patch("/agent/lottery-config/threshold")
async def agent_update_lottery_threshold(payload: LotteryThresholdUpdateRequest, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        value = LotteryConfigDB.set_threshold(owner_id, payload.threshold_amount)
        return success_response("抽奖门槛已更新", {"threshold_amount": value})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"代理更新抽奖门槛失败: {e}")
        return error_response("更新抽奖门槛失败", 500)


@app.patch("/admin/lottery-config/enabled")
async def admin_update_lottery_enabled(payload: LotteryEnabledUpdateRequest, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        is_enabled = LotteryConfigDB.set_enabled(owner_id, payload.is_enabled)
        return success_response("抽奖启用状态已更新", {"is_enabled": is_enabled})
    except Exception as e:
        logger.error(f"更新抽奖启用状态失败: {e}")
        return error_response("更新抽奖启用状态失败", 500)


@app.patch("/agent/lottery-config/enabled")
async def agent_update_lottery_enabled(payload: LotteryEnabledUpdateRequest, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        is_enabled = LotteryConfigDB.set_enabled(owner_id, payload.is_enabled)
        return success_response("抽奖启用状态已更新", {"is_enabled": is_enabled})
    except Exception as e:
        logger.error(f"代理更新抽奖启用状态失败: {e}")
        return error_response("更新抽奖启用状态失败", 500)


@app.get("/admin/auto-gifts")
async def admin_get_auto_gifts(request: Request):
    admin = get_current_admin_required_from_cookie(request)
    owner_id = get_owner_id_for_staff(admin)
    try:
        items = AutoGiftDB.list_items(owner_id)
        return success_response("获取满额赠品配置成功", {"items": items})
    except Exception as e:
        logger.error(f"读取满额赠品配置失败: {e}")
        return error_response("读取满额赠品配置失败", 500)


@app.put("/admin/auto-gifts")
async def admin_update_auto_gifts(payload: AutoGiftUpdateRequest, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        items = payload.items or []
        unique: set = set()
        normalized: List[Dict[str, Optional[str]]] = []
        for item in items:
            key = (item.product_id, item.variant_id or None)
            if key in unique:
                continue
            unique.add(key)
            normalized.append({'product_id': item.product_id, 'variant_id': item.variant_id})
        AutoGiftDB.replace_items(owner_id, normalized)
        refreshed = AutoGiftDB.list_items(owner_id)
        return success_response("满额赠品配置已更新", {"items": refreshed})
    except Exception as e:
        logger.error(f"更新满额赠品配置失败: {e}")
        return error_response("更新满额赠品配置失败", 500)


@app.get("/admin/auto-gifts/search")
async def admin_search_auto_gift_items(request: Request, query: Optional[str] = None, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    try:
        owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
        results = _search_inventory_for_selector(query, staff=admin, owner_override=owner_id)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"搜索满额赠品候选失败: {e}")
        return error_response("搜索满额赠品候选失败", 500)


@app.get("/agent/auto-gifts")
async def agent_get_auto_gifts(request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        items = AutoGiftDB.list_items(owner_id)
        return success_response("获取满额赠品配置成功", {"items": items})
    except Exception as e:
        logger.error(f"代理读取满额赠品配置失败: {e}")
        return error_response("读取满额赠品配置失败", 500)


@app.put("/agent/auto-gifts")
async def agent_update_auto_gifts(payload: AutoGiftUpdateRequest, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        items = payload.items or []
        unique: set = set()
        normalized: List[Dict[str, Optional[str]]] = []
        for item in items:
            key = (item.product_id, item.variant_id or None)
            if key in unique:
                continue
            unique.add(key)
            normalized.append({'product_id': item.product_id, 'variant_id': item.variant_id})
        AutoGiftDB.replace_items(owner_id, normalized)
        refreshed = AutoGiftDB.list_items(owner_id)
        return success_response("满额赠品配置已更新", {"items": refreshed})
    except Exception as e:
        logger.error(f"代理更新满额赠品配置失败: {e}")
        return error_response("更新满额赠品配置失败", 500)


@app.get("/agent/auto-gifts/search")
async def agent_search_auto_gift_items(request: Request, query: Optional[str] = None):
    agent, _ = require_agent_with_scope(request)
    try:
        results = _search_inventory_for_selector(query, staff=agent)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"代理搜索满额赠品候选失败: {e}")
        return error_response("搜索满额赠品候选失败", 500)


@app.get("/auto-gifts")
async def public_get_auto_gifts():
    try:
        # 公共接口需要获取所有可用的自动赠品，不限制owner_id
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM auto_gift_items ORDER BY created_at ASC')
            rows = [dict(r) for r in cursor.fetchall() or []]
            
            if not rows:
                return success_response("获取满额赠品成功", {"items": []})

            product_ids = {row['product_id'] for row in rows if row.get('product_id')}
            variant_ids = {row['variant_id'] for row in rows if row.get('variant_id')}

            product_map = {}
            if product_ids:
                placeholders = ','.join('?' * len(product_ids))
                cursor.execute(f'SELECT * FROM products WHERE id IN ({placeholders})', list(product_ids))
                product_map = {row['id']: dict(row) for row in cursor.fetchall() or []}

            variant_map = {}
            if variant_ids:
                placeholders = ','.join('?' * len(variant_ids))
                cursor.execute(f'SELECT * FROM product_variants WHERE id IN ({placeholders})', list(variant_ids))
                variant_map = {row['id']: dict(row) for row in cursor.fetchall() or []}

            items = []
            for row in rows:
                product_id = row.get('product_id')
                variant_id = row.get('variant_id')
                
                product_info = product_map.get(product_id) if product_id else None
                variant_info = variant_map.get(variant_id) if variant_id else None
                
                if not product_info:
                    continue
                
                # 计算库存和可用性
                if variant_id and variant_info:
                    stock = variant_info.get('stock', 0) or 0
                    product_name = f"{product_info.get('name', '')}（{variant_info.get('name', '')}）"
                else:
                    stock = product_info.get('stock', 0) or 0
                    product_name = product_info.get('name', '')
                
                available = (product_info.get('is_active', 1) == 1) and (stock > 0)
                
                item = {
                    'id': row.get('id'),
                    'product_id': product_id,
                    'variant_id': variant_id,
                    'product_name': product_name,
                    'stock': stock,
                    'available': available,
                    'available_stock': stock
                }
                
                if available:  # 只返回可用的赠品
                    items.append(item)
        
        return success_response("获取满额赠品成功", {"items": items})
    except Exception as e:
        logger.error(f"获取满额赠品失败: {e}")
        return error_response("获取满额赠品失败", 500)


@app.get("/gift-thresholds")
async def public_get_gift_thresholds(request: Request):
    """获取当前配送范围内启用的满额门槛配置"""
    try:
        scope = resolve_shopping_scope(request)
        owner_id = get_owner_id_from_scope(scope)
        thresholds = GiftThresholdDB.list_all(owner_id=owner_id, include_inactive=False)

        simplified_thresholds = []
        for threshold in thresholds:
            available_items = [item for item in threshold.get('items', []) if item.get('available')]
            selected_product_name = ''
            if available_items:
                sorted_items = sorted(available_items, key=lambda x: x.get('stock', 0), reverse=True)
                chosen_item = sorted_items[0]
                name = chosen_item.get('product_name', '')
                if chosen_item.get('variant_name'):
                    name += f"（{chosen_item.get('variant_name')}）"
                selected_product_name = name

            simplified_thresholds.append({
                'threshold_amount': threshold.get('threshold_amount'),
                'gift_products': threshold.get('gift_products', 0) == 1,
                'gift_coupon': threshold.get('gift_coupon', 0) == 1,
                'coupon_amount': threshold.get('coupon_amount', 0),
                'products_count': len(available_items),
                'selected_product_name': selected_product_name
            })

        return success_response("获取满额门槛配置成功", {
            "thresholds": simplified_thresholds,
            "owner_id": owner_id
        })
    except Exception as e:
        logger.error(f"获取满额门槛配置失败: {e}")
        return error_response("获取满额门槛配置失败", 500)


@app.get("/delivery-config")
async def get_delivery_config(request: Request):
    """获取当前配送范围内的配送费配置"""
    try:
        scope = resolve_shopping_scope(request)
        owner_id = get_owner_id_from_scope(scope)
        delivery_config = DeliverySettingsDB.get_delivery_config(owner_id)
        return success_response("获取配送费配置成功", {"delivery_config": delivery_config})
    except Exception as e:
        logger.error(f"获取配送费配置失败: {e}")
        return error_response("获取配送费配置失败", 500)


@app.get("/admin/delivery-settings")
async def admin_get_delivery_settings(request: Request, owner_id: Optional[str] = None):
    """获取配送费设置（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        settings = DeliverySettingsDB.get_settings(owner_id)
        return success_response("获取配送费设置成功", {"settings": settings})
    except Exception as e:
        logger.error(f"获取配送费设置失败: {e}")
        return error_response("获取配送费设置失败", 500)


@app.post("/admin/delivery-settings")
async def admin_create_or_update_delivery_settings(payload: DeliverySettingsCreate, request: Request, owner_id: Optional[str] = None):
    """创建或更新配送费设置（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        if payload.delivery_fee < 0:
            return error_response("配送费不能为负数", 400)

        if payload.free_delivery_threshold < 0:
            return error_response("免配送费门槛不能为负数", 400)

        setting_id = DeliverySettingsDB.create_or_update_settings(
            owner_id=owner_id,
            delivery_fee=payload.delivery_fee,
            free_delivery_threshold=payload.free_delivery_threshold
        )

        settings = DeliverySettingsDB.get_settings(owner_id)
        return success_response("配送费设置保存成功", {"settings": settings})
    except Exception as e:
        logger.error(f"保存配送费设置失败: {e}")
        return error_response("保存配送费设置失败", 500)


@app.get("/agent/delivery-settings")
async def agent_get_delivery_settings(request: Request):
    """获取配送费设置（代理）"""
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        settings = DeliverySettingsDB.get_settings(owner_id)
        return success_response("获取配送费设置成功", {"settings": settings})
    except Exception as e:
        logger.error(f"获取配送费设置失败: {e}")
        return error_response("获取配送费设置失败", 500)


@app.post("/agent/delivery-settings")
async def agent_create_or_update_delivery_settings(payload: DeliverySettingsCreate, request: Request):
    """创建或更新配送费设置（代理）"""
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        if payload.delivery_fee < 0:
            return error_response("配送费不能为负数", 400)

        if payload.free_delivery_threshold < 0:
            return error_response("免配送费门槛不能为负数", 400)

        setting_id = DeliverySettingsDB.create_or_update_settings(
            owner_id=owner_id,
            delivery_fee=payload.delivery_fee,
            free_delivery_threshold=payload.free_delivery_threshold
        )

        settings = DeliverySettingsDB.get_settings(owner_id)
        return success_response("配送费设置保存成功", {"settings": settings})
    except Exception as e:
        logger.error(f"保存配送费设置失败: {e}")
        return error_response("保存配送费设置失败", 500)


@app.post("/admin/lottery-prizes")
async def admin_create_lottery_prize(payload: LotteryPrizeInput, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        prize_id = _persist_lottery_prize_from_payload(payload, owner_id)
        prize = next((p for p in LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True) if p.get('id') == prize_id), None)
        return success_response("抽奖奖项已创建", {"prize": prize})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"创建抽奖奖项失败: {e}")
        return error_response("创建抽奖奖项失败", 500)


@app.post("/agent/lottery-prizes")
async def agent_create_lottery_prize(payload: LotteryPrizeInput, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        prize_id = _persist_lottery_prize_from_payload(payload, owner_id)
        prize = next((p for p in LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True) if p.get('id') == prize_id), None)
        return success_response("抽奖奖项已创建", {"prize": prize})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"代理创建抽奖奖项失败: {e}")
        return error_response("创建抽奖奖项失败", 500)


@app.put("/admin/lottery-prizes/{prize_id}")
async def admin_update_lottery_prize(prize_id: str, payload: LotteryPrizeInput, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        updated_id = _persist_lottery_prize_from_payload(payload, owner_id, override_id=prize_id)
        prize = next((p for p in LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True) if p.get('id') == updated_id), None)
        if not prize:
            return error_response("奖项不存在", 404)
        return success_response("抽奖奖项已更新", {"prize": prize})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"更新抽奖奖项失败: {e}")
        return error_response("更新抽奖奖项失败", 500)


@app.put("/agent/lottery-prizes/{prize_id}")
async def agent_update_lottery_prize(prize_id: str, payload: LotteryPrizeInput, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        updated_id = _persist_lottery_prize_from_payload(payload, owner_id, override_id=prize_id)
        prize = next((p for p in LotteryDB.list_prizes(owner_id=owner_id, include_inactive=True) if p.get('id') == updated_id), None)
        if not prize:
            return error_response("奖项不存在", 404)
        return success_response("抽奖奖项已更新", {"prize": prize})
    except ValueError as ve:
        return error_response(str(ve), 400)
    except Exception as e:
        logger.error(f"代理更新抽奖奖项失败: {e}")
        return error_response("更新抽奖奖项失败", 500)


@app.delete("/admin/lottery-prizes/{prize_id}")
async def admin_delete_lottery_prize(prize_id: str, request: Request, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        ok = LotteryDB.delete_prize(prize_id, owner_id)
        if not ok:
            return error_response("奖项不存在", 404)
        return success_response("抽奖奖项已删除")
    except Exception as e:
        logger.error(f"删除抽奖奖项失败: {e}")
        return error_response("删除抽奖奖项失败", 500)


@app.delete("/agent/lottery-prizes/{prize_id}")
async def agent_delete_lottery_prize(prize_id: str, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        ok = LotteryDB.delete_prize(prize_id, owner_id)
        if not ok:
            return error_response("奖项不存在", 404)
        return success_response("抽奖奖项已删除")
    except Exception as e:
        logger.error(f"代理删除抽奖奖项失败: {e}")
        return error_response("删除抽奖奖项失败", 500)


@app.get("/admin/lottery-prizes/search")
async def admin_search_lottery_prize_items(request: Request, query: Optional[str] = None, owner_id: Optional[str] = None):
    admin = get_current_admin_required_from_cookie(request)
    try:
        owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
        results = _search_inventory_for_selector(query, staff=admin, owner_override=owner_id)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"搜索抽奖候选商品失败: {e}")
        return error_response("搜索抽奖候选商品失败", 500)


@app.get("/agent/lottery-prizes/search")
async def agent_search_lottery_prize_items(request: Request, query: Optional[str] = None):
    agent, _ = require_agent_with_scope(request)
    try:
        results = _search_inventory_for_selector(query, staff=agent)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"代理搜索抽奖候选商品失败: {e}")
        return error_response("搜索抽奖候选商品失败", 500)


# ==================== 满额门槛配置管理（管理员） ====================

@app.get("/admin/gift-thresholds")
async def admin_get_gift_thresholds(request: Request, include_inactive: bool = False, owner_id: Optional[str] = None):
    """获取满额门槛配置列表（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        thresholds = GiftThresholdDB.list_all(owner_id=owner_id, include_inactive=include_inactive)
        return success_response("获取满额门槛配置成功", {"thresholds": thresholds})
    except Exception as e:
        logger.error(f"获取满额门槛配置失败: {e}")
        return error_response("获取满额门槛配置失败", 500)


@app.get("/agent/gift-thresholds")
async def agent_get_gift_thresholds(request: Request, include_inactive: bool = False):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        thresholds = GiftThresholdDB.list_all(owner_id=owner_id, include_inactive=include_inactive)
        return success_response("获取满额门槛配置成功", {"thresholds": thresholds})
    except Exception as e:
        logger.error(f"代理获取满额门槛配置失败: {e}")
        return error_response("获取满额门槛配置失败", 500)


@app.post("/admin/gift-thresholds")
async def admin_create_gift_threshold(payload: GiftThresholdCreate, request: Request, owner_id: Optional[str] = None):
    """创建满额门槛配置（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        if payload.threshold_amount <= 0:
            return error_response("门槛金额必须大于0", 400)

        if payload.gift_coupon and payload.coupon_amount <= 0:
            return error_response("优惠券金额必须大于0", 400)

        normalized_limit = None
        if payload.per_order_limit is not None:
            raw_limit = int(payload.per_order_limit)
            if raw_limit < 0:
                return error_response("每单赠品上限必须为正整数或留空", 400)
            normalized_limit = normalize_per_order_limit(raw_limit)

        # 创建门槛配置
        threshold_id = GiftThresholdDB.create_threshold(
            owner_id=owner_id,
            threshold_amount=payload.threshold_amount,
            gift_products=payload.gift_products,
            gift_coupon=payload.gift_coupon,
            coupon_amount=payload.coupon_amount if payload.gift_coupon else 0.0,
            per_order_limit=normalized_limit
        )

        # 添加商品到门槛
        if payload.items and payload.gift_products:
            items_data = []
            for item in payload.items:
                if item.product_id:
                    items_data.append({
                        'product_id': item.product_id,
                        'variant_id': item.variant_id
                    })
            if items_data:
                GiftThresholdDB.add_items_to_threshold(threshold_id, owner_id, items_data)

        threshold = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        return success_response("满额门槛配置创建成功", {"threshold": threshold})
    except Exception as e:
        logger.error(f"创建满额门槛配置失败: {e}")
        return error_response("创建满额门槛配置失败", 500)


@app.post("/agent/gift-thresholds")
async def agent_create_gift_threshold(payload: GiftThresholdCreate, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        if payload.threshold_amount <= 0:
            return error_response("门槛金额必须大于0", 400)

        if payload.gift_coupon and payload.coupon_amount <= 0:
            return error_response("优惠券金额必须大于0", 400)

        normalized_limit = None
        if payload.per_order_limit is not None:
            raw_limit = int(payload.per_order_limit)
            if raw_limit < 0:
                return error_response("每单赠品上限必须为正整数或留空", 400)
            normalized_limit = normalize_per_order_limit(raw_limit)

        threshold_id = GiftThresholdDB.create_threshold(
            owner_id=owner_id,
            threshold_amount=payload.threshold_amount,
            gift_products=payload.gift_products,
            gift_coupon=payload.gift_coupon,
            coupon_amount=payload.coupon_amount if payload.gift_coupon else 0.0,
            per_order_limit=normalized_limit
        )

        if payload.items and payload.gift_products:
            items_data = []
            for item in payload.items:
                if item.product_id:
                    items_data.append({
                        'product_id': item.product_id,
                        'variant_id': item.variant_id
                    })
            if items_data:
                GiftThresholdDB.add_items_to_threshold(threshold_id, owner_id, items_data)

        threshold = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        return success_response("满额门槛配置创建成功", {"threshold": threshold})
    except Exception as e:
        logger.error(f"代理创建满额门槛配置失败: {e}")
        return error_response("创建满额门槛配置失败", 500)


@app.put("/admin/gift-thresholds/{threshold_id}")
async def admin_update_gift_threshold(threshold_id: str, payload: GiftThresholdUpdate, request: Request, owner_id: Optional[str] = None):
    """更新满额门槛配置（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        existing = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        if not existing:
            return error_response("门槛配置不存在", 404)
        
        if payload.threshold_amount is not None and payload.threshold_amount <= 0:
            return error_response("门槛金额必须大于0", 400)
        
        if payload.gift_coupon and payload.coupon_amount is not None and payload.coupon_amount <= 0:
            return error_response("优惠券金额必须大于0", 400)

        per_order_limit_param = None
        if 'per_order_limit' in payload.__fields_set__:
            raw_limit_value = payload.per_order_limit
            if raw_limit_value is None:
                per_order_limit_param = 0
            else:
                raw_limit = int(raw_limit_value)
                if raw_limit < 0:
                    return error_response("每单赠品上限必须为正整数或留空", 400)
                normalized_limit = normalize_per_order_limit(raw_limit)
                per_order_limit_param = normalized_limit if normalized_limit is not None else 0
        
        # 更新基础配置
        GiftThresholdDB.update_threshold(
            threshold_id=threshold_id,
            owner_id=owner_id,
            threshold_amount=payload.threshold_amount,
            gift_products=payload.gift_products,
            gift_coupon=payload.gift_coupon,
            coupon_amount=payload.coupon_amount,
            per_order_limit=per_order_limit_param,
            is_active=payload.is_active
        )

        # 更新商品列表
        if payload.items is not None:
            items_data = []
            for item in payload.items:
                if item.product_id:
                    items_data.append({
                        'product_id': item.product_id,
                        'variant_id': item.variant_id
                    })
            GiftThresholdDB.add_items_to_threshold(threshold_id, owner_id, items_data)

        threshold = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        return success_response("满额门槛配置更新成功", {"threshold": threshold})
    except Exception as e:
        logger.error(f"更新满额门槛配置失败: {e}")
        return error_response("更新满额门槛配置失败", 500)


@app.put("/agent/gift-thresholds/{threshold_id}")
async def agent_update_gift_threshold(threshold_id: str, payload: GiftThresholdUpdate, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        existing = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        if not existing:
            return error_response("门槛配置不存在", 404)

        if payload.threshold_amount is not None and payload.threshold_amount <= 0:
            return error_response("门槛金额必须大于0", 400)

        if payload.gift_coupon and payload.coupon_amount is not None and payload.coupon_amount <= 0:
            return error_response("优惠券金额必须大于0", 400)

        per_order_limit_param = None
        if 'per_order_limit' in payload.__fields_set__:
            raw_limit_value = payload.per_order_limit
            if raw_limit_value is None:
                per_order_limit_param = 0
            else:
                raw_limit = int(raw_limit_value)
                if raw_limit < 0:
                    return error_response("每单赠品上限必须为正整数或留空", 400)
                normalized_limit = normalize_per_order_limit(raw_limit)
                per_order_limit_param = normalized_limit if normalized_limit is not None else 0

        GiftThresholdDB.update_threshold(
            threshold_id=threshold_id,
            owner_id=owner_id,
            threshold_amount=payload.threshold_amount,
            gift_products=payload.gift_products,
            gift_coupon=payload.gift_coupon,
            coupon_amount=payload.coupon_amount,
            per_order_limit=per_order_limit_param,
            is_active=payload.is_active
        )

        if payload.items is not None:
            items_data = []
            for item in payload.items:
                if item.product_id:
                    items_data.append({
                        'product_id': item.product_id,
                        'variant_id': item.variant_id
                    })
            GiftThresholdDB.add_items_to_threshold(threshold_id, owner_id, items_data)

        threshold = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        return success_response("满额门槛配置更新成功", {"threshold": threshold})
    except Exception as e:
        logger.error(f"代理更新满额门槛配置失败: {e}")
        return error_response("更新满额门槛配置失败", 500)


@app.delete("/admin/gift-thresholds/{threshold_id}")
async def admin_delete_gift_threshold(threshold_id: str, request: Request, owner_id: Optional[str] = None):
    """删除满额门槛配置（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
    try:
        existing = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        if not existing:
            return error_response("门槛配置不存在", 404)

        success = GiftThresholdDB.delete_threshold(threshold_id, owner_id)
        if not success:
            return error_response("删除满额门槛配置失败", 500)

        return success_response("满额门槛配置删除成功")
    except Exception as e:
        logger.error(f"删除满额门槛配置失败: {e}")
        return error_response("删除满额门槛配置失败", 500)


@app.delete("/agent/gift-thresholds/{threshold_id}")
async def agent_delete_gift_threshold(threshold_id: str, request: Request):
    agent, _ = require_agent_with_scope(request)
    owner_id = get_owner_id_for_staff(agent)
    try:
        existing = GiftThresholdDB.get_by_id(threshold_id, owner_id)
        if not existing:
            return error_response("门槛配置不存在", 404)

        success = GiftThresholdDB.delete_threshold(threshold_id, owner_id)
        if not success:
            return error_response("删除满额门槛配置失败", 500)

        return success_response("满额门槛配置删除成功")
    except Exception as e:
        logger.error(f"代理删除满额门槛配置失败: {e}")
        return error_response("删除满额门槛配置失败", 500)


@app.get("/admin/gift-thresholds/search")
async def admin_search_gift_threshold_items(request: Request, query: Optional[str] = None, owner_id: Optional[str] = None):
    """搜索满额门槛赠品候选商品（管理员）"""
    admin = get_current_admin_required_from_cookie(request)
    try:
        owner_id, _ = resolve_single_owner_for_staff(admin, owner_id)
        results = _search_inventory_for_selector(query, staff=admin, owner_override=owner_id)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"搜索满额门槛赠品候选失败: {e}")
        return error_response("搜索满额门槛赠品候选失败", 500)


@app.get("/agent/gift-thresholds/search")
async def agent_search_gift_threshold_items(request: Request, query: Optional[str] = None):
    """搜索满额门槛赠品候选商品（代理）"""
    agent, _ = require_agent_with_scope(request)
    try:
        results = _search_inventory_for_selector(query, staff=agent)
        return success_response("搜索成功", {"items": results})
    except Exception as e:
        logger.error(f"代理搜索满额门槛赠品候选失败: {e}")
        return error_response("搜索满额门槛赠品候选失败", 500)


class OrderDeleteRequest(BaseModel):
    order_ids: Optional[List[str]] = None

@app.delete("/admin/orders/{order_id}")
async def admin_delete_orders(order_id: str, request: Request, delete_request: Optional[OrderDeleteRequest] = None):
    """删除订单（支持单个或批量）"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        scope = build_staff_scope(staff)
        if delete_request and delete_request.order_ids:
            ids = delete_request.order_ids
            if len(ids) > 500:
                return error_response("批量删除数量不能超过500笔订单", 400)
            from database import OrderDB
            # 预取订单，用于返还锁定的优惠券
            try:
                orders_before = [OrderDB.get_order_by_id(i) for i in ids]
            except Exception:
                orders_before = []
            accessible_ids: List[str] = []
            for od in orders_before:
                if not od:
                    continue
                if not staff_can_access_order(staff, od, scope):
                    return error_response("无权删除部分订单", 403)
                accessible_ids.append(od['id'])
            if not accessible_ids:
                return success_response("未找到可删除的订单", {
                    "deleted_count": 0,
                    "deleted_ids": [],
                    "not_found_ids": ids
                })
            result = OrderDB.batch_delete_orders(accessible_ids)
            if not result.get("success"):
                return error_response(result.get("message", "批量删除失败"), 400)
            # 恢复库存和返还相关优惠券
            try:
                for od in (orders_before or []):
                    if not od:
                        continue
                    try:
                        payment_status = od.get("payment_status") or "pending"
                        
                        # 如果是已支付订单，恢复库存
                        if payment_status == "succeeded":
                            try:
                                restore_ok = OrderDB.restore_stock_from_order(od.get("id"))
                                if not restore_ok:
                                    logger.warning(f"批量删除时恢复库存失败: order_id={od.get('id')}")
                            except Exception as e:
                                logger.warning(f"批量删除时恢复库存异常: {e}")
                        
                        # 返还优惠券（仅对未支付订单）
                        if payment_status != "succeeded":
                            c_id = od.get("coupon_id")
                            d_amt = float(od.get("discount_amount") or 0)
                            if c_id and d_amt > 0:
                                CouponDB.unlock_for_order(c_id, od.get("id"))
                    except Exception:
                        pass
            except Exception as e:
                logger.warning(f"批量删除后处理失败: {e}")
            return success_response(result.get("message", "批量删除成功"), result)
        else:
            from database import OrderDB
            # 单笔：删除前恢复库存和返还优惠券
            try:
                od = OrderDB.get_order_by_id(order_id)
                if not staff_can_access_order(staff, od, scope):
                    return error_response("无权删除此订单", 403)
                
                if od:
                    payment_status = od.get("payment_status") or "pending"
                    
                    # 如果是已支付订单，恢复库存
                    if payment_status == "succeeded":
                        try:
                            restore_ok = OrderDB.restore_stock_from_order(order_id)
                            if not restore_ok:
                                logger.warning(f"删除订单时恢复库存失败: order_id={order_id}")
                        except Exception as e:
                            logger.warning(f"删除订单时恢复库存异常: {e}")
                    
                    # 返还优惠券（仅未支付订单）
                    if payment_status != "succeeded":
                        c_id = od.get("coupon_id")
                        d_amt = float(od.get("discount_amount") or 0)
                        if c_id and d_amt > 0:
                            CouponDB.unlock_for_order(c_id, order_id)
            except Exception as e:
                logger.warning(f"单笔删除前处理失败: {e}")
            ok = OrderDB.delete_order(order_id)
            if not ok:
                return error_response("删除订单失败或订单不存在", 400)
            return success_response("订单删除成功")
    except Exception as e:
        logger.error(f"删除订单失败: {e}")
        return error_response("删除订单失败", 500)

@app.patch("/admin/orders/{order_id}/status")
async def update_order_status(
    order_id: str,
    status_request: OrderStatusUpdateRequest,
    request: Request
):
    """更新订单状态（管理员）"""
    staff = get_current_staff_required_from_cookie(request)

    try:
        # 验证状态值
        valid_statuses = ['pending', 'confirmed', 'shipped', 'delivered', 'cancelled']
        if status_request.status not in valid_statuses:
            return error_response("无效的订单状态", 400)
        
        # 检查订单是否存在
        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        scope = build_staff_scope(staff)
        if not staff_can_access_order(staff, order, scope):
            return error_response("无权操作该订单", 403)

        # 更新状态
        success = OrderDB.update_order_status(order_id, status_request.status)
        if not success:
            return error_response("更新订单状态失败", 500)
        
        return success_response("订单状态更新成功", {"order_id": order_id, "new_status": status_request.status})
    
    except Exception as e:
        logger.error(f"更新订单状态失败: {e}")
        return error_response("更新订单状态失败", 500)

@app.get("/admin/order-stats")
async def get_order_statistics(request: Request, agent_id: Optional[str] = None):
    """获取订单统计信息（管理员）"""
    staff = get_current_staff_required_from_cookie(request)

    try:
        scope = build_staff_scope(staff)
        (
            selected_agent_id,
            selected_address_ids,
            selected_building_ids,
            exclude_address_ids,
            exclude_building_ids,
            resolved_filter
        ) = resolve_staff_order_scope(staff, scope, agent_id)

        stats = OrderDB.get_order_stats(
            agent_id=selected_agent_id,
            address_ids=selected_address_ids,
            building_ids=selected_building_ids,
            exclude_address_ids=exclude_address_ids,
            exclude_building_ids=exclude_building_ids
        )
        stats["scope"] = scope
        stats["selected_agent_filter"] = resolved_filter
        return success_response("获取订单统计成功", stats)

    except Exception as e:
        logger.error(f"获取订单统计失败: {e}")
        return error_response("获取订单统计失败", 500)

@app.get("/admin/dashboard-stats")
async def get_dashboard_statistics(
    request: Request, 
    period: str = 'week',
    range_start: Optional[str] = None,
    range_end: Optional[str] = None,
    agent_id: Optional[str] = None
):
    """获取仪表盘详细统计信息（管理员）"""
    staff = get_current_staff_required_from_cookie(request)

    try:
        if period not in ['day', 'week', 'month']:
            period = 'week'

        scope = build_staff_scope(staff)
        (
            selected_agent_id,
            selected_address_ids,
            selected_building_ids,
            _,
            _,
            selected_filter
        ) = resolve_staff_order_scope(staff, scope, agent_id)
        filter_admin_orders = scope.get('filter_admin_orders', False)
        if selected_filter and selected_filter != 'self':
            filter_admin_orders = False

        stats = OrderDB.get_dashboard_stats(
            period,
            agent_id=selected_agent_id,
            address_ids=selected_address_ids,
            building_ids=selected_building_ids,
            filter_admin_orders=filter_admin_orders,
            top_range_start=range_start,
            top_range_end=range_end
        )
        stats["scope"] = scope
        stats["selected_agent_id"] = selected_agent_id
        return success_response("获取仪表盘统计成功", stats)
    
    except Exception as e:
        logger.error(f"获取仪表盘统计失败: {e}")
        return error_response("获取仪表盘统计失败", 500)


@app.get("/agent/dashboard-stats")
async def get_agent_dashboard_statistics(
    request: Request, 
    period: str = 'week',
    range_start: Optional[str] = None,
    range_end: Optional[str] = None
):
    """获取仪表盘详细统计信息（代理）"""
    _agent, scope = require_agent_with_scope(request)

    try:
        if period not in ['day', 'week', 'month']:
            period = 'week'

        stats = OrderDB.get_dashboard_stats(
            period,
            agent_id=scope.get('agent_id'),
            address_ids=scope.get('address_ids'),
            building_ids=scope.get('building_ids'),
            top_range_start=range_start,
            top_range_end=range_end
        )
        stats["scope"] = scope
        return success_response("获取仪表盘统计成功", stats)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"代理获取仪表盘统计失败: {e}")
        return error_response("获取仪表盘统计失败", 500)

@app.get("/admin/customers")
async def get_customers_with_purchases(request: Request, limit: Optional[int] = 5, offset: Optional[int] = 0, agent_id: Optional[str] = None):
    """获取购买过商品的客户列表（管理员）"""
    staff = get_current_staff_required_from_cookie(request)

    try:
        # 参数验证和限制
        try:
            limit_val = int(limit or 5)
        except Exception:
            limit_val = 5
        if limit_val <= 0:
            limit_val = 5
        if limit_val > 50:  # 限制单次最多返回50个
            limit_val = 50
            
        try:
            offset_val = int(offset or 0)
        except Exception:
            offset_val = 0
        if offset_val < 0:
            offset_val = 0

        scope = build_staff_scope(staff)
        (
            selected_agent_id,
            selected_address_ids,
            selected_building_ids,
            _,
            _,
            selected_filter
        ) = resolve_staff_order_scope(staff, scope, agent_id)
        filter_admin_orders = scope.get('filter_admin_orders', False)
        if selected_filter and selected_filter != 'self':
            filter_admin_orders = False
        customers_data = OrderDB.get_customers_with_purchases(
            limit=limit_val,
            offset=offset_val,
            agent_id=selected_agent_id,
            address_ids=selected_address_ids,
            building_ids=selected_building_ids,
            filter_admin_orders=filter_admin_orders
        )
        customers_data['scope'] = scope
        return success_response("获取客户列表成功", customers_data)
    
    except Exception as e:
        logger.error(f"获取客户列表失败: {e}")
        return error_response("获取客户列表失败", 500)

# ==================== AI聊天路由 ====================

from ai_chat import stream_chat


def _serialize_chat_thread(thread: Dict[str, Any]) -> Dict[str, Any]:
    if not thread:
        return {}
    
    # 返回完整的自定义标题（如果有）和预览（前8个字符）
    # 前端负责显示逻辑：
    # - 如果有自定义标题且超过8个字符，显示前7个字符+"..."
    # - 否则显示完整标题或预览
    custom_title = (thread.get("title") or "").strip()
    preview = (thread.get("first_message_preview") or "").strip()
    
    return {
        "id": thread.get("id"),
        "title": custom_title if custom_title else None,  # 完整的自定义标题或null
        "preview": preview[:8] if preview else None,  # 消息预览的前8个字符
        "created_at": thread.get("created_at"),
        "updated_at": thread.get("updated_at"),
        "last_message_at": thread.get("last_message_at"),
        "is_archived": bool(thread.get("is_archived")),
    }


def _serialize_chat_message(record: Dict[str, Any]) -> Dict[str, Any]:
    if not record:
        return {}
    content = record.get("content")
    payload = {
        "id": record.get("id"),
        "role": record.get("role"),
        "content": content,
        "raw_content": content,
        "timestamp": record.get("timestamp"),
        "thread_id": record.get("thread_id"),
        "tool_call_id": record.get("tool_call_id"),
    }
    # 对于 assistant 角色，尝试解析 JSON 格式的内容（可能包含 tool_calls）
    if payload["role"] == "assistant":
        payload["thinking_content"] = ""
        if isinstance(content, str):
            try:
                parsed = json.loads(content)
                if isinstance(parsed, dict):
                    # 提取 tool_calls（如果存在）
                    if "tool_calls" in parsed:
                        payload["tool_calls"] = parsed.get("tool_calls")

                    # 提取或重置 content 字段
                    if "content" in parsed:
                        # 新格式：JSON 中包含 content 字段
                        payload["content"] = parsed.get("content") or ""
                    elif "tool_calls" in parsed:
                        # 旧格式：只有 tool_calls，没有 content 字段
                        payload["content"] = ""
                    # 解析思维链内容
                    thinking_value = parsed.get("thinking_content")
                    if thinking_value is None:
                        thinking_value = ""
                    if isinstance(thinking_value, str):
                        payload["thinking_content"] = thinking_value
                    else:
                        payload["thinking_content"] = str(thinking_value)
            except Exception:
                # 不是 JSON 格式，保持原始内容
                pass
    return payload


@app.get("/ai/models")
async def list_ai_models():
    """返回可用模型列表及其能力，用于前端渲染模型选择器。"""
    configs = get_settings().model_order
    logger.info(f"/ai/models API调用 - 配置中的模型数量: {len(configs)}")
    logger.info(f"/ai/models API调用 - 配置中的模型列表: {[(cfg.name, cfg.label) for cfg in configs]}")
    result = {
        "models": [
            {
                "model": cfg.name,
                "name": cfg.label,
                "supports_thinking": cfg.supports_thinking,
            }
            for cfg in configs
        ]
    }
    logger.info(f"/ai/models API调用 - 返回结果中的模型数量: {len(result['models'])}")
    return result


@app.get("/ai/chats")
async def list_chat_history(request: Request, limit: int = 100):
    """列出当前用户的聊天会话"""
    user = get_current_user_required_from_cookie(request)
    try:
        safe_limit = max(1, min(limit, 200))
        threads = ChatLogDB.list_threads(user['id'], limit=safe_limit)
        return {"chats": [_serialize_chat_thread(thread) for thread in threads]}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"获取聊天历史失败: {exc}")
        raise HTTPException(status_code=500, detail="无法获取聊天历史")


@app.post("/ai/chats")
async def create_chat_history(payload: ChatThreadCreateRequest, request: Request):
    """创建新的聊天会话"""
    user = get_current_user_required_from_cookie(request)
    try:
        thread = ChatLogDB.create_thread(user['id'], title=payload.title)
        return {"chat": _serialize_chat_thread(thread)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.error(f"创建聊天会话失败: {exc}")
        raise HTTPException(status_code=500, detail="创建聊天会话失败")


@app.get("/ai/chats/{chat_id}")
async def get_chat_history(chat_id: str, request: Request):
    """获取指定聊天会话及其消息"""
    user = get_current_user_required_from_cookie(request)
    try:
        thread = ChatLogDB.get_thread_for_user(user['id'], chat_id)
        if not thread:
            raise HTTPException(status_code=401, detail="无权访问该会话")
        messages = ChatLogDB.get_thread_messages(user['id'], chat_id, limit=800)
        return {
            "chat": _serialize_chat_thread(thread),
            "messages": [_serialize_chat_message(msg) for msg in messages]
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"获取聊天会话失败: {exc}")
        raise HTTPException(status_code=500, detail="获取聊天会话失败")


@app.patch("/ai/chats/{chat_id}")
async def rename_chat_history(chat_id: str, payload: ChatThreadUpdateRequest, request: Request):
    """重命名聊天会话"""
    user = get_current_user_required_from_cookie(request)
    try:
        updated = ChatLogDB.rename_thread(user['id'], chat_id, payload.title or "")
        if not updated:
            raise HTTPException(status_code=401, detail="无权更新该会话")
        thread = ChatLogDB.get_thread_for_user(user['id'], chat_id)
        return {"chat": _serialize_chat_thread(thread)}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"更新聊天会话失败: {exc}")
        raise HTTPException(status_code=500, detail="更新聊天会话失败")

@app.post("/ai/chat")
async def ai_chat(
    request: ChatRequest,
    http_request: Request
):
    """AI聊天接口（支持未登录用户，但功能受限）"""
    try:
        # 尝试从Cookie获取用户信息，允许未登录
        user = None
        try:
            user = get_current_user_from_cookie(http_request)
            logger.info(f"AI聊天请求 - 用户ID: {user['id'] if user else 'anonymous'}")
        except Exception as e:
            logger.info(f"AI聊天请求 - 用户未登录: {e}")
        
        # 验证并解析会话ID
        conversation_id = (request.conversation_id or "").strip() or None
        if conversation_id and not user:
            raise HTTPException(status_code=401, detail="需要登录才能访问指定对话")
        if user:
            if not conversation_id:
                raise HTTPException(status_code=400, detail="缺少会话ID")
            thread = ChatLogDB.get_thread_for_user(user['id'], conversation_id)
            if not thread:
                raise HTTPException(status_code=401, detail="无权访问该会话")
        else:
            conversation_id = None

        # 转换消息，保留所有必要字段（role, content, tool_calls, tool_call_id）
        messages = []
        for msg in request.messages:
            message_dict = {"role": msg.role, "content": msg.content}
            # 保留 tool_calls 和 tool_call_id（如果存在）
            if msg.tool_calls is not None:
                message_dict["tool_calls"] = msg.tool_calls
            if msg.tool_call_id is not None:
                message_dict["tool_call_id"] = msg.tool_call_id
            messages.append(message_dict)
        
        selected_model = (request.model or "").strip()
        return await stream_chat(user, messages, http_request, selected_model, conversation_id)
    except Exception as e:
        logger.error(f"AI聊天失败: {e}")
        raise HTTPException(status_code=500, detail="AI聊天服务暂时不可用")

 # ==================== 微信扫码支付相关（人工确认） ====================

@app.post("/orders/{order_id}/mark-paid")
async def mark_order_paid_pending(order_id: str, request: Request):
    """用户扫码后手动标记为待验证（processing）"""
    user = get_current_user_required_from_cookie(request)
    try:
        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        if order["student_id"] != user["id"]:
            return error_response("无权操作此订单", 403)

        current = order.get("payment_status") or "pending"
        if current == "succeeded":
            return error_response("订单已支付，无需重复操作", 400)
        if current == "processing":
            return success_response("订单已处于待验证状态")

        # 允许从 pending/failed 进入 processing
        if current not in ["pending", "failed"]:
            return error_response("当前订单支付状态不允许此操作", 400)

        ok = OrderDB.update_payment_status(order_id, "processing")
        if not ok:
            return error_response("更新订单支付状态失败", 500)
        return success_response("已标记为待验证", {"order_id": order_id, "payment_status": "processing"})
    except Exception as e:
        logger.error(f"用户标记订单待验证失败: {e}")
        return error_response("操作失败", 500)

# ==================== 抽奖功能 ====================

@app.post("/orders/{order_id}/lottery/draw")
async def draw_lottery(order_id: str, request: Request):
    """订单点击“已付款”后触发抽奖（订单商品金额满10元；每单一次）。"""
    user = get_current_user_required_from_cookie(request)
    try:
        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        if order["student_id"] != user["id"]:
            return error_response("无权操作此订单", 403)

        # 仅计算非抽奖项的小计（老订单不会有 is_lottery 字段，默认参与计算）
        items = order.get("items") or []
        items_subtotal = 0.0
        for it in items:
            if isinstance(it, dict) and (it.get("is_lottery") or it.get("is_auto_gift")):
                continue
            try:
                items_subtotal += float(it.get("subtotal", 0) or 0)
            except Exception:
                pass
        owner_id = LotteryConfigDB.normalize_owner(order.get('agent_id'))
        lottery_enabled = LotteryConfigDB.get_enabled(owner_id)
        if not lottery_enabled:
            return error_response("抽奖功能已禁用", 400)
        threshold_amount = LotteryConfigDB.get_threshold(owner_id)
        if items_subtotal < threshold_amount:
            return error_response(f"本单商品金额未满{threshold_amount:.2f}元，不参与抽奖", 400)

        # 每个订单仅允许一次抽奖
        existing = LotteryDB.get_draw_by_order(order_id)
        if existing:
            prize_name = existing.get("prize_name")
            prize_detail = None
            if prize_name and prize_name != "谢谢参与":
                prize_detail = {
                    "display_name": prize_name,
                    "product_id": existing.get("prize_product_id"),
                    "product_name": existing.get("prize_product_name"),
                    "variant_id": existing.get("prize_variant_id"),
                    "variant_name": existing.get("prize_variant_name"),
                    "group_id": existing.get("prize_group_id"),
                }
            return success_response("抽奖已完成", {
                "prize_name": prize_name,
                "already_drawn": True,
                "names": [prize_name] if prize_name else [],
                "prize": prize_detail,
                "threshold_amount": threshold_amount
            })

        prize_groups = LotteryDB.get_active_prizes_for_draw(owner_id)
        names = [p.get("display_name") for p in prize_groups if p.get("display_name")]
        weights = [max(0.0, float(p.get("weight") or 0)) for p in prize_groups]
        sum_w = sum(weights)
        is_fraction = sum_w <= 1.000001
        scale = 1.0 if is_fraction else 100.0
        # 剩余“谢谢参与”概率
        leftover = max(0.0, scale - sum_w)
        total_w = sum_w + leftover
        if total_w <= 0:
            return error_response("抽奖配置权重无效", 500)

        rnd = random.random() * total_w
        acc = 0.0
        selected_group = None
        for group, weight in zip(prize_groups, weights):
            if weight <= 0:
                continue
            acc += weight
            if rnd <= acc:
                selected_group = group
                break

        selected_item = None
        if selected_group:
            available_items = [item for item in selected_group.get("items", []) if item.get("available")]
            total_stock = sum(max(0, int(item.get("stock") or 0)) for item in available_items)
            if total_stock > 0:
                # 选择库存数量最多的商品，而不是随机选择
                max_stock = 0
                selected_item = None
                for item in available_items:
                    stock_val = max(0, int(item.get("stock") or 0))
                    if stock_val > max_stock:
                        max_stock = stock_val
                        selected_item = item
            if not selected_item:
                selected_group = None

        prize_payload = None
        prize_product_id = None
        prize_variant_id = None
        prize_product_name = None
        prize_variant_name = None
        prize_unit_price = 0.0
        prize_group_id = None

        if selected_group is None:
            selected_name = "谢谢参与"
        else:
            selected_name = selected_group.get("display_name") or ""
            prize_product_id = selected_item.get("product_id")
            prize_variant_id = selected_item.get("variant_id")
            prize_product_name = selected_item.get("product_name")
            prize_variant_name = selected_item.get("variant_name")
            try:
                prize_unit_price = float(selected_item.get("retail_price") or 0.0)
            except Exception:
                prize_unit_price = 0.0
            prize_group_id = selected_group.get("id")
            prize_payload = {
                "display_name": selected_name,
                "product_id": prize_product_id,
                "product_name": prize_product_name,
                "variant_id": prize_variant_id,
                "variant_name": prize_variant_name,
                "group_id": prize_group_id
            }

        LotteryDB.create_draw(
            order_id,
            user["id"],
            selected_name,
            prize_product_id,
            1,
            owner_id=owner_id,
            prize_group_id=prize_group_id,
            prize_product_name=prize_product_name,
            prize_variant_id=prize_variant_id,
            prize_variant_name=prize_variant_name,
            prize_unit_price=prize_unit_price
        )

        thanks_prob_percent = (leftover / scale) * 100.0 if total_w > 0 else 0.0
        return success_response("抽奖完成", {
            "prize_name": selected_name,
            "already_drawn": False,
            "names": names,
            "thanks_probability": round(thanks_prob_percent, 2),
            "prize": prize_payload,
            "threshold_amount": threshold_amount
        })
    except Exception as e:
        logger.error(f"抽奖失败: {e}")
        return error_response("抽奖失败", 500)

@app.get("/rewards/eligible")
async def get_eligible_rewards(
    request: Request,
    owner_id: Optional[str] = None,
    restrict_owner: Optional[bool] = False
):
    """获取当前用户可用（未消费）的抽奖奖品列表"""
    user = get_current_user_required_from_cookie(request)
    try:
        normalized_owner: Optional[str]
        if owner_id is None:
            normalized_owner = None
        else:
            value = owner_id.strip()
            normalized_owner = None if value.lower() in {"", "none", "null", "undefined"} else value

        if isinstance(restrict_owner, bool):
            restrict_flag = restrict_owner
        else:
            restrict_flag = str(restrict_owner).strip().lower() in {"1", "true", "yes"}

        rewards = RewardDB.get_eligible_rewards(
            user["id"],
            normalized_owner,
            restrict_owner=restrict_flag
        ) or []
        return success_response("获取奖品成功", {"rewards": rewards})
    except Exception as e:
        logger.error(f"获取奖品失败: {e}")
        return error_response("获取奖品失败", 500)

@app.patch("/admin/orders/{order_id}/payment-status")
async def admin_update_payment_status(order_id: str, payload: PaymentStatusUpdateRequest, request: Request):
    """管理员更新订单支付状态：pending/processing/succeeded/failed"""
    staff = get_current_staff_required_from_cookie(request)
    try:
        new_status = payload.payment_status
        # 允许管理员将支付状态设置为 pending/processing/succeeded/failed
        if new_status not in ["pending", "processing", "succeeded", "failed"]:
            return error_response("无效的支付状态", 400)

        order = OrderDB.get_order_by_id(order_id)
        if not order:
            return error_response("订单不存在", 404)
        scope = build_staff_scope(staff)
        if not staff_can_access_order(staff, order, scope):
            return error_response("无权操作该订单", 403)

        # 成功付款：扣减库存、更新状态、清空购物车
        if new_status == "succeeded":
            ok, missing_items = OrderDB.complete_payment_and_update_stock(order_id)
            if not ok:
                message = "处理支付成功失败，可能库存不足或状态异常"
                details = {}
                if missing_items:
                    message = f"以下商品库存不足：{'、'.join(missing_items)}"
                    details["out_of_stock_items"] = missing_items
                return error_response(message, 400, details)
            order_owner_id = LotteryConfigDB.normalize_owner(order.get('agent_id'))
            try:
                # 清空该用户购物车
                CartDB.update_cart(order["student_id"], {})  # student_id会自动解析为user_id
                # 缓存用户最近的收货信息
                if isinstance(order.get("shipping_info"), dict):
                    try:
                        UserProfileDB.upsert_shipping(order["student_id"], order["shipping_info"])  # student_id会自动解析为user_id
                    except Exception as e:
                        logger.warning(f"缓存用户收货信息失败: {e}")
                # 若该订单曾进行抽奖，确认成功后生成可用奖品（排除谢谢参与）
                try:
                    draw = LotteryDB.get_draw_by_order(order_id)
                    if draw and draw.get("prize_name") != "谢谢参与":
                        RewardDB.add_reward_from_order(
                            user_identifier=order["student_id"],  # student_id会自动解析为user_id
                            prize_name=draw.get("prize_name"),
                            prize_product_id=draw.get("prize_product_id"),
                            quantity=int(draw.get("prize_quantity") or 1),
                            source_order_id=order_id,
                            owner_id=order_owner_id,
                            prize_group_id=draw.get("prize_group_id"),
                            prize_product_name=draw.get("prize_product_name"),
                            prize_variant_id=draw.get("prize_variant_id"),
                            prize_variant_name=draw.get("prize_variant_name"),
                            prize_unit_price=draw.get("prize_unit_price")
                        )
                except Exception as e:
                    logger.warning(f"生成抽奖奖品失败: {e}")
                # 发放满额优惠券（根据订单金额重新计算）
                try:
                    # 计算订单商品小计（不含运费、不含抽奖奖品和满额赠品）
                    items = order.get("items") or []
                    items_subtotal = 0.0
                    for item in items:
                        if isinstance(item, dict) and not (item.get("is_lottery") or item.get("is_auto_gift")):
                            try:
                                items_subtotal += float(item.get("subtotal", 0) or 0)
                            except Exception:
                                pass
                    
                    # 获取适用的门槛配置并发放优惠券
                    applicable_thresholds = GiftThresholdDB.get_applicable_thresholds(items_subtotal, order_owner_id)
                    for threshold in applicable_thresholds:
                        gift_coupon = threshold.get('gift_coupon', 0) == 1
                        coupon_amount = threshold.get('coupon_amount', 0)
                        applicable_times = threshold.get('applicable_times', 0)
                        threshold_amount = threshold.get('threshold_amount', 0)
                        
                        if gift_coupon and coupon_amount > 0 and applicable_times > 0:
                            for _ in range(applicable_times):
                                coupon_ids = CouponDB.issue_coupons(
                                    user_identifier=order["student_id"],  # student_id会自动解析为user_id
                                    amount=coupon_amount,
                                    quantity=1,
                                    expires_at=None,
                                    owner_id=order_owner_id
                                )
                                if coupon_ids:
                                    logger.info(f"支付成功后为用户 {order['student_id']} 发放满额优惠券 {coupon_amount} 元（门槛{threshold_amount}）")
                except Exception as e:
                    logger.warning(f"发放满额优惠券失败: {e}")
                
                # 若本单使用了优惠券，支付成功后删除该券（已使用不再显示）
                try:
                    c_id = order.get("coupon_id")
                    d_amt = float(order.get("discount_amount") or 0)
                    if c_id and d_amt > 0:
                        CouponDB.delete_coupon(c_id)
                except Exception as e:
                    logger.warning(f"删除已用优惠券失败: {e}")
            except Exception as e:
                logger.warning(f"清空购物车失败: {e}")
            return success_response("已标记为已支付", {"order_id": order_id, "payment_status": "succeeded"})

        # 失败、待验证或回退为未付款：恢复库存并更新支付状态
        # 先检查当前状态，如果从succeeded回退，需要恢复库存
        current_status = order.get("payment_status")
        if current_status == "succeeded" and new_status in ["pending", "processing", "failed"]:
            # 从成功状态回退，需要恢复库存
            try:
                restore_ok = OrderDB.restore_stock_from_order(order_id)
                if not restore_ok:
                    logger.warning(f"恢复库存失败，但继续处理状态更新: order_id={order_id}")
            except Exception as e:
                logger.warning(f"恢复库存异常: {e}")
        
        ok = OrderDB.update_payment_status(order_id, new_status)
        if not ok:
            return error_response("更新支付状态失败", 500)
        # 若标记为未付款或失败，返还被锁定的优惠券
        try:
            if new_status in ["pending", "failed"]:
                c_id = order.get("coupon_id")
                d_amt = float(order.get("discount_amount") or 0)
                if c_id and d_amt > 0:
                    CouponDB.unlock_for_order(c_id, order_id)
        except Exception as e:
            logger.warning(f"返还优惠券失败: {e}")
        return success_response("支付状态已更新", {"order_id": order_id, "payment_status": new_status})
    except Exception as e:
        logger.error(f"管理员更新支付状态失败: {e}")
        return error_response("更新支付状态失败", 500)

@app.get("/healthz")
async def health_check():
    """健康检查"""
    return success_response("服务运行正常")

# ============== 用户资料（收货信息缓存） ==============

@app.get("/profile/shipping")
async def get_profile_shipping(request: Request):
    user = get_current_user_required_from_cookie(request)
    try:
        prof = UserProfileDB.get_shipping(user["id"])
        return success_response("获取收货资料成功", {"shipping": prof})
    except Exception as e:
        logger.error(f"获取收货资料失败: {e}")
        return error_response("获取收货资料失败", 500)


@app.post("/profile/location")
async def update_profile_location(payload: LocationUpdateRequest, request: Request):
    user = get_current_user_required_from_cookie(request)
    try:
        address = AddressDB.get_by_id(payload.address_id)
        if not address or int(address.get('enabled', 1) or 1) != 1:
            return error_response("地址不存在或未启用", 400)

        building = BuildingDB.get_by_id(payload.building_id)
        if (not building or building.get('address_id') != address.get('id')
                or int(building.get('enabled', 1) or 1) != 1):
            return error_response("楼栋不存在或未启用", 400)

        assignment = AgentAssignmentDB.get_agent_for_building(building.get('id'))
        agent_id = assignment.get('agent_id') if assignment else None

        existing = UserProfileDB.get_shipping(user['id']) or {}
        dormitory_name = address.get('name') or ''
        building_name = building.get('name') or ''
        room = existing.get('room') or ''
        full_address = f"{dormitory_name} {building_name} {room}".strip()

        updated_profile = {
            'name': existing.get('name') or '',
            'phone': existing.get('phone') or '',
            'room': room,
            'dormitory': dormitory_name,
            'building': building_name,
            'full_address': full_address,
            'address_id': address.get('id'),
            'building_id': building.get('id'),
            'agent_id': agent_id,
        }

        UserProfileDB.upsert_shipping(user['id'], updated_profile)
        try:
            CartDB.update_cart(user['id'], {})
        except Exception as e:
            logger.warning(f"切换地址时清空购物车失败: {e}")

        return success_response("配送地址已更新", {"shipping": updated_profile})
    except Exception as e:
        logger.error(f"更新配送地址失败: {e}")
        return error_response("更新配送地址失败", 500)

# 在所有API路由之后添加静态文件路由，避免冲突
@app.get("/logo.{extension}")
async def serve_logo(extension: str):
    """Serve logo files"""
    filename = f"logo.{extension}"
    file_path = os.path.join(public_dir, filename)
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    media_type = mimetypes.guess_type(file_path)[0]
    return FileResponse(
        file_path, 
        media_type=media_type,
        headers={
            "Cache-Control": f"public, max-age={STATIC_CACHE_MAX_AGE}, immutable",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
    )

@app.get("/payment_qr_{payment_id}.{extension}")
async def serve_payment_qr(payment_id: str, extension: str):
    """Serve payment QR code files"""
    filename = f"payment_qr_{payment_id}.{extension}"
    file_path = os.path.join(public_dir, filename)
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    media_type = mimetypes.guess_type(file_path)[0]
    return FileResponse(
        file_path, 
        media_type=media_type,
        headers={
            "Cache-Control": f"public, max-age={STATIC_CACHE_MAX_AGE}, immutable",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
    )

@app.get("/{filename}.txt")
async def serve_txt_files(filename: str):
    """Serve any .txt files from public directory"""
    full_filename = f"{filename}.txt"
    file_path = os.path.join(public_dir, full_filename)
    
    # Security check: prevent directory traversal
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path, 
        media_type="text/plain",
        headers={
            "Cache-Control": f"public, max-age={STATIC_CACHE_MAX_AGE}, immutable",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
    )

if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host=settings.backend_host,
        port=settings.backend_port,
        reload=settings.is_development,
        log_level=settings.log_level.lower()
    )
